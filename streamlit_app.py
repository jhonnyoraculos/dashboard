from __future__ import annotations

import base64
import hashlib
import html
import os
import re
import subprocess
import unicodedata
from io import BytesIO
from datetime import date, datetime
from pathlib import Path
from zoneinfo import ZoneInfo

os.environ.setdefault("JR_SKIP_WARM_CACHE", "1")

import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from PIL import Image, ImageDraw, ImageFont

import app as backend


JR_BLUE = "#1C2D6B"
JR_RED = "#BE1E2D"
MUTED = "#6B7280"
CARD_BORDER = "#c2d2f3"
LOGO_PATH = Path(__file__).parent / "static" / "logo-jr.png"
CURRENT_YEAR = date.today().year
APP_VERSION = "deploy-ranking-defaults-monthly-v1"
BR_TZ = ZoneInfo("America/Sao_Paulo")

PLOTLY_CONFIG = {
    "responsive": True,
    "displaylogo": False,
    "toImageButtonOptions": {"format": "png", "scale": 2},
    "modeBarButtonsToRemove": ["lasso2d", "select2d"],
}

MONTH_NAMES = [
    "Janeiro",
    "Fevereiro",
    "Março",
    "Abril",
    "Maio",
    "Junho",
    "Julho",
    "Agosto",
    "Setembro",
    "Outubro",
    "Novembro",
    "Dezembro",
]
MONTH_ABBR = ["jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out", "nov", "dez"]
YEAR_SERIES_COLORS = [JR_BLUE, JR_RED, "#0F766E", "#D97706", "#7C3AED", "#4B5563", "#9CA3AF"]

ROUTES = {
    "combustivel": backend.data_comb,
    "manutencao": backend.data_manu,
    "hoteis": backend.data_hoteis,
    "pedagio": backend.data_pedagio,
    "vex": backend.data_vex,
    "frota": backend.data_frota,
    "overview": backend.data_overview,
}

DASHBOARD_META = {
    "combustivel": {"label": "Combustível", "color": JR_BLUE, "supports_plate": True},
    "manutencao": {"label": "Manutenção", "color": JR_RED, "supports_plate": True},
    "hoteis": {"label": "Hotéis", "color": "#0F766E", "supports_plate": False},
    "pedagio": {"label": "Pedágio/IPVA", "color": "#D97706", "supports_plate": True},
    "vex": {"label": "Vex", "color": "#7C3AED", "supports_plate": True},
    "frota": {"label": "Ranking da frota", "color": JR_BLUE, "supports_plate": True},
}
COMPARE_ALLOWED_ROUTES = {"combustivel", "manutencao", "pedagio"}

COMPARE_SERIES = {
    "combustivel": {
        "monthly": ("custo_mensal", "Mes", "Custo"),
        "weekly": ("gasto_semana", "Dia", "Custo"),
        "plate": ("gasto_por_placa", "PLACA", "Custo"),
    },
    "manutencao": {
        "monthly": ("custo_mensal", "Mes", "Custo"),
        "weekly": ("custo_semana", "Dia", "Custo"),
        "plate": ("gasto_por_placa", "PLACA", "Custo"),
    },
    "hoteis": {
        "monthly": ("valor_mensal", "Mes", "Valor"),
        "weekly": ("valor_semana", "Dia", "Valor"),
        "plate": None,
    },
    "pedagio": {
        "monthly": ("custo_mensal", "Mes", "Custo"),
        "weekly": ("custo_semana", "Dia", "Custo"),
        "plate": ("gasto_por_placa", "PLACA", "Custo"),
    },
    "vex": {
        "monthly": ("mensal_total", "Mes", "Valor"),
        "weekly": None,
        "plate": ("gasto_por_placa", "PLACA", "Valor"),
    },
}

RANK_ORDER_OPTIONS = {
    "total": "Gasto total",
    "combustivel": "Combustível",
    "manutencao": "Manutenção",
    "pedagio": "Pedágio/IPVA",
    "peso": "Peso",
}


def logo_data_uri() -> str:
    if not LOGO_PATH.exists():
        return ""
    encoded = base64.b64encode(LOGO_PATH.read_bytes()).decode("ascii")
    return f"data:image/png;base64,{encoded}"


def inject_css() -> None:
    st.markdown(
        f"""
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800&display=swap');

        :root {{
          --jr-blue: {JR_BLUE};
          --jr-red: {JR_RED};
          --muted: {MUTED};
          --card-border: {CARD_BORDER};
          --radius: 14px;
        }}

        html, body, [class*="css"] {{
          font-family: "Inter", sans-serif;
        }}

        html,
        body,
        .stApp,
        [data-testid="stAppViewContainer"],
        [data-testid="stMain"],
        [data-testid="stMainBlockContainer"] {{
          max-width: 100%;
          overflow-x: clip;
        }}

        .stApp {{
          background: linear-gradient(180deg, #fbfdff 0%, #f5f7fc 100%);
          color: var(--jr-blue);
          position: relative;
          overflow-x: hidden;
        }}

        .stApp::before,
        .stApp::after {{
          content: "";
          position: fixed;
          inset: -18vh -18vw;
          pointer-events: none;
          z-index: 0;
          background-repeat: repeat;
          will-change: transform;
        }}

        .stApp::before {{
          opacity: .72;
          filter: drop-shadow(0 0 5px rgba(28,45,107,.28));
          background-image:
            radial-gradient(circle at 7% 14%, rgba(28,45,107,.32) 0 1px, transparent 2.1px),
            radial-gradient(circle at 18% 78%, rgba(28,45,107,.20) 0 1.4px, transparent 2.5px),
            radial-gradient(circle at 31% 43%, rgba(255,255,255,.78) 0 1px, transparent 2px),
            radial-gradient(circle at 42% 19%, rgba(28,45,107,.26) 0 1.2px, transparent 2.4px),
            radial-gradient(circle at 55% 68%, rgba(28,45,107,.18) 0 1.5px, transparent 2.7px),
            radial-gradient(circle at 69% 32%, rgba(255,255,255,.68) 0 1px, transparent 2.2px),
            radial-gradient(circle at 76% 91%, rgba(28,45,107,.25) 0 1px, transparent 2.2px),
            radial-gradient(circle at 88% 53%, rgba(28,45,107,.18) 0 1.6px, transparent 2.8px),
            radial-gradient(circle at 96% 23%, rgba(28,45,107,.30) 0 1.1px, transparent 2.3px);
          background-size: 620px 540px;
          animation: jr-particles-drift 52s linear infinite;
        }}

        .stApp::after {{
          opacity: .64;
          filter: drop-shadow(0 0 7px rgba(190,30,45,.30));
          background-image:
            radial-gradient(circle at 11% 37%, rgba(190,30,45,.24) 0 1.2px, transparent 2.4px),
            radial-gradient(circle at 21% 9%, rgba(190,30,45,.15) 0 1px, transparent 2.2px),
            radial-gradient(circle at 34% 84%, rgba(190,30,45,.20) 0 1.4px, transparent 2.7px),
            radial-gradient(circle at 48% 28%, rgba(255,255,255,.72) 0 1px, transparent 2.1px),
            radial-gradient(circle at 58% 61%, rgba(190,30,45,.18) 0 1.1px, transparent 2.3px),
            radial-gradient(circle at 73% 16%, rgba(190,30,45,.24) 0 1.5px, transparent 2.8px),
            radial-gradient(circle at 81% 74%, rgba(255,255,255,.62) 0 1px, transparent 2.1px),
            radial-gradient(circle at 93% 47%, rgba(190,30,45,.17) 0 1.3px, transparent 2.5px);
          background-size: 760px 610px;
          animation: jr-particles-float 71s linear infinite;
        }}

        @keyframes jr-particles-drift {{
          from {{ transform: translate3d(-5vw, -4vh, 0) rotate(.001deg); }}
          to {{ transform: translate3d(16vw, 13vh, 0) rotate(.001deg); }}
        }}

        @keyframes jr-particles-float {{
          from {{ transform: translate3d(14vw, 11vh, 0) rotate(.001deg); }}
          to {{ transform: translate3d(-18vw, -14vh, 0) rotate(.001deg); }}
        }}

        @media (prefers-reduced-motion: reduce) {{
          .stApp::before,
          .stApp::after {{
            animation: none;
          }}
        }}

        [data-testid="stAppViewContainer"],
        [data-testid="stMain"],
        [data-testid="stMainBlockContainer"],
        .block-container {{
          background: transparent;
          position: relative;
          z-index: 1;
        }}

        header[data-testid="stHeader"],
        #MainMenu,
        footer {{
          visibility: hidden;
          height: 0;
        }}

        .block-container {{
          padding: 0 24px 40px;
          max-width: 1400px;
        }}

        .jr-topbar {{
          display: flex;
          align-items: center;
          justify-content: space-between;
          gap: 16px;
          flex-wrap: wrap;
          padding: 18px 32px;
          margin: 0 0 0;
          width: auto;
          box-sizing: border-box;
          background: var(--jr-blue);
          box-shadow:
            0 4px 12px rgba(0,0,0,.2),
            0 0 0 100vmax var(--jr-blue);
          clip-path: inset(0 -100vmax);
          position: sticky;
          top: 0;
          z-index: 10;
        }}

        .jr-brand {{
          display: flex;
          align-items: center;
          gap: 14px;
        }}

        .jr-logo {{
          width: 44px;
          height: 44px;
          object-fit: contain;
        }}

        .jr-topbar h1 {{
          margin: 0;
          color: #fff;
          font-size: 22px;
          line-height: 1.2;
          font-weight: 800;
          letter-spacing: .2px;
        }}

        .jr-back {{
          display: inline-flex;
          align-items: center;
          justify-content: center;
          min-height: 38px;
          padding: 8px 16px;
          border-radius: 8px;
          background: var(--jr-red);
          color: #fff !important;
          font-size: 14px;
          font-weight: 700;
          text-decoration: none !important;
          box-shadow: 0 3px 8px rgba(0,0,0,.12);
        }}

        .st-key-comb_filterbar,
        .st-key-manu_filterbar,
        .st-key-hotel_filterbar,
        .st-key-ped_filterbar,
        .st-key-vex_filterbar,
        .st-key-rank_filterbar {{
          background: var(--jr-blue);
          margin: 0 0 58px;
          width: auto;
          padding: 0 32px 18px;
          box-sizing: border-box;
          box-shadow:
            0 4px 12px rgba(0,0,0,.2),
            0 0 0 100vmax var(--jr-blue);
          clip-path: inset(0 -100vmax);
          position: sticky;
          top: 80px;
          z-index: 9;
          isolation: isolate;
          overflow: visible;
        }}

        .st-key-comb_filterbar > div,
        .st-key-manu_filterbar > div,
        .st-key-hotel_filterbar > div,
        .st-key-ped_filterbar > div,
        .st-key-vex_filterbar > div,
        .st-key-rank_filterbar > div {{
          position: relative;
          z-index: 1;
        }}

        .st-key-comb_filterbar label,
        .st-key-manu_filterbar label,
        .st-key-hotel_filterbar label,
        .st-key-ped_filterbar label,
        .st-key-vex_filterbar label,
        .st-key-rank_filterbar label {{
          display: none !important;
        }}

        .st-key-comb_filterbar div[data-baseweb="select"] > div,
        .st-key-manu_filterbar div[data-baseweb="select"] > div,
        .st-key-hotel_filterbar div[data-baseweb="select"] > div,
        .st-key-ped_filterbar div[data-baseweb="select"] > div,
        .st-key-vex_filterbar div[data-baseweb="select"] > div,
        .st-key-rank_filterbar div[data-baseweb="select"] > div {{
          min-height: 38px;
          border-radius: 8px;
          border: 1px solid rgba(255,255,255,.25);
          background: #fff;
          box-shadow: 0 3px 8px rgba(0,0,0,.12);
        }}

        .st-key-comb_filterbar [data-testid="stHorizontalBlock"],
        .st-key-manu_filterbar [data-testid="stHorizontalBlock"],
        .st-key-hotel_filterbar [data-testid="stHorizontalBlock"],
        .st-key-ped_filterbar [data-testid="stHorizontalBlock"],
        .st-key-vex_filterbar [data-testid="stHorizontalBlock"],
        .st-key-rank_filterbar [data-testid="stHorizontalBlock"] {{
          gap: 12px;
          align-items: stretch;
          flex-wrap: wrap !important;
        }}

        .st-key-comb_filterbar [data-testid="column"],
        .st-key-manu_filterbar [data-testid="column"],
        .st-key-hotel_filterbar [data-testid="column"],
        .st-key-ped_filterbar [data-testid="column"],
        .st-key-vex_filterbar [data-testid="column"],
        .st-key-rank_filterbar [data-testid="column"] {{
          flex: 1 1 150px !important;
          min-width: 145px !important;
          max-width: none !important;
        }}

        .filter-back {{
          display: inline-flex;
          width: 100%;
          min-height: 38px;
          align-items: center;
          justify-content: center;
          border-radius: 8px;
          background: var(--jr-red);
          color: #fff !important;
          font-size: 14px;
          font-weight: 800;
          text-decoration: none !important;
          box-shadow: 0 3px 8px rgba(190,30,45,.25);
          white-space: nowrap;
        }}

        .home-wrapper {{
          max-width: 1100px;
          margin: 0 auto;
          padding: 0 0 48px;
          display: flex;
          flex-direction: column;
          gap: 48px;
        }}

        .home-header {{
          max-width: 1100px;
          margin: 0 auto 48px;
          background:
            linear-gradient(145deg, rgba(255,255,255,.78), rgba(255,255,255,.52)),
            rgba(255,255,255,.58);
          border: 1px solid rgba(255,255,255,.78);
          border-radius: 18px;
          padding: 40px;
          box-shadow:
            0 22px 54px rgba(16,24,40,.10),
            inset 0 1px 0 rgba(255,255,255,.78);
          backdrop-filter: blur(22px) saturate(145%);
          -webkit-backdrop-filter: blur(22px) saturate(145%);
          display: flex;
          flex-direction: column;
          gap: 28px;
          position: relative;
          overflow: hidden;
          isolation: isolate;
        }}

        .home-header::before {{
          content: "";
          position: absolute;
          left: -140px;
          bottom: -130px;
          width: 520px;
          height: 330px;
          background:
            radial-gradient(ellipse at 28% 72%, rgba(190,30,45,.24), rgba(190,30,45,.12) 34%, transparent 68%),
            radial-gradient(ellipse at 58% 52%, rgba(190,30,45,.13), transparent 62%);
          filter: blur(3px);
          z-index: 0;
        }}

        .home-header::after {{
          content: "";
          position: absolute;
          right: -60px;
          top: -60px;
          width: 220px;
          height: 220px;
          background-image: radial-gradient(circle, rgba(190,30,45,.18) 1.2px, transparent 1.8px);
          background-size: 18px 18px;
          opacity: .52;
          transform: rotate(12deg);
          filter: drop-shadow(0 0 6px rgba(190,30,45,.22));
          z-index: 0;
        }}

        .home-brand {{
          display: flex;
          align-items: flex-start;
          gap: 24px;
          position: relative;
          z-index: 1;
        }}

        .home-logo {{
          width: 72px;
          height: 72px;
          border-radius: 16px;
          box-shadow: 0 8px 22px rgba(0,0,0,0.12);
          object-fit: contain;
        }}

        .home-eyebrow {{
          margin: 0;
          font-size: 13px;
          text-transform: uppercase;
          letter-spacing: .18em;
          color: var(--muted);
          font-weight: 600;
        }}

        .home-header h1 {{
          margin: 4px 0 12px;
          font-size: 34px;
          font-weight: 800;
          color: var(--jr-blue);
        }}

        .home-subtitle {{
          margin: 0;
          font-size: 16px;
          color: #4B5563;
          max-width: 560px;
        }}

        .home-cta,
        .home-outline {{
          align-self: flex-start;
          border-radius: 999px;
          text-decoration: none !important;
          font-weight: 700;
          position: relative;
          z-index: 1;
        }}

        .home-cta {{
          background: var(--jr-red);
          color: #fff !important;
          padding: 12px 24px;
          box-shadow: 0 10px 24px rgba(190,30,45,0.25);
        }}

        .home-header-actions {{
          position: absolute;
          top: 24px;
          right: 24px;
          display: flex;
          flex-direction: column;
          align-items: flex-end;
          gap: 8px;
          z-index: 2;
        }}

        .home-last-update {{
          margin: 0;
          padding: 4px 9px;
          border-radius: 999px;
          background: rgba(255,255,255,.70);
          border: 1px solid rgba(28,45,107,.10);
          color: rgba(28,45,107,.62);
          font-size: 10.5px;
          line-height: 1.25;
          font-weight: 700;
          white-space: nowrap;
          backdrop-filter: blur(10px);
        }}

        .home-admin-link {{
          min-height: 38px;
          display: inline-flex;
          align-items: center;
          justify-content: center;
          padding: 0 16px;
          border-radius: 999px;
          border: 1.5px solid rgba(28,45,107,.16);
          background: rgba(255,255,255,.86);
          color: var(--jr-blue) !important;
          font-size: 13px;
          font-weight: 800;
          text-decoration: none !important;
          box-shadow: 0 10px 24px rgba(16,24,40,.10);
          backdrop-filter: blur(10px);
        }}

        .st-key-cadastro_shell {{
          max-width: 1120px;
          margin: 28px auto 48px;
          padding: 28px;
          border: 1px solid rgba(255,255,255,.74);
          border-radius: var(--radius);
          background: rgba(255,255,255,.76);
          box-shadow: 0 8px 22px rgba(16,24,40,.10);
          backdrop-filter: blur(14px);
        }}

        .st-key-cadastro_shell [data-testid="stForm"] {{
          border: 1.5px solid var(--card-border);
          border-radius: var(--radius);
          background: rgba(255,255,255,.82);
          padding: 22px;
          box-shadow: 0 12px 28px rgba(16,24,40,.10);
        }}

        .st-key-cadastro_shell .stButton > button,
        .st-key-cadastro_shell [data-testid="stFormSubmitButton"] button {{
          min-height: 42px;
          border-radius: 999px;
          font-weight: 800;
        }}

        .home-total-section {{
          background: rgba(244,247,253,0.68);
          border: 1px solid rgba(255,255,255,0.7);
          border-radius: var(--radius);
          padding: 32px;
          box-shadow: 0 6px 18px rgba(0,0,0,0.08);
          backdrop-filter: blur(12px);
        }}

        .st-key-home_total_section {{
          max-width: 1100px;
          margin: 0 auto 28px;
          background: rgba(244,247,253,0.68);
          border: 1px solid rgba(255,255,255,0.7);
          border-radius: var(--radius);
          padding: 32px;
          box-shadow: 0 6px 18px rgba(0,0,0,0.08);
          backdrop-filter: blur(12px);
        }}

        .st-key-home_total_section div[data-baseweb="select"] > div {{
          min-height: 34px;
          border-radius: 14px;
          border: 1px solid rgba(28,45,107,0.15);
          background: #fff;
          box-shadow: 0 4px 10px rgba(0,0,0,0.05);
        }}

        .st-key-home_total_section .stButton > button {{
          min-height: 34px;
          border-radius: 999px;
          box-shadow: 0 8px 20px rgba(190,30,45,0.2);
        }}

        .home-export-bar {{
          display: grid;
          grid-template-columns: repeat(2, minmax(0, 1fr));
          gap: 10px;
          margin: 6px 0 12px;
        }}

        .home-export-btn {{
          min-height: 40px;
          display: inline-flex;
          align-items: center;
          justify-content: center;
          border: 1.5px solid var(--card-border);
          background: #f8f9ff;
          color: var(--jr-blue);
          font-weight: 800;
          border-radius: 12px;
          box-shadow: 0 8px 16px rgba(16,24,40,0.08);
        }}

        .home-total-grid,
        .kpis {{
          display: grid;
          grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
          gap: 20px;
        }}

        .home-total-card,
        .kpi {{
          background:
            linear-gradient(150deg, rgba(255,255,255,.92), rgba(255,255,255,.66)),
            rgba(255,255,255,.70);
          border: 1px solid rgba(194,210,243,.72);
          border-radius: 16px;
          padding: 18px 24px;
          display: flex;
          flex-direction: column;
          gap: 4px;
          box-shadow:
            0 18px 44px rgba(16,24,40,.10),
            inset 0 1px 0 rgba(255,255,255,.82);
          backdrop-filter: blur(18px) saturate(140%);
          -webkit-backdrop-filter: blur(18px) saturate(140%);
        }}

        .home-total-card {{
          border: 0;
          box-shadow: 0 14px 32px rgba(16,24,40,0.12);
        }}

        .kpis {{
          margin: 18px 0 28px;
          gap: 28px;
        }}

        .kpi-sections {{
          display: grid;
          grid-template-columns: repeat(2, minmax(0, 1fr));
          gap: 24px;
          margin: 18px 0 28px;
        }}

        .kpi-section--summary {{
          grid-column: 1 / -1;
        }}

        .kpi-section {{
          border: 1px solid rgba(194,210,243,.9);
          border-radius: var(--radius);
          padding: 18px;
          background: rgba(255,255,255,.46);
          box-shadow: 0 12px 28px rgba(16,24,40,0.08);
        }}

        .kpi-section-title {{
          display: flex;
          align-items: center;
          gap: 10px;
          margin: 0 0 16px;
          color: var(--jr-blue);
          font-size: 13px;
          font-weight: 800;
          text-transform: uppercase;
          letter-spacing: .05em;
        }}

        .kpi-section-title::before {{
          content: "";
          width: 32px;
          height: 4px;
          border-radius: 999px;
          background: var(--section-accent, var(--jr-red));
        }}

        .kpi-section .kpis {{
          margin: 0;
          gap: 18px;
        }}

        .kpi {{
          padding: 24px;
          min-height: 116px;
          align-items: center;
          justify-content: center;
          text-align: center;
          position: relative;
          overflow: hidden;
        }}

        .kpi::before {{
          content: "";
          position: absolute;
          inset: 0 0 auto 0;
          height: 4px;
          background: var(--kpi-accent, var(--jr-red));
        }}

        .home-total-label,
        .kpi-title {{
          margin: 0;
          font-size: 12px;
          letter-spacing: .05em;
          text-transform: uppercase;
          color: var(--muted);
          font-weight: 700;
          width: 100%;
          text-align: center;
        }}

        .home-total-value {{
          margin: 0;
          font-size: 28px;
          font-weight: 800;
          color: var(--jr-blue);
          line-height: 1.2;
        }}

        .home-total-status {{
          margin: 0;
          font-size: 12px;
          color: #4B5563;
        }}

        .home-filter-row {{
          margin-top: 24px;
        }}

        .kpi-value {{
          font-size: clamp(18px, 1.8vw, 26px);
          font-weight: 800;
          color: var(--kpi-accent, var(--jr-red));
          line-height: 1.05;
          margin-top: 8px;
          overflow-wrap: anywhere;
          width: 100%;
          text-align: center;
        }}

        .home-grid {{
          max-width: 1100px;
          margin: 0 auto;
          display: grid;
          gap: 28px;
          grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
        }}

        .home-card {{
          background:
            linear-gradient(145deg, rgba(255,255,255,.76), rgba(255,255,255,.48)),
            rgba(255,255,255,.58);
          border: 1px solid rgba(255,255,255,.82);
          border-radius: 18px;
          padding: 32px;
          box-shadow:
            0 22px 48px rgba(16,24,40,.10),
            inset 0 1px 0 rgba(255,255,255,.86);
          backdrop-filter: blur(22px) saturate(145%);
          -webkit-backdrop-filter: blur(22px) saturate(145%);
          display: flex;
          flex-direction: column;
          gap: 20px;
          position: relative;
          overflow: hidden;
          min-height: 330px;
          color: inherit !important;
          text-decoration: none !important;
          cursor: pointer;
          transition:
            transform .18s ease,
            box-shadow .18s ease,
            border-color .18s ease,
            background .18s ease;
        }}

        .home-card:hover {{
          transform: translateY(-4px);
          border-color: rgba(194,210,243,.96);
          box-shadow:
            0 28px 64px rgba(16,24,40,.14),
            inset 0 1px 0 rgba(255,255,255,.90);
        }}

        .home-card:focus-visible {{
          outline: 3px solid rgba(190,30,45,.26);
          outline-offset: 3px;
        }}

        .home-card::after {{
          content: "";
          position: absolute;
          inset: 0;
          background:
            radial-gradient(circle at 88% 8%, rgba(255,255,255,.72), transparent 18%),
            linear-gradient(140deg, rgba(28,45,107,0.06), transparent 65%);
          pointer-events: none;
        }}

        .home-card > * {{
          position: relative;
          z-index: 1;
        }}

        .home-chip {{
          display: inline-flex;
          align-self: flex-start;
          background: rgba(28,45,107,0.08);
          border: 1px solid rgba(255,255,255,.72);
          color: var(--jr-blue);
          padding: 6px 14px;
          border-radius: 999px;
          font-size: 13px;
          font-weight: 700;
          letter-spacing: .04em;
          margin-bottom: 12px;
        }}

        .home-card h2 {{
          margin: 0;
          font-size: 22px;
          color: var(--jr-blue);
        }}

        .home-card-text,
        .home-list,
        .home-footer p {{
          color: #4B5563;
          font-size: 15px;
        }}

        .home-list {{
          margin: 0;
          padding-left: 20px;
          display: flex;
          flex-direction: column;
          gap: 6px;
        }}

        .home-link {{
          margin-top: auto;
          align-self: flex-start;
          font-weight: 800;
          color: var(--jr-blue) !important;
          text-decoration: none !important;
          transition: transform .18s ease, color .18s ease;
        }}

        .home-card:hover .home-link {{
          transform: translateX(3px);
          color: var(--jr-red) !important;
        }}

        .home-footer {{
          max-width: 1100px;
          margin: 48px auto 0;
          background:
            linear-gradient(145deg, rgba(255,255,255,.76), rgba(255,255,255,.48)),
            rgba(255,255,255,.58);
          border: 1px solid rgba(255,255,255,.82);
          border-radius: 18px;
          padding: 36px;
          box-shadow:
            0 22px 48px rgba(16,24,40,.10),
            inset 0 1px 0 rgba(255,255,255,.86);
          backdrop-filter: blur(22px) saturate(145%);
          -webkit-backdrop-filter: blur(22px) saturate(145%);
        }}

        .home-footer h3 {{
          margin: 0 0 16px;
          font-size: 20px;
          color: var(--jr-blue);
        }}

        div[data-testid="stVerticalBlockBorderWrapper"] {{
          background:
            linear-gradient(150deg, rgba(255,255,255,.92), rgba(255,255,255,.68)),
            rgba(255,255,255,.70);
          border: 1px solid rgba(194,210,243,.76);
          border-radius: 16px;
          box-shadow:
            0 18px 44px rgba(16,24,40,.10),
            inset 0 1px 0 rgba(255,255,255,.84);
          backdrop-filter: blur(18px) saturate(140%);
          -webkit-backdrop-filter: blur(18px) saturate(140%);
        }}

        .chart-title {{
          margin: 0 0 8px;
          font-size: 12px;
          font-weight: 800;
          color: var(--jr-blue);
          position: relative;
          padding-bottom: 10px;
        }}

        .chart-title::after {{
          content: "";
          position: absolute;
          left: 0;
          bottom: 0;
          width: 32px;
          height: 3px;
          background: var(--jr-red);
          border-radius: 2px;
        }}

        .footer-note {{
          color: var(--muted);
          font-size: 13px;
          text-align: center;
          padding: 24px;
          margin-top: 40px;
          border-top: 1px solid #e5e7eb;
        }}

        div[data-testid="stSelectbox"] label,
        div[data-testid="stMultiSelect"] label {{
          color: var(--jr-blue);
          font-weight: 800;
          font-size: 12px;
          text-transform: uppercase;
          letter-spacing: .04em;
        }}

        .stButton > button {{
          background: var(--jr-red);
          color: #fff;
          border: none;
          border-radius: 10px;
          font-weight: 800;
          box-shadow: 0 3px 8px rgba(190,30,45,.25);
          white-space: nowrap;
        }}

        .stButton > button:hover {{
          color: #fff;
          border: none;
          background: #a51924;
        }}

        [class*="_export_"] button,
        [class*="_toggle_"] button {{
          background: #f8f9ff !important;
          color: var(--jr-blue) !important;
          border: 1.5px solid var(--card-border) !important;
          border-radius: 12px !important;
          box-shadow: 0 8px 16px rgba(16,24,40,0.08) !important;
        }}

        [class*="_download_"] button {{
          background: var(--jr-blue) !important;
          color: #fff !important;
          border: none !important;
          border-radius: 12px !important;
        }}

        [class*="_dashboard_controls"] {{
          margin: 14px 0 22px;
          padding: 0;
          border: 1px solid rgba(194,210,243,.85);
          border-radius: 14px;
          background: rgba(255,255,255,.70);
          box-shadow: 0 14px 34px rgba(16,24,40,.10);
          backdrop-filter: blur(12px);
          overflow: hidden;
        }}

        [class*="_dashboard_controls"] details {{
          border: 0 !important;
          background: transparent !important;
        }}

        [class*="_dashboard_controls"] summary {{
          min-height: 46px;
          padding: 0 16px !important;
          color: var(--jr-blue) !important;
          font-weight: 800 !important;
          border-bottom: 1px solid rgba(194,210,243,.72);
        }}

        [class*="_dashboard_controls"] details > div {{
          padding: 14px 16px 16px;
        }}

        .control-row-label {{
          margin: 2px 0 8px;
          color: var(--muted);
          font-size: 11px;
          font-weight: 800;
          letter-spacing: .06em;
          text-transform: uppercase;
        }}

        .control-row-spacer {{
          height: 12px;
        }}

        [class*="_dashboard_controls"] [data-testid="stHorizontalBlock"] {{
          gap: 10px;
          flex-wrap: wrap !important;
          align-items: stretch;
        }}

        [class*="_dashboard_controls"] [data-testid="column"] {{
          display: flex;
          flex-direction: column;
          justify-content: stretch;
        }}

        [class*="_dashboard_controls"] .stButton,
        [class*="_dashboard_controls"] [data-testid="stDownloadButton"],
        [class*="_dashboard_controls"] [data-testid="stCheckbox"] {{
          width: 100%;
        }}

        [class*="_dashboard_controls"] .stButton > button,
        [class*="_dashboard_controls"] [data-testid="stDownloadButton"] button {{
          min-height: 40px;
        }}

        [class*="_dashboard_controls"] [data-testid="stCheckbox"] label {{
          width: 100%;
          min-height: 44px;
          display: inline-flex;
          align-items: center;
          justify-content: flex-start;
          gap: 8px;
          padding: 8px 12px;
          border: 1.5px solid var(--card-border);
          border-radius: 12px;
          background: rgba(248,249,255,.92);
          color: var(--jr-blue);
          box-shadow: 0 6px 12px rgba(16,24,40,0.08);
          font-weight: 800;
          box-sizing: border-box;
        }}

        [class*="_dashboard_controls"] [data-testid="stCheckbox"] p {{
          font-size: 12px;
          line-height: 1.2;
          font-weight: 800;
          color: var(--jr-blue);
        }}

        .ranking-header {{
          display: grid;
          grid-template-columns: .45fr .9fr repeat(4, 1.05fr) .85fr .75fr .8fr;
          gap: 12px;
          align-items: center;
          padding: 12px 16px;
          margin: 8px 0 10px;
          border: 1px solid rgba(194,210,243,.88);
          border-radius: 12px;
          background: rgba(255,255,255,.74);
          color: var(--muted);
          font-size: 11px;
          font-weight: 900;
          letter-spacing: .06em;
          text-transform: uppercase;
          box-shadow: 0 10px 24px rgba(16,24,40,.08);
        }}

        .st-key-frota_ranking_table [data-testid="stExpander"] {{
          border: 1px solid rgba(194,210,243,.95);
          border-radius: 12px;
          background: rgba(255,255,255,.84);
          box-shadow: 0 10px 24px rgba(16,24,40,.08);
          overflow: hidden;
        }}

        .st-key-frota_ranking_table [data-testid="stExpander"] details {{
          border: 0 !important;
          background: transparent !important;
        }}

        .st-key-frota_ranking_table [data-testid="stExpander"] summary {{
          min-height: 60px;
          padding: 0 18px !important;
          border-bottom: 1px solid rgba(194,210,243,.58);
        }}

        .st-key-frota_ranking_table [data-testid="stExpander"] summary p {{
          color: var(--jr-blue);
          font-weight: 900;
          font-size: 16px;
          line-height: 1.35;
        }}

        .ranking-detail-grid {{
          display: grid;
          grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
          gap: 12px;
          padding: 4px 0 8px;
        }}

        .ranking-detail-card {{
          min-height: 74px;
          padding: 12px;
          border: 1px solid rgba(194,210,243,.82);
          border-radius: 10px;
          background: linear-gradient(145deg, #fff, #f9fbff);
        }}

        .ranking-detail-label {{
          margin: 0 0 6px;
          color: var(--muted);
          font-size: 10px;
          font-weight: 900;
          letter-spacing: .06em;
          text-transform: uppercase;
        }}

        .ranking-detail-value {{
          margin: 0;
          color: var(--jr-red);
          font-size: 19px;
          line-height: 1.15;
          font-weight: 900;
        }}

        .ranking-row-summary,
        .ranking-versus-grid {{
          display: grid;
          grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
          gap: 12px;
          margin: 4px 0 14px;
        }}

        .ranking-row-summary {{
          grid-template-columns: repeat(auto-fit, minmax(130px, 1fr));
        }}

        .ranking-summary-item,
        .ranking-versus-card {{
          border: 1px solid rgba(194,210,243,.9);
          border-radius: 12px;
          background: #fff;
          box-shadow: 0 8px 18px rgba(16,24,40,.08);
          overflow: hidden;
        }}

        .ranking-summary-item {{
          padding: 12px;
        }}

        .ranking-summary-label {{
          margin: 0 0 5px;
          color: var(--muted);
          font-size: 10px;
          font-weight: 900;
          text-transform: uppercase;
          letter-spacing: .06em;
        }}

        .ranking-summary-value {{
          margin: 0;
          color: var(--jr-blue);
          font-size: 17px;
          font-weight: 900;
          line-height: 1.15;
        }}

        .ranking-versus-title {{
          margin: 18px 0 10px;
          color: var(--jr-blue);
          font-size: 18px;
          font-weight: 900;
        }}

        .ranking-versus-card {{
          padding: 16px;
          border-top: 4px solid var(--jr-blue);
        }}

        .ranking-versus-card h3 {{
          margin: 0 0 12px;
          color: var(--jr-blue);
          font-size: 18px;
          font-weight: 900;
        }}

        .ranking-versus-metric {{
          display: flex;
          justify-content: space-between;
          gap: 12px;
          padding: 8px 0;
          border-top: 1px solid rgba(229,231,235,.95);
          color: var(--jr-blue);
          font-size: 13px;
          font-weight: 800;
        }}

        .ranking-versus-metric span:first-child {{
          color: var(--muted);
          font-weight: 800;
        }}

        .ranking-versus-metric span:last-child {{
          color: var(--jr-red);
          text-align: right;
          font-weight: 900;
        }}

        .ranking-difference {{
          margin: 0 0 16px;
          padding: 16px;
          border: 1.5px solid rgba(194,210,243,.95);
          border-top: 4px solid var(--jr-red);
          border-radius: 12px;
          background: rgba(255,255,255,.92);
          box-shadow: 0 10px 24px rgba(16,24,40,.10);
        }}

        .ranking-difference h3 {{
          margin: 0 0 12px;
          color: var(--jr-blue);
          font-size: 17px;
          font-weight: 900;
        }}

        .ranking-difference-grid {{
          display: grid;
          grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
          gap: 10px;
        }}

        .ranking-difference-item {{
          padding: 12px;
          border-radius: 10px;
          background: #f8faff;
          border: 1px solid rgba(194,210,243,.72);
        }}

        .ranking-difference-label {{
          margin: 0 0 6px;
          color: var(--muted);
          font-size: 10px;
          font-weight: 900;
          letter-spacing: .06em;
          text-transform: uppercase;
        }}

        .ranking-difference-value {{
          margin: 0;
          color: var(--jr-red);
          font-size: 20px;
          line-height: 1.15;
          font-weight: 900;
        }}

        .ranking-difference-note {{
          margin: 4px 0 0;
          color: var(--jr-blue);
          font-size: 12px;
          font-weight: 800;
        }}

        .ranking-empty {{
          margin: 18px 0;
          padding: 18px;
          border: 1px solid rgba(194,210,243,.9);
          border-radius: 12px;
          background: rgba(255,255,255,.72);
          color: var(--muted);
          font-weight: 700;
        }}

        .dominance-panel {{
          margin-top: 14px;
          padding: 16px;
          border: 1px solid rgba(194,210,243,.80);
          border-radius: 14px;
          background:
            linear-gradient(150deg, rgba(255,255,255,.88), rgba(255,255,255,.62)),
            rgba(255,255,255,.70);
          box-shadow: inset 0 1px 0 rgba(255,255,255,.84);
        }}

        .dominance-title {{
          margin: 0 0 6px;
          color: var(--jr-blue);
          font-size: 15px;
          font-weight: 900;
        }}

        .dominance-note {{
          margin: 0 0 14px;
          color: var(--muted);
          font-size: 13px;
          font-weight: 700;
        }}

        .dominance-cities {{
          display: grid;
          gap: 8px;
          max-height: 360px;
          overflow-y: auto;
          padding-right: 4px;
        }}

        .dominance-city-row {{
          display: grid;
          grid-template-columns: minmax(130px, 1.25fr) minmax(82px, .75fr) repeat(3, minmax(84px, auto));
          gap: 10px;
          align-items: center;
          padding: 10px 12px;
          border: 1px solid rgba(194,210,243,.62);
          border-radius: 10px;
          background: rgba(255,255,255,.64);
        }}

        .dominance-city-row--head {{
          background: rgba(28,45,107,.06);
          border-color: rgba(194,210,243,.84);
        }}

        .dominance-city-name {{
          color: var(--jr-blue);
          font-weight: 900;
        }}

        .dominance-city-plate {{
          color: var(--jr-blue);
          font-size: 12px;
          font-weight: 900;
          white-space: nowrap;
        }}

        .dominance-city-metric {{
          color: var(--jr-red);
          font-size: 12px;
          font-weight: 900;
          text-align: right;
          white-space: nowrap;
        }}

        @media (max-width: 780px) {{
          .block-container {{
            padding: 0 10px 30px;
          }}

          .jr-topbar {{
            margin: 0;
            width: auto;
            padding: 10px 12px;
            position: static;
          }}

          .st-key-comb_filterbar,
          .st-key-manu_filterbar,
          .st-key-hotel_filterbar,
          .st-key-ped_filterbar,
          .st-key-vex_filterbar,
          .st-key-rank_filterbar {{
            margin: 0 -10px 18px;
            width: auto;
            padding: 8px 10px 10px;
            position: static;
            overflow: visible;
          }}

          .st-key-comb_filterbar [data-testid="stHorizontalBlock"],
          .st-key-manu_filterbar [data-testid="stHorizontalBlock"],
          .st-key-hotel_filterbar [data-testid="stHorizontalBlock"],
          .st-key-ped_filterbar [data-testid="stHorizontalBlock"],
          .st-key-vex_filterbar [data-testid="stHorizontalBlock"],
          .st-key-rank_filterbar [data-testid="stHorizontalBlock"] {{
            flex-wrap: nowrap !important;
            align-items: stretch;
            gap: 8px;
            overflow-x: auto;
            overflow-y: visible;
            padding: 2px 2px 8px;
            scroll-snap-type: x proximity;
            -webkit-overflow-scrolling: touch;
            scrollbar-width: thin;
          }}

          .st-key-comb_filterbar [data-testid="column"],
          .st-key-manu_filterbar [data-testid="column"],
          .st-key-hotel_filterbar [data-testid="column"],
          .st-key-ped_filterbar [data-testid="column"],
          .st-key-vex_filterbar [data-testid="column"],
          .st-key-rank_filterbar [data-testid="column"] {{
            flex: 0 0 clamp(132px, 43vw, 190px) !important;
            min-width: clamp(132px, 43vw, 190px) !important;
            max-width: clamp(132px, 43vw, 190px) !important;
            scroll-snap-align: start;
          }}

          .st-key-comb_filterbar div[data-baseweb="select"] > div,
          .st-key-manu_filterbar div[data-baseweb="select"] > div,
          .st-key-hotel_filterbar div[data-baseweb="select"] > div,
          .st-key-ped_filterbar div[data-baseweb="select"] > div,
          .st-key-vex_filterbar div[data-baseweb="select"] > div,
          .st-key-rank_filterbar div[data-baseweb="select"] > div {{
            min-height: 36px;
          }}

          .st-key-comb_filterbar span[data-baseweb="tag"],
          .st-key-manu_filterbar span[data-baseweb="tag"],
          .st-key-hotel_filterbar span[data-baseweb="tag"],
          .st-key-ped_filterbar span[data-baseweb="tag"],
          .st-key-vex_filterbar span[data-baseweb="tag"],
          .st-key-rank_filterbar span[data-baseweb="tag"] {{
            max-width: 74px;
            min-height: 22px;
          }}

          .st-key-comb_filterbar span[data-baseweb="tag"] span,
          .st-key-manu_filterbar span[data-baseweb="tag"] span,
          .st-key-hotel_filterbar span[data-baseweb="tag"] span,
          .st-key-ped_filterbar span[data-baseweb="tag"] span,
          .st-key-vex_filterbar span[data-baseweb="tag"] span,
          .st-key-rank_filterbar span[data-baseweb="tag"] span {{
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
            font-size: 11px;
          }}

          .jr-topbar h1 {{
            font-size: 17px;
          }}

          .home-wrapper {{
            padding: 20px 0 36px;
            gap: 24px;
          }}

          .home-header,
          .home-total-section,
          .home-card,
          .home-footer {{
            padding: 20px;
          }}

          .home-brand {{
            flex-direction: column;
          }}

          .home-header-actions {{
            position: relative;
            top: auto;
            right: auto;
            align-self: flex-start;
            align-items: flex-start;
          }}

          .home-header h1 {{
            font-size: 28px;
          }}

          .home-total-grid,
          .kpi-sections,
          .kpis,
          .home-grid,
          .home-export-bar {{
            grid-template-columns: 1fr;
          }}

          [class*="_dashboard_controls"] {{
            margin: 10px 0 18px;
          }}

          [class*="_dashboard_controls"] summary {{
            min-height: 42px;
            padding: 0 12px !important;
          }}

          [class*="_dashboard_controls"] details > div {{
            padding: 12px;
          }}

          [class*="_dashboard_controls"] [data-testid="column"] {{
            flex: 1 1 calc(50% - 8px) !important;
            min-width: min(180px, 100%) !important;
          }}

          [class*="_dashboard_controls"] [data-testid="stCheckbox"] label {{
            min-height: 40px;
            padding: 7px 10px;
          }}

          .ranking-header {{
            display: none;
          }}

          .st-key-frota_ranking_table [data-testid="stExpander"] summary p {{
            font-size: 12px;
          }}

          .dominance-city-row {{
            grid-template-columns: 1fr;
            gap: 4px;
          }}

          .dominance-city-metric {{
            text-align: left;
          }}
        }}

        @media (max-width: 560px) {{
          .st-key-comb_filterbar [data-testid="column"],
          .st-key-manu_filterbar [data-testid="column"],
          .st-key-hotel_filterbar [data-testid="column"],
          .st-key-ped_filterbar [data-testid="column"],
          .st-key-vex_filterbar [data-testid="column"],
          .st-key-rank_filterbar [data-testid="column"] {{
            flex-basis: 150px !important;
            min-width: 150px !important;
            max-width: 150px !important;
          }}

          .filter-back,
          .st-key-comb_filterbar .stButton > button,
          .st-key-manu_filterbar .stButton > button,
          .st-key-hotel_filterbar .stButton > button,
          .st-key-ped_filterbar .stButton > button,
          .st-key-vex_filterbar .stButton > button,
          .st-key-rank_filterbar .stButton > button {{
            min-height: 36px;
            padding-left: 8px;
            padding-right: 8px;
            font-size: 12px;
          }}

          [class*="_dashboard_controls"] [data-testid="column"] {{
            flex: 1 1 100% !important;
            min-width: 100% !important;
          }}
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )


def _freeze_route_params(params: dict[str, object] | None = None) -> tuple[tuple[str, object], ...]:
    params = params or {}
    clean_items: list[tuple[str, object]] = []
    for key, value in params.items():
        if value is None:
            continue
        if isinstance(value, str) and value in ("", "Todos"):
            continue
        if isinstance(value, (list, tuple, set)):
            values = tuple(item for item in value if item not in (None, "", "Todos"))
            if values:
                clean_items.append((key, values))
        else:
            clean_items.append((key, value))
    return tuple(sorted(clean_items))


@st.cache_data(ttl=30, show_spinner=False)
def _route_json_cached(route: str, frozen_params: tuple[tuple[str, object], ...], version: str) -> dict:
    func = ROUTES[route]
    params = {
        key: list(value) if isinstance(value, tuple) else value
        for key, value in frozen_params
    }
    return func(params) or {}


def clear_cached_reads() -> None:
    try:
        _route_json_cached.clear()
    except Exception:
        pass


def route_json(route: str, params: dict[str, object] | None = None) -> dict:
    return _route_json_cached(route, _freeze_route_params(params), APP_VERSION)


def route_json_uncached(route: str, params: dict[str, object] | None = None) -> dict:
    func = ROUTES[route]
    clean_params: dict[str, object] = {}
    for key, value in _freeze_route_params(params):
        clean_params[key] = list(value) if isinstance(value, tuple) else value
    return func(clean_params) or {}


def page_param() -> str:
    value = st.query_params.get("page", "home")
    if isinstance(value, list):
        value = value[0] if value else "home"
    return str(value or "home").lower()


def navigate(page: str) -> None:
    st.query_params["page"] = page
    st.rerun()


def fmt_brl(value: object, *, compact_threshold: float = 1000.0) -> str:
    try:
        number = float(value or 0)
    except (TypeError, ValueError):
        number = 0.0
    if abs(number) >= compact_threshold:
        integer = int(number) if number >= 0 else -int(abs(number))
        formatted = f"{integer:,}".replace(",", ".")
        return f"R$ {formatted}"
    formatted = f"{number:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {formatted}"


def fmt_brl_compact(value: object) -> str:
    try:
        number = float(value or 0)
    except (TypeError, ValueError):
        number = 0.0
    formatted = f"{number:,.0f}".replace(",", ".")
    return f"R$ {formatted}"


def fmt_brl_big(value: object, *, threshold: float = 1000.0) -> str:
    return fmt_brl(value, compact_threshold=threshold)


def fmt_num(value: object, decimals: int = 0) -> str:
    try:
        number = float(value or 0)
    except (TypeError, ValueError):
        number = 0.0
    formatted = f"{number:,.{decimals}f}".replace(",", "X").replace(".", ",").replace("X", ".")
    if decimals == 0:
        return formatted.split(",")[0]
    return formatted


def fmt_peso(value: object) -> str:
    try:
        number = float(value or 0)
    except (TypeError, ValueError):
        number = 0.0
    if abs(number) >= 1000:
        tons = int(number / 1000) if number >= 0 else -int(abs(number) / 1000)
        return f"{fmt_num(tons)} ton"
    return f"{fmt_num(number, 2)} kg"


@st.cache_data(ttl=600, show_spinner=False)
def last_update_label(version: str) -> str:
    root = Path(__file__).resolve().parent
    dt: datetime | None = None
    try:
        output = subprocess.check_output(
            ["git", "-C", str(root), "log", "-1", "--format=%cI"],
            text=True,
            stderr=subprocess.DEVNULL,
            timeout=2,
        ).strip()
        if output:
            dt = datetime.fromisoformat(output.replace("Z", "+00:00"))
    except Exception:
        dt = None
    if dt is None:
        mtimes = [Path(__file__).stat().st_mtime]
        backend_path = root / "app.py"
        if backend_path.exists():
            mtimes.append(backend_path.stat().st_mtime)
        dt = datetime.fromtimestamp(max(mtimes), tz=BR_TZ)
    return dt.astimezone(BR_TZ).strftime("%d/%m/%Y %H:%M")


def bar_value_text(value: object, *, currency: bool = True) -> str:
    try:
        number = float(value or 0)
    except (TypeError, ValueError):
        number = 0.0
    if abs(number) < 0.005:
        return ""
    return fmt_brl_compact(number) if currency else fmt_num(number)


def h(text: object) -> str:
    escaped = html.escape(clean_text(text), quote=True)
    return escaped.encode("ascii", "xmlcharrefreplace").decode("ascii")


def clean_text(text: object) -> str:
    value = str(text if text is not None else "")
    if any(marker in value for marker in ("Ã", "Â", "â€", "â€¢")):
        try:
            value = value.encode("latin1").decode("utf-8")
        except UnicodeError:
            replacements = {
                "Ã¡": "á",
                "Ã¢": "â",
                "Ã£": "ã",
                "Ã©": "é",
                "Ãª": "ê",
                "Ã­": "í",
                "Ã³": "ó",
                "Ã´": "ô",
                "Ãµ": "õ",
                "Ãº": "ú",
                "Ã§": "ç",
                "Â©": "©",
                "â€¢": "•",
            }
            for bad, good in replacements.items():
                value = value.replace(bad, good)
    broken = r"[\ufffd?]"

    def same_case(match: re.Match, replacement: str) -> str:
        original = match.group(0)
        if original.isupper():
            return replacement.upper()
        if original[:1].isupper():
            return replacement[:1].upper() + replacement[1:]
        return replacement

    word_replacements = [
        (fr"combust{broken}vel", "combustível"),
        (fr"manuten{broken}+o", "manutenção"),
        (fr"ped{broken}gio", "pedágio"),
        (fr"hot{broken}is", "hotéis"),
        (fr"servi{broken}os", "serviços"),
        (fr"lan{broken}amento", "lançamento"),
        (fr"edi{broken}+o", "edição"),
        (fr"gr{broken}fico", "gráfico"),
        (fr"p{broken}gina", "página"),
        (fr"m{broken}dia", "média"),
        (fr"m{broken}dio", "médio"),
        (fr"m{broken}s", "mês"),
        (fr"{broken}ltimos", "últimos"),
    ]
    for pattern, replacement in word_replacements:
        value = re.sub(pattern, lambda match, repl=replacement: same_case(match, repl), value, flags=re.IGNORECASE)

    value = re.sub(fr"\s*{broken}\s*", " - ", value)
    value = re.sub(r"\s{2,}", " ", value)
    return value.strip()


def selected_or_default(options: list, current: object | None = None, *, default_current_year: bool = False) -> object:
    all_options = ["Todos", *[item for item in options if item != "Todos"]]
    if current in all_options:
        return current
    if default_current_year and CURRENT_YEAR in all_options:
        return CURRENT_YEAR
    if default_current_year and str(CURRENT_YEAR) in all_options:
        return str(CURRENT_YEAR)
    return "Todos"


def month_label(value: object) -> str:
    if value == "Todos":
        return "Todos"
    try:
        month = int(value)
    except (TypeError, ValueError):
        return str(value)
    if 1 <= month <= 12:
        return f"{month:02d} - {MONTH_NAMES[month - 1]}"
    return str(value)


def month_filter_label(value: object) -> str:
    if value == "Todos":
        return "Todos"
    key = parse_month_key(value)
    if not key:
        return str(value)
    year, month = key
    return f"{MONTH_ABBR[month - 1]}/{str(year)[-2:]}"


def select_all_label(label: str):
    return lambda value: f"{label} (Todos)" if value == "Todos" else str(value)


def unique_filter_options(options: list[object]) -> list[object]:
    cleaned: list[object] = []
    seen: set[str] = set()
    for item in options or []:
        if item is None or item == "Todos":
            continue
        text = str(item).strip()
        if not text or text.lower() in {"nan", "none", "nat", "<na>"}:
            continue
        key = text.upper()
        if key in seen:
            continue
        seen.add(key)
        cleaned.append(item)
    return cleaned


def normalize_multiselect(values: list[object], previous: list[object] | None = None) -> list[object]:
    values = list(values or [])
    previous = list(previous or [])
    if not values:
        return ["Todos"]
    if "Todos" in values and len(values) > 1:
        if "Todos" in previous and len(previous) == 1:
            return [item for item in values if item != "Todos"] or ["Todos"]
        if "Todos" not in previous:
            return ["Todos"]
        return [item for item in values if item != "Todos"] or ["Todos"]
    return values


def sync_multiselect_selection(key: str) -> None:
    previous_key = f"{key}__previous"
    normalized = normalize_multiselect(
        st.session_state.get(key, []),
        st.session_state.get(previous_key, ["Todos"]),
    )
    st.session_state[key] = normalized
    st.session_state[previous_key] = normalized


def query_mes(values: list[object]) -> list[str]:
    cleaned = normalize_multiselect(values)
    if cleaned == ["Todos"]:
        return ["Todos"]
    return [str(item) for item in cleaned]


def topbar(title: str, *, back: bool = True) -> None:
    logo = logo_data_uri()
    back_html = '<a class="jr-back" href="?page=home" target="_self">&larr; Voltar</a>' if back else ""
    st.markdown(
        f"""
        <header class="jr-topbar">
          <div class="jr-brand">
            <img class="jr-logo" src="{logo}" alt="JR">
            <h1>{h(title)}</h1>
          </div>
          {back_html}
        </header>
        """,
        unsafe_allow_html=True,
    )


def render_kpis(items: list[tuple]) -> None:
    def card_html(item: tuple) -> str:
        label, value = item[0], item[1]
        color = item[2] if len(item) > 2 and item[2] else JR_RED
        return f'<div class="kpi" style="--kpi-accent: {h(color)}"><div class="kpi-title">{h(label)}</div><div class="kpi-value">{h(value)}</div></div>'

    has_sections = any(len(item) > 3 and item[3] for item in items)
    if not has_sections:
        st.html(f'<section class="kpis">{"".join(card_html(item) for item in items)}</section>')
        return

    sections: dict[str, dict[str, object]] = {}
    for item in items:
        section = clean_text(item[3] if len(item) > 3 and item[3] else "Resumo").strip() or "Resumo"
        color = item[2] if len(item) > 2 and item[2] else JR_RED
        sections.setdefault(section, {"color": color, "items": []})
        sections[section]["items"].append(item)

    section_html = []
    for section, payload in sections.items():
        color = payload["color"]
        cards = "".join(card_html(item) for item in payload["items"])
        summary_class = " kpi-section--summary" if section == "Comparativo geral" else ""
        section_html.append(
            f'<section class="kpi-section{summary_class}" style="--section-accent: {h(color)}">'
            f'<h3 class="kpi-section-title">{h(section)}</h3>'
            f'<div class="kpis">{cards}</div>'
            f'</section>'
        )
    st.html(f'<div class="kpi-sections">{"".join(section_html)}</div>')


def render_home_totals(cards: list[tuple[str, str, str]]) -> None:
    body = []
    for label, value, status in cards:
        body.append(
            f'<div class="home-total-card home-total-card--secondary"><p class="home-total-label">{h(label)}</p><p class="home-total-value">{h(value)}</p><p class="home-total-status">{h(status)}</p></div>'
        )
    st.markdown(f'<div class="home-total-grid">{"".join(body)}</div>', unsafe_allow_html=True)


def apply_theme(fig: go.Figure, *, height: int = 360, margin: dict | None = None) -> go.Figure:
    fig.update_layout(
        paper_bgcolor="#fff",
        plot_bgcolor="#fff",
        font={"family": "Inter, sans-serif", "color": JR_BLUE},
        colorway=[JR_BLUE, JR_RED, "#4B5563", "#9CA3AF"],
        height=height,
        margin=margin or {"l": 60, "r": 30, "t": 28, "b": 62},
        hovermode="closest",
    )
    fig.update_xaxes(gridcolor="#e5e7eb", zerolinecolor="#e5e7eb", linecolor="#e5e7eb")
    fig.update_yaxes(gridcolor="#e5e7eb", zerolinecolor="#e5e7eb", linecolor="#e5e7eb")
    return fig


def figure_meta(fig: go.Figure) -> dict:
    return dict(fig.layout.meta) if isinstance(fig.layout.meta, dict) else {}


def update_figure_meta(fig: go.Figure, **values) -> go.Figure:
    meta = figure_meta(fig)
    meta.update(values)
    fig.update_layout(meta=meta)
    return fig


def _compare_chart_full_width(series_count: int) -> bool:
    return series_count >= 3


def parse_month_key(value: object, fallback_year: object | None = None) -> tuple[int, int] | None:
    if value is None:
        return None
    text = str(value).strip().lower()
    if not text:
        return None
    parsed = pd.to_datetime(text, errors="coerce")
    if pd.notna(parsed):
        return int(parsed.year), int(parsed.month)
    for idx, abbr in enumerate(MONTH_ABBR, start=1):
        if abbr in text:
            year = None
            digits = "".join(ch if ch.isdigit() else " " for ch in text).split()
            if digits:
                year = int(digits[-1])
                if year < 100:
                    year += 2000
            elif fallback_year not in (None, "Todos"):
                try:
                    year = int(fallback_year)
                except (TypeError, ValueError):
                    year = None
            if year:
                return year, idx
    return None


def format_month_for_chart(value: object, *, include_year: bool = True, fallback_year: object | None = None) -> str:
    key = parse_month_key(value, fallback_year)
    if not key:
        return str(value)
    year, month = key
    label = MONTH_ABBR[month - 1]
    return f"{label}/{str(year)[-2:]}" if include_year else label


def sorted_series(data: dict, label_key: str, value_key: str, *, include_year: bool = True, fallback_year=None) -> tuple[list[str], list[float]]:
    labels = list(data.get(label_key, []) or [])
    values = list(data.get(value_key, []) or [])
    rows = []
    for index, label in enumerate(labels):
        value = values[index] if index < len(values) else 0
        key = parse_month_key(label, fallback_year)
        order = (key[0] * 100 + key[1]) if key and include_year else (key[1] if key else index)
        rows.append((order, index, label, float(value or 0)))
    rows.sort(key=lambda item: (item[0], item[1]))
    return [format_month_for_chart(item[2], include_year=include_year, fallback_year=fallback_year) for item in rows], [item[3] for item in rows]


def yearly_month_series(data: dict, label_key: str, value_key: str, *, fallback_year=None) -> list[dict]:
    labels = list(data.get(label_key, []) or [])
    values = list(data.get(value_key, []) or [])
    grouped: dict[int, dict[int, float]] = {}
    for index, label in enumerate(labels):
        key = parse_month_key(label, fallback_year)
        if not key:
            continue
        year, month = key
        value = values[index] if index < len(values) else 0
        try:
            number = float(value or 0)
        except (TypeError, ValueError):
            number = 0.0
        grouped.setdefault(year, {})
        grouped[year][month] = grouped[year].get(month, 0.0) + number

    series = []
    for index, year in enumerate(sorted(grouped)):
        months = sorted(grouped[year])
        series.append(
            {
                "label": str(year),
                "color": YEAR_SERIES_COLORS[index % len(YEAR_SERIES_COLORS)],
                "raw_labels": [f"{year}-{month:02d}" for month in months],
                "values": [grouped[year][month] for month in months],
            }
        )
    return series


def yearly_month_line_chart(data: dict, label_key: str, value_key: str, *, fallback_year=None) -> go.Figure:
    return multi_line_chart(yearly_month_series(data, label_key, value_key, fallback_year=fallback_year), include_year=False, fallback_year=fallback_year)


def yearly_month_bar_chart(
    data: dict,
    label_key: str,
    value_key: str,
    *,
    fallback_year=None,
    currency: bool = True,
) -> go.Figure:
    return multi_bar_chart(
        yearly_month_series(data, label_key, value_key, fallback_year=fallback_year),
        currency=currency,
        label_mode="month",
        fallback_year=fallback_year,
    )


def line_chart(labels: list[str], values: list[float]) -> go.Figure:
    values = [float(value or 0) for value in values]
    if values:
        min_value = min(values)
        max_value = max(values)
        padding = (max_value - min_value) * 0.14 if max_value != min_value else max(1, abs(max_value) * 0.08)
        y_range = [max(0, min_value - padding), max_value + padding]
    else:
        y_range = None
    fig = go.Figure(
        go.Scatter(
            x=labels,
            y=values,
            mode="lines+markers+text",
            line={"color": JR_BLUE, "width": 3},
            marker={"color": JR_RED, "size": 7},
            text=[fmt_brl_compact(value) for value in values],
            textposition="top center",
            textfont={"size": 10, "color": JR_BLUE},
            cliponaxis=False,
            hovertemplate="<b>%{x}</b><br>R$ %{y:,.2f}<extra></extra>",
        )
    )
    fig.update_xaxes(tickangle=-30, type="category")
    fig.update_yaxes(title="R$", range=y_range, automargin=True)
    return apply_theme(fig, height=370, margin={"l": 60, "r": 60, "t": 58, "b": 60})


def bar_chart(
    labels: list,
    values: list,
    *,
    horizontal: bool = False,
    sort_desc: bool = False,
    currency: bool = True,
    show_text: bool = False,
    height: int | None = None,
    marker_colors: list[str] | None = None,
) -> go.Figure:
    rows = [(str(label), float(value or 0)) for label, value in zip(labels or [], values or [])]
    if sort_desc:
        rows.sort(key=lambda item: item[1], reverse=True)
    labels_clean = [item[0] for item in rows]
    values_clean = [item[1] for item in rows]
    text = None
    if show_text:
        text = [bar_value_text(value, currency=currency) for value in values_clean]

    if horizontal:
        chart_height = height or max(320, 95 + len(labels_clean) * 30)
        fig = go.Figure(
            go.Bar(
                x=values_clean,
                y=labels_clean,
                orientation="h",
                marker={"color": marker_colors or JR_BLUE},
                text=text,
                textposition="outside" if show_text else "none",
                textfont={"size": 11, "color": JR_BLUE},
                cliponaxis=False,
                hovertemplate="<b>%{y}</b><br>R$ %{x:,.2f}<extra></extra>" if currency else "<b>%{y}</b><br>%{x:,.0f}<extra></extra>",
            )
        )
        max_value = max(values_clean) if values_clean else 0
        fig.update_xaxes(
            range=[0, max_value * (1.32 if show_text else 1.08)] if max_value else None,
            tickprefix="R$ " if currency else "",
            rangemode="tozero",
        )
        fig.update_yaxes(autorange="reversed", automargin=True, tickfont={"size": 11})
        fig = apply_theme(fig, height=chart_height, margin={"l": 130, "r": 95 if show_text else 45, "t": 20, "b": 45})
        fig.update_layout(meta={"jr_horizontal_bar": True, "jr_row_count": len(labels_clean)})
        return fig

    chart_height = height or 340
    fig = go.Figure(
        go.Bar(
            x=labels_clean,
            y=values_clean,
            marker={"color": marker_colors or JR_BLUE},
            text=text,
            textposition="outside" if show_text else "none",
            cliponaxis=False,
            hovertemplate="<b>%{x}</b><br>R$ %{y:,.2f}<extra></extra>" if currency else "<b>%{x}</b><br>%{y:,.0f}<extra></extra>",
        )
    )
    max_value = max(values_clean) if values_clean else 0
    fig.update_xaxes(tickangle=-30, automargin=True, type="category")
    fig.update_yaxes(
        title="R$" if currency else "",
        tickprefix="R$ " if currency else "",
        range=[0, max_value * (1.18 if show_text else 1.08)] if max_value else None,
        rangemode="tozero",
    )
    return apply_theme(fig, height=chart_height)


def pie_chart(labels: list, values: list) -> go.Figure:
    fig = go.Figure(
        go.Pie(
            labels=labels or [],
            values=[float(value or 0) for value in values or []],
            hole=0.45,
            textinfo="percent",
            textposition="inside",
            hovertemplate="<b>%{label}</b><br>R$ %{value:,.2f}<extra></extra>",
        )
    )
    fig.update_layout(showlegend=True, legend={"orientation": "h", "x": 0.5, "xanchor": "center", "y": -0.2})
    return apply_theme(fig, height=360, margin={"l": 16, "r": 16, "t": 30, "b": 90})


def peso_pie_chart(labels: list, values: list, city_counts: list | None = None, city_summaries: dict | None = None) -> go.Figure:
    values_clean = [float(value or 0) for value in values or []]
    counts = list(city_counts or [])
    summaries = city_summaries or {}
    labels_clean = [str(label) for label in labels or []]
    customdata = []
    for index, value in enumerate(values_clean):
        label = labels_clean[index] if index < len(labels_clean) else ""
        count = int(counts[index] or 0) if index < len(counts) else 0
        if count > 0 and summaries.get(label):
            customdata.append(f"{fmt_peso(value)} em {count} cidade(s)<br>Cidades: {h(summaries[label])}")
        elif count > 0:
            customdata.append(f"{fmt_peso(value)} em {count} cidade(s)")
        else:
            customdata.append(fmt_peso(value))
    fig = go.Figure(
        go.Pie(
            labels=labels or [],
            values=values_clean,
            customdata=customdata,
            hole=0.45,
            textinfo="percent",
            textposition="inside",
            hovertemplate="<b>%{label}</b><br>%{customdata}<extra></extra>",
            marker={"colors": [JR_BLUE, JR_RED, "#D97706", "#64748B", "#3158B7", "#E66C7A", "#8892A6"]},
        )
    )
    fig.update_layout(showlegend=True, legend={"orientation": "h", "x": 0.5, "xanchor": "center", "y": -0.2})
    return apply_theme(fig, height=380, margin={"l": 16, "r": 16, "t": 30, "b": 92})


def _series_color(route: str) -> str:
    return DASHBOARD_META.get(route, {}).get("color", JR_BLUE)


def _series_label(route: str) -> str:
    return DASHBOARD_META.get(route, {}).get("label", route.title())


def _ordered_month_labels(series: list[dict], *, include_year: bool, fallback_year) -> list[str]:
    labels: dict[str, tuple[int, int]] = {}
    fallback_order = 0
    for item in series:
        for raw_label in item.get("raw_labels", []):
            formatted = format_month_for_chart(raw_label, include_year=include_year, fallback_year=fallback_year)
            key = parse_month_key(raw_label, fallback_year)
            if key:
                order = ((key[0] * 100 + key[1]) if include_year else key[1], 0)
            else:
                fallback_order += 1
                order = (999999, fallback_order)
            labels.setdefault(formatted, order)
    return [label for label, _ in sorted(labels.items(), key=lambda item: item[1])]


def _ordered_day_labels(series: list[dict]) -> list[str]:
    labels: dict[str, tuple[int, int, int]] = {}
    fallback_order = 0
    for item in series:
        for raw_label in item.get("raw_labels", []):
            text = clean_text(raw_label).strip()
            parts = text.split("/")
            try:
                day = int(parts[0])
                month = int(parts[1]) if len(parts) > 1 else 0
                order = (month, day, 0)
            except (TypeError, ValueError):
                fallback_order += 1
                order = (99, 99, fallback_order)
            labels.setdefault(text, order)
    return [label for label, _ in sorted(labels.items(), key=lambda item: item[1])]


def _series_value_map(
    labels: list,
    values: list,
    *,
    label_formatter=None,
) -> dict[str, float]:
    mapped: dict[str, float] = {}
    for index, raw_label in enumerate(labels or []):
        label = label_formatter(raw_label) if label_formatter else clean_text(raw_label).strip()
        value = values[index] if index < len(values or []) else 0
        try:
            number = float(value or 0)
        except (TypeError, ValueError):
            number = 0.0
        mapped[label] = mapped.get(label, 0.0) + number
    return mapped


def multi_line_chart(series: list[dict], *, include_year: bool = True, fallback_year=None) -> go.Figure:
    clean_series = [item for item in series if item.get("raw_labels")]
    series_count = len(clean_series)
    if not clean_series:
        return line_chart([], [])
    if series_count == 1:
        item = clean_series[0]
        labels, values = sorted_series(
            {"Mes": item.get("raw_labels", []), "Valor": item.get("values", [])},
            "Mes",
            "Valor",
            include_year=include_year,
            fallback_year=fallback_year,
        )
        return line_chart(labels, values)

    ordered_labels = _ordered_month_labels(clean_series, include_year=include_year, fallback_year=fallback_year)
    fig = go.Figure()
    all_values = []
    for item in clean_series:
        value_map = _series_value_map(
            item.get("raw_labels", []),
            item.get("values", []),
            label_formatter=lambda raw: format_month_for_chart(raw, include_year=include_year, fallback_year=fallback_year),
        )
        values = [value_map.get(label, 0.0) for label in ordered_labels]
        all_values.extend(values)
        color = item.get("color", JR_BLUE)
        fig.add_trace(
            go.Scatter(
                x=ordered_labels,
                y=values,
                name=item.get("label", ""),
                mode="lines+markers",
                line={"color": color, "width": 3},
                marker={"color": color, "size": 7},
                hovertemplate="<b>%{fullData.name}</b><br>%{x}<br>R$ %{y:,.2f}<extra></extra>",
            )
        )
    max_value = max(all_values) if all_values else 0
    fig.update_xaxes(tickangle=-30, type="category")
    fig.update_yaxes(title="R$", range=[0, max_value * 1.12] if max_value else None, automargin=True)
    fig.update_layout(showlegend=True, legend={"orientation": "h", "x": 0, "y": 1.14})
    fig = apply_theme(fig, height=430 if _compare_chart_full_width(series_count) else 370, margin={"l": 60, "r": 40, "t": 78, "b": 70})
    return update_figure_meta(fig, jr_compare_series_count=series_count, jr_full_width=_compare_chart_full_width(series_count))


def multi_bar_chart(
    series: list[dict],
    *,
    horizontal: bool = False,
    sort_desc: bool = False,
    currency: bool = True,
    height: int | None = None,
    label_mode: str = "day",
    fallback_year=None,
) -> go.Figure:
    clean_series = [item for item in series if item.get("raw_labels")]
    series_count = len(clean_series)
    if not clean_series:
        return bar_chart([], [], horizontal=horizontal, currency=currency)
    if series_count == 1:
        item = clean_series[0]
        labels = item.get("raw_labels", [])
        values = item.get("values", [])
        if label_mode == "month":
            labels, values = sorted_series(
                {"Mes": labels, "Valor": values},
                "Mes",
                "Valor",
                include_year=False,
                fallback_year=fallback_year,
            )
        return bar_chart(
            labels,
            values,
            horizontal=horizontal,
            sort_desc=sort_desc,
            currency=currency,
            show_text=horizontal or label_mode == "month",
            height=height,
            marker_colors=[item.get("color", JR_BLUE)] * len(labels),
        )

    if horizontal:
        totals: dict[str, float] = {}
        maps = []
        for item in clean_series:
            value_map = _series_value_map(item.get("raw_labels", []), item.get("values", []))
            maps.append(value_map)
            for label, value in value_map.items():
                totals[label] = totals.get(label, 0.0) + value
        ordered_labels = list(totals.keys())
        if sort_desc:
            ordered_labels.sort(key=lambda label: totals.get(label, 0.0), reverse=True)
        row_height = 54 if _compare_chart_full_width(series_count) else 34
        chart_height = height or max(500 if _compare_chart_full_width(series_count) else 360, 92 + len(ordered_labels) * row_height)
        fig = go.Figure()
        max_value = 0.0
        text_size = 13 if _compare_chart_full_width(series_count) else 10
        for item, value_map in zip(clean_series, maps):
            values = [value_map.get(label, 0.0) for label in ordered_labels]
            text = [bar_value_text(value, currency=currency) for value in values]
            max_value = max(max_value, max(values) if values else 0.0)
            fig.add_trace(
                go.Bar(
                    x=values,
                    y=ordered_labels,
                    orientation="h",
                    name=item.get("label", ""),
                    marker={"color": item.get("color", JR_BLUE)},
                    text=text,
                    textposition="outside",
                    textfont={"size": text_size, "color": item.get("color", JR_BLUE)},
                    cliponaxis=False,
                    hovertemplate="<b>%{fullData.name}</b><br>%{y}<br>R$ %{x:,.2f}<extra></extra>"
                    if currency
                    else "<b>%{fullData.name}</b><br>%{y}<br>%{x:,.0f}<extra></extra>",
                )
            )
        fig.update_layout(
            barmode="group",
            bargap=0.28 if _compare_chart_full_width(series_count) else 0.18,
            bargroupgap=0.08,
            showlegend=True,
            legend={"orientation": "h", "x": 0.02, "y": 1.03, "yanchor": "bottom", "font": {"size": 13}},
        )
        fig.update_xaxes(tickprefix="R$ " if currency else "", range=[0, max_value * 1.34] if max_value else None, rangemode="tozero")
        fig.update_yaxes(autorange="reversed", automargin=True, tickfont={"size": 11})
        fig = apply_theme(fig, height=chart_height, margin={"l": 140, "r": 120, "t": 34, "b": 50})
        return update_figure_meta(
            fig,
            jr_horizontal_bar=True,
            jr_row_count=len(ordered_labels),
            jr_compare_series_count=series_count,
            jr_full_width=_compare_chart_full_width(series_count),
        )

    label_formatter = None
    if label_mode == "month":
        ordered_labels = _ordered_month_labels(clean_series, include_year=False, fallback_year=fallback_year)
        label_formatter = lambda raw: format_month_for_chart(raw, include_year=False, fallback_year=fallback_year)
    else:
        ordered_labels = _ordered_day_labels(clean_series)
    if not ordered_labels:
        ordered_labels = [label for item in clean_series for label in item.get("raw_labels", [])]
    fig = go.Figure()
    max_value = 0.0
    text_size = 13 if _compare_chart_full_width(series_count) else 10
    for item in clean_series:
        value_map = _series_value_map(item.get("raw_labels", []), item.get("values", []), label_formatter=label_formatter)
        values = [value_map.get(label, 0.0) for label in ordered_labels]
        text = [bar_value_text(value, currency=currency) for value in values]
        max_value = max(max_value, max(values) if values else 0.0)
        fig.add_trace(
            go.Bar(
                x=ordered_labels,
                y=values,
                name=item.get("label", ""),
                marker={"color": item.get("color", JR_BLUE)},
                text=text,
                textposition="outside",
                textfont={"size": text_size, "color": item.get("color", JR_BLUE)},
                cliponaxis=False,
                hovertemplate="<b>%{fullData.name}</b><br>%{x}<br>R$ %{y:,.2f}<extra></extra>"
                if currency
                else "<b>%{fullData.name}</b><br>%{x}<br>%{y:,.0f}<extra></extra>",
            )
        )
    fig.update_layout(
        barmode="group",
        bargap=0.24 if _compare_chart_full_width(series_count) else 0.18,
        bargroupgap=0.08,
        showlegend=True,
        legend={"orientation": "h", "x": 0.02, "y": 1.05, "yanchor": "bottom", "font": {"size": 13 if _compare_chart_full_width(series_count) else 11}},
    )
    fig.update_xaxes(tickangle=-30, automargin=True, type="category")
    fig.update_yaxes(
        title="R$" if currency else "",
        tickprefix="R$ " if currency else "",
        range=[0, max_value * (1.38 if _compare_chart_full_width(series_count) else 1.26)] if max_value else None,
        rangemode="tozero",
    )
    chart_height = height or (460 if _compare_chart_full_width(series_count) else 360)
    fig = apply_theme(fig, height=chart_height, margin={"l": 66, "r": 52, "t": 54 if _compare_chart_full_width(series_count) else 88, "b": 74})
    return update_figure_meta(fig, jr_compare_series_count=series_count, jr_full_width=_compare_chart_full_width(series_count))


def report_font(size: int, *, bold: bool = False) -> ImageFont.ImageFont:
    candidates = [
        "arialbd.ttf" if bold else "arial.ttf",
        "segoeuib.ttf" if bold else "segoeui.ttf",
        "DejaVuSans-Bold.ttf" if bold else "DejaVuSans.ttf",
        "LiberationSans-Bold.ttf" if bold else "LiberationSans-Regular.ttf",
        "FreeSansBold.ttf" if bold else "FreeSans.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
        "/usr/share/fonts/truetype/freefont/FreeSansBold.ttf" if bold else "/usr/share/fonts/truetype/freefont/FreeSans.ttf",
        "C:/Windows/Fonts/arialbd.ttf" if bold else "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/segoeuib.ttf" if bold else "C:/Windows/Fonts/segoeui.ttf",
    ]
    for candidate in candidates:
        try:
            return ImageFont.truetype(candidate, size=size)
        except OSError:
            continue
    try:
        return ImageFont.load_default(size=size)
    except TypeError:
        return ImageFont.load_default()


def draw_wrapped_text(
    draw: ImageDraw.ImageDraw,
    text: str,
    xy: tuple[int, int],
    *,
    font: ImageFont.ImageFont,
    fill: str,
    max_width: int,
    line_gap: int = 6,
) -> int:
    words = clean_text(text).split()
    lines: list[str] = []
    current = ""
    for word in words:
        trial = f"{current} {word}".strip()
        if draw.textbbox((0, 0), trial, font=font)[2] <= max_width or not current:
            current = trial
        else:
            lines.append(current)
            current = word
    if current:
        lines.append(current)
    x, y = xy
    line_height = draw.textbbox((0, 0), "Ag", font=font)[3] + line_gap
    for line in lines:
        draw.text((x, y), line, font=font, fill=fill)
        y += line_height
    return y


def draw_centered_wrapped_text(
    draw: ImageDraw.ImageDraw,
    text: str,
    center_x: int,
    y: int,
    *,
    font: ImageFont.ImageFont,
    fill: str,
    max_width: int,
    line_gap: int = 6,
) -> int:
    words = clean_text(text).split()
    lines: list[str] = []
    current = ""
    for word in words:
        trial = f"{current} {word}".strip()
        if draw.textbbox((0, 0), trial, font=font)[2] <= max_width or not current:
            current = trial
        else:
            lines.append(current)
            current = word
    if current:
        lines.append(current)
    line_height = draw.textbbox((0, 0), "Ag", font=font)[3] + line_gap
    for line in lines:
        bbox = draw.textbbox((0, 0), line, font=font)
        draw.text((center_x - (bbox[2] - bbox[0]) // 2, y), line, font=font, fill=fill)
        y += line_height
    return y


def horizontal_bar_row_count(fig: go.Figure) -> int:
    meta = figure_meta(fig)
    if meta.get("jr_horizontal_bar"):
        try:
            return int(meta.get("jr_row_count") or 0)
        except (TypeError, ValueError):
            return 0
    for trace in fig.data:
        if getattr(trace, "type", "") == "bar" and getattr(trace, "orientation", None) == "h":
            return len(getattr(trace, "y", []) or [])
    return 0


def chart_scroll_height(fig: go.Figure) -> int | None:
    row_count = horizontal_bar_row_count(fig)
    if row_count <= 10:
        return None
    meta = figure_meta(fig)
    try:
        series_count = int(meta.get("jr_compare_series_count") or 1)
    except (TypeError, ValueError):
        series_count = 1
    return 560 if series_count >= 3 else 420


def chart_prefers_full_width(fig: go.Figure) -> bool:
    return bool(figure_meta(fig).get("jr_full_width"))


def export_ready_figure(fig: go.Figure, *, max_horizontal_rows: int = 10) -> go.Figure:
    export_fig = go.Figure(fig)
    for trace in export_fig.data:
        if getattr(trace, "type", "") != "bar" or getattr(trace, "orientation", None) != "h":
            continue
        labels = list(getattr(trace, "y", []) or [])
        if len(labels) <= max_horizontal_rows:
            continue
        trace.y = labels[:max_horizontal_rows]
        trace.x = list(getattr(trace, "x", []) or [])[:max_horizontal_rows]
        if getattr(trace, "text", None) is not None:
            trace.text = list(trace.text)[:max_horizontal_rows]
        marker_color = getattr(trace.marker, "color", None)
        if isinstance(marker_color, (list, tuple)):
            trace.marker.color = list(marker_color)[:max_horizontal_rows]
        export_fig.update_layout(height=max(420, 95 + max_horizontal_rows * 30))
    return export_fig


def compose_home_export_image(
    cards: list[tuple[str, str, str]],
    *,
    include_header: bool,
) -> Image.Image:
    width = 1500
    margin = 54
    gap = 28
    bg = "#ffffff"
    panel = "#ffffff"
    line = "#cbd9ff"
    shadow = "#e4ebf8"
    card_cols = min(3, len(cards)) if cards else 1
    card_width = (width - margin * 2 - gap * (card_cols - 1)) // card_cols
    card_height = 205
    header_height = 140 if include_header else 0
    height = margin + header_height + card_height + margin
    canvas = Image.new("RGB", (width, height), bg)
    draw = ImageDraw.Draw(canvas)
    y = margin

    if include_header:
        if LOGO_PATH.exists():
            try:
                logo = Image.open(LOGO_PATH).convert("RGBA").resize((64, 64))
                canvas.paste(logo, (margin, y), logo)
            except Exception:
                logo = None
        draw.text((margin + 84, y + 4), "Dashboards operacionais", font=report_font(36, bold=True), fill=JR_BLUE)
        draw_wrapped_text(
            draw,
            "Monitoramento consolidado do JR Dashboard.",
            (margin + 84, y + 56),
            font=report_font(20),
            fill=MUTED,
            max_width=760,
        )
        y += header_height

    for index, (label, value, status) in enumerate(cards):
        x = margin + index * (card_width + gap)
        draw.rounded_rectangle(
            (x + 6, y + 8, x + card_width + 6, y + card_height + 8),
            radius=18,
            fill=shadow,
        )
        draw.rounded_rectangle(
            (x, y, x + card_width, y + card_height),
            radius=18,
            fill=panel,
            outline=line,
            width=2,
        )
        center_x = x + card_width // 2
        draw_centered_wrapped_text(
            draw,
            clean_text(label).upper(),
            center_x,
            y + 26,
            font=report_font(20, bold=True),
            fill=MUTED,
            max_width=card_width - 60,
            line_gap=5,
        )
        value_text = clean_text(value)
        value_font = report_font(30, bold=True)
        value_bbox = draw.textbbox((0, 0), value_text, font=value_font)
        draw.text((center_x - (value_bbox[2] - value_bbox[0]) // 2, y + 95), value_text, font=value_font, fill=JR_BLUE)
        draw_centered_wrapped_text(
            draw,
            status,
            center_x,
            y + 138,
            font=report_font(16),
            fill="#374151",
            max_width=card_width - 52,
            line_gap=4,
        )

    return canvas


def compose_export_image(
    kpis: list[tuple[str, str]],
    charts: list[tuple[str, go.Figure]],
    *,
    title: str,
    include_cards: bool,
    include_charts: bool,
) -> Image.Image:
    _ = title
    width = 1800
    margin = 58
    gap = 28
    bg = "#ffffff"
    panel = "#ffffff"
    line = "#cbd9ff"
    text = JR_BLUE
    shadow = "#e4ebf8"
    card_cols = min(3, len(kpis)) if include_cards and kpis else 1
    card_width = (width - margin * 2 - gap * (card_cols - 1)) // card_cols if card_cols else 0
    card_height = 174
    chart_cols = 2
    chart_card_width = (width - margin * 2 - gap) // chart_cols
    chart_inner_width = chart_card_width - 56
    chart_images: list[tuple[str, Image.Image]] = []

    if include_charts:
        for chart_title, fig in charts:
            try:
                export_fig = export_ready_figure(fig)
                desired_height = max(540, min(780, int(export_fig.layout.height or 460) + 110))
                pie_label_count = 0
                for trace in export_fig.data:
                    if getattr(trace, "type", "") == "pie":
                        pie_label_count = max(pie_label_count, len(getattr(trace, "labels", []) or []))
                if pie_label_count > 10:
                    export_fig.update_layout(showlegend=False)
                    export_fig.update_traces(textinfo="percent", textposition="inside")
                    desired_height = max(desired_height, 600)
                export_fig.update_layout(
                    width=chart_inner_width,
                    height=desired_height,
                    paper_bgcolor="#ffffff",
                    plot_bgcolor="#ffffff",
                    font={"family": "Arial, sans-serif", "size": 17, "color": JR_BLUE},
                    margin={"l": 80, "r": 42, "t": 34, "b": 84},
                )
                export_fig.update_xaxes(tickfont={"size": 16}, title_font={"size": 18})
                export_fig.update_yaxes(tickfont={"size": 16}, title_font={"size": 18})
                export_fig.update_traces(textfont={"size": 15})
                png = export_fig.to_image(
                    format="png",
                    width=chart_inner_width,
                    height=desired_height,
                    scale=1,
                )
                chart_images.append((chart_title, Image.open(BytesIO(png)).convert("RGB")))
            except Exception:
                desired_height = 540
                placeholder = Image.new("RGB", (chart_inner_width, desired_height), "#ffffff")
                draw = ImageDraw.Draw(placeholder)
                draw.text(
                    (32, 32),
                    f"Não foi possível renderizar: {clean_text(chart_title)}",
                    font=report_font(26, bold=True),
                    fill=JR_RED,
                )
                chart_images.append((chart_title, placeholder))

    rows = 0
    if include_cards and kpis:
        rows = (len(kpis) + card_cols - 1) // card_cols
    height = margin
    if rows:
        height += rows * card_height + (rows - 1) * gap
        if include_charts and chart_images:
            height += gap + 10
    if include_charts and chart_images:
        chart_row_heights = []
        for row_start in range(0, len(chart_images), chart_cols):
            row_items = chart_images[row_start : row_start + chart_cols]
            chart_row_heights.append(max(image.height for _, image in row_items) + 92)
        height += sum(chart_row_heights) + gap * max(0, len(chart_row_heights) - 1)
    height += margin
    if height < 500:
        height = 500

    canvas = Image.new("RGB", (width, height), bg)
    draw = ImageDraw.Draw(canvas)
    y = margin

    if include_cards and kpis:
        for index, (label, value) in enumerate(kpis):
            row = index // card_cols
            col = index % card_cols
            x = margin + col * (card_width + gap)
            card_y = y + row * (card_height + gap)
            draw.rounded_rectangle(
                (x + 6, card_y + 8, x + card_width + 6, card_y + card_height + 8),
                radius=20,
                fill=shadow,
            )
            draw.rounded_rectangle(
                (x, card_y, x + card_width, card_y + card_height),
                radius=20,
                fill=panel,
                outline=line,
                width=2,
            )
            draw_wrapped_text(
                draw,
                clean_text(label).upper(),
                (x + 30, card_y + 28),
                font=report_font(22, bold=True),
                fill=MUTED,
                max_width=card_width - 60,
                line_gap=4,
            )
            value_size = 42
            value_text = clean_text(value)
            if len(value_text) > 18:
                value_size = 36
            if len(value_text) > 28:
                value_size = 31
            value_font = report_font(value_size, bold=True)
            value_bbox = draw.textbbox((0, 0), value_text, font=value_font)
            value_x = x + (card_width - (value_bbox[2] - value_bbox[0])) // 2
            draw.text((value_x, card_y + 104), value_text, font=value_font, fill=JR_RED)
        y += rows * card_height + (rows - 1) * gap
        if include_charts and chart_images:
            y += gap + 10

    if include_charts and chart_images:
        for row_start in range(0, len(chart_images), chart_cols):
            row_items = chart_images[row_start : row_start + chart_cols]
            row_height = max(image.height for _, image in row_items) + 92
            for col, (chart_title, image) in enumerate(row_items):
                x = margin + col * (chart_card_width + gap)
                chart_y = y
                draw.rounded_rectangle(
                    (x + 6, chart_y + 8, x + chart_card_width + 6, chart_y + row_height + 8),
                    radius=22,
                    fill=shadow,
                )
                draw.rounded_rectangle(
                    (x, chart_y, x + chart_card_width, chart_y + row_height),
                    radius=22,
                    fill=panel,
                    outline=line,
                    width=2,
                )
                draw.text((x + 28, chart_y + 24), clean_text(chart_title), font=report_font(30, bold=True), fill=text)
                canvas.paste(image, (x + 28, chart_y + 72))
            y += row_height + gap
        y -= gap

    return canvas.crop((0, 0, width, min(height, y + margin)))


def image_to_png_bytes(image: Image.Image) -> bytes:
    buffer = BytesIO()
    image.save(buffer, format="PNG")
    return buffer.getvalue()


def image_to_pdf_bytes(image: Image.Image) -> bytes:
    buffer = BytesIO()
    image.convert("RGB").save(buffer, format="PDF", resolution=120.0)
    return buffer.getvalue()


def export_file_name(prefix: str, scope: str, ext: str) -> str:
    return f"{prefix}-{scope}.{ext}"


def chart_card(title: str, fig: go.Figure) -> None:
    with st.container(border=True):
        st.html(f'<div class="chart-title">{h(title)}</div>')
        scroll_height = chart_scroll_height(fig)
        if scroll_height:
            with st.container(height=scroll_height, border=False):
                st.plotly_chart(fig, width="stretch", config=PLOTLY_CONFIG)
        else:
            st.plotly_chart(fig, width="stretch", config=PLOTLY_CONFIG)


def chart_grid(charts: list[tuple[str, go.Figure]]) -> None:
    index = 0
    while index < len(charts):
        title, fig = charts[index]
        if chart_prefers_full_width(fig):
            chart_card(title, fig)
            index += 1
            continue

        if index + 1 < len(charts) and not chart_prefers_full_width(charts[index + 1][1]):
            cols = st.columns(2)
            for col, item in zip(cols, charts[index : index + 2]):
                with col:
                    chart_card(item[0], item[1])
            index += 2
        else:
            chart_card(title, fig)
            index += 1


def kpi_control_label(item: tuple) -> str:
    label = clean_text(item[1])
    section = clean_text(item[4] if len(item) > 4 and item[4] else "").strip()
    return f"{section} - {label}" if section else label


def dashboard_controls(
    key_prefix: str,
    *,
    title: str,
    kpis: list[tuple],
    charts: list[tuple[str, str, go.Figure]],
) -> tuple[bool, bool, list[tuple], list[tuple[str, str, go.Figure]]]:
    hide_cards_key = f"{key_prefix}_hide_cards"
    hide_charts_key = f"{key_prefix}_hide_charts"
    export_ready_key = f"{key_prefix}_export_ready"
    st.session_state.setdefault(hide_cards_key, False)
    st.session_state.setdefault(hide_charts_key, False)
    ready_file = st.session_state.get(export_ready_key)
    if ready_file and ready_file.get("version") != APP_VERSION:
        st.session_state.pop(export_ready_key, None)

    with st.container(key=f"{key_prefix}_dashboard_controls"):
        with st.expander("Controles", expanded=False):
            export_jobs = [
                ("cards_png", "Cards PNG", "cards", "png"),
                ("cards_pdf", "Cards PDF", "cards", "pdf"),
                ("charts_png", "Gráficos PNG", "graficos", "png"),
                ("charts_pdf", "Gráficos PDF", "graficos", "pdf"),
                ("page_png", "Página PNG", "pagina", "png"),
                ("page_pdf", "Página PDF", "pagina", "pdf"),
            ]
            selected_kpis_preview = [
                (kpi_control_label(item), item[2])
                for item in kpis
                if st.session_state.get(f"{key_prefix}_visible_{item[0]}", True)
            ]
            selected_charts_preview = [
                (chart_title, fig)
                for item_id, chart_title, fig in charts
                if st.session_state.get(f"{key_prefix}_visible_{item_id}", True)
            ]

            st.markdown('<div class="control-row-label">Exportar</div>', unsafe_allow_html=True)
            buttons = st.columns(6)
            for col, (job_id, label, scope, ext) in zip(buttons, export_jobs):
                with col:
                    if st.button(clean_text(label), key=f"{key_prefix}_export_{job_id}", width="stretch"):
                        include_cards = scope == "cards" or (scope == "pagina" and not st.session_state[hide_cards_key])
                        include_charts = scope == "graficos" or (scope == "pagina" and not st.session_state[hide_charts_key])
                        if not include_cards and not include_charts:
                            st.warning("Nada selecionado para exportar.")
                        elif include_cards and not selected_kpis_preview and not include_charts:
                            st.warning("Nenhum card selecionado para exportar.")
                        elif include_charts and not selected_charts_preview and not include_cards:
                            st.warning("Nenhum gráfico selecionado para exportar.")
                        elif include_cards and include_charts and not selected_kpis_preview and not selected_charts_preview:
                            st.warning("Nenhum item selecionado para exportar.")
                        else:
                            with st.spinner("Gerando arquivo..."):
                                image = compose_export_image(
                                    selected_kpis_preview,
                                    selected_charts_preview,
                                    title=title,
                                    include_cards=include_cards,
                                    include_charts=include_charts,
                                )
                                data = image_to_png_bytes(image) if ext == "png" else image_to_pdf_bytes(image)
                                st.session_state[export_ready_key] = {
                                    "data": data,
                                    "file_name": export_file_name(key_prefix, scope, ext),
                                    "mime": "image/png" if ext == "png" else "application/pdf",
                                    "version": APP_VERSION,
                                }

            ready = st.session_state.get(export_ready_key)
            if ready:
                st.download_button(
                    "Baixar arquivo gerado",
                    data=ready["data"],
                    file_name=ready["file_name"],
                    mime=ready["mime"],
                    key=f"{key_prefix}_download_ready",
                    width="stretch",
                )

            st.markdown('<div class="control-row-spacer"></div><div class="control-row-label">Exibir</div>', unsafe_allow_html=True)
            toggles = st.columns(2)
            with toggles[0]:
                toggle_label = "Mostrar cards" if st.session_state[hide_cards_key] else "Ocultar cards"
                if st.button(clean_text(toggle_label), key=f"{key_prefix}_toggle_cards", width="stretch"):
                    st.session_state[hide_cards_key] = not st.session_state[hide_cards_key]
                    st.rerun()
            with toggles[1]:
                toggle_label = "Mostrar gráficos" if st.session_state[hide_charts_key] else "Ocultar gráficos"
                if st.button(clean_text(toggle_label), key=f"{key_prefix}_toggle_charts", width="stretch"):
                    st.session_state[hide_charts_key] = not st.session_state[hide_charts_key]
                    st.rerun()

            st.markdown('<div class="control-row-spacer"></div><div class="control-row-label">Itens</div>', unsafe_allow_html=True)
            checkbox_items = [("cards", item[0], kpi_control_label(item)) for item in kpis]
            checkbox_items.extend(("charts", item_id, label) for item_id, label, _ in charts)
            checkbox_cols = 4
            for index in range(0, len(checkbox_items), checkbox_cols):
                cols = st.columns(checkbox_cols)
                for col, (_, item_id, label) in zip(cols, checkbox_items[index : index + checkbox_cols]):
                    state_key = f"{key_prefix}_visible_{item_id}"
                    st.session_state.setdefault(state_key, True)
                    with col:
                        st.checkbox(clean_text(label), key=state_key)

    show_cards = not st.session_state[hide_cards_key]
    show_charts = not st.session_state[hide_charts_key]
    visible_kpis = [item for item in kpis if st.session_state.get(f"{key_prefix}_visible_{item[0]}", True)]
    visible_charts = [item for item in charts if st.session_state.get(f"{key_prefix}_visible_{item[0]}", True)]
    return show_cards, show_charts, visible_kpis, visible_charts


def render_controlled_dashboard(
    key_prefix: str,
    *,
    title: str,
    kpis: list[tuple],
    charts: list[tuple[str, str, go.Figure]],
) -> None:
    show_cards, show_charts, visible_kpis, visible_charts = dashboard_controls(
        key_prefix,
        title=title,
        kpis=kpis,
        charts=charts,
    )
    if show_cards and visible_kpis:
        render_kpis([(item[1], item[2], item[3] if len(item) > 3 else None, item[4] if len(item) > 4 else None) for item in visible_kpis])
    if show_charts and visible_charts:
        chart_grid([(chart_title, fig) for _, chart_title, fig in visible_charts])


def filter_controls(
    route: str,
    *,
    extra_filters: list[tuple[str, str, list]],
    key_prefix: str,
    all_data: dict | None = None,
) -> tuple[dict[str, object], dict]:
    all_data = all_data if all_data is not None else route_json(route, {"ano": "Todos", "mes": ["Todos"]})
    years = all_data.get("anos", []) or []
    year_options = ["Todos", *years]
    year_default = selected_or_default(years, st.session_state.get(f"{key_prefix}_ano"), default_current_year=True)
    year_index = year_options.index(year_default) if year_default in year_options else 0

    with st.container(key=f"{key_prefix}_filterbar"):
        filter_widths = [1.1 if len(label) > 10 else 1.0 for _, label, _ in extra_filters]
        compare_options = [key for key in DASHBOARD_META if key in COMPARE_ALLOWED_ROUTES and key != route]
        widths = [0.85, 1.25, *filter_widths, 1.35, 1.05, 0.8]
        filter_cols = st.columns(widths)
        with filter_cols[0]:
            ano = st.selectbox(
                "Ano",
                year_options,
                index=year_index,
                key=f"{key_prefix}_ano",
                format_func=select_all_label("Ano"),
                label_visibility="collapsed",
            )

        month_data = all_data if ano == "Todos" else route_json(route, {"ano": ano, "mes": ["Todos"]})
        meses = unique_filter_options(month_data.get("meses", []) or [])
        mes_options = ["Todos", *meses]
        mes_key = f"{key_prefix}_mes"
        mes_previous_key = f"{mes_key}__previous"
        mes_state_exists = mes_key in st.session_state
        current_meses = st.session_state.get(mes_key, ["Todos"])
        current_meses = [item for item in current_meses if item in mes_options] or ["Todos"]
        if mes_state_exists and st.session_state.get(mes_key) != current_meses:
            st.session_state[mes_key] = current_meses
            st.session_state[mes_previous_key] = current_meses
        if mes_previous_key not in st.session_state:
            st.session_state[mes_previous_key] = current_meses
        with filter_cols[1]:
            mes_kwargs = {
                "key": mes_key,
                "on_change": sync_multiselect_selection,
                "args": (mes_key,),
                "format_func": month_filter_label,
                "label_visibility": "collapsed",
            }
            if not mes_state_exists:
                mes_kwargs["default"] = current_meses
            meses_selected = st.multiselect("Mês", mes_options, **mes_kwargs)
            meses_selected = normalize_multiselect(meses_selected, st.session_state.get(mes_previous_key, ["Todos"]))

        params: dict[str, object] = {"ano": None if ano == "Todos" else ano, "mes": query_mes(meses_selected)}
        for idx, (param_name, label, options) in enumerate(extra_filters, start=2):
            options = ["Todos", *unique_filter_options(options)]
            current = st.session_state.get(f"{key_prefix}_{param_name}", "Todos")
            if current not in options:
                current = "Todos"
            with filter_cols[idx]:
                params[param_name] = st.selectbox(
                    label,
                    options,
                    index=options.index(current),
                    key=f"{key_prefix}_{param_name}",
                    format_func=select_all_label(label),
                    label_visibility="collapsed",
                )

        compare_key = f"{key_prefix}_compare"
        compare_state_exists = compare_key in st.session_state
        compare_current = [item for item in st.session_state.get(compare_key, []) if item in compare_options]
        if compare_state_exists and st.session_state.get(compare_key) != compare_current:
            st.session_state[compare_key] = compare_current
        compare_kwargs = {
            "key": compare_key,
            "format_func": lambda value: DASHBOARD_META.get(value, {}).get("label", value),
            "label_visibility": "collapsed",
            "placeholder": "Comparar",
        }
        if not compare_state_exists:
            compare_kwargs["default"] = []
        with filter_cols[-3]:
            compare_selected = st.multiselect("Comparar", compare_options, **compare_kwargs)

        with filter_cols[-2]:
            if st.button("Limpar filtros", key=f"{key_prefix}_clear", width="stretch"):
                for state_key in list(st.session_state.keys()):
                    if state_key.startswith(f"{key_prefix}_"):
                        del st.session_state[state_key]
                st.rerun()
        with filter_cols[-1]:
            st.markdown('<a class="filter-back" href="?page=home" target="_self">&larr; Voltar</a>', unsafe_allow_html=True)

    for key, value in list(params.items()):
        if value == "Todos":
            params[key] = None
    all_data = dict(all_data)
    all_data["_compare"] = [item for item in compare_selected if item in compare_options]
    return params, all_data


def compare_query_params(route: str, params: dict[str, object]) -> dict[str, object]:
    query: dict[str, object] = {}
    if params.get("ano") is not None:
        query["ano"] = params.get("ano")
    if params.get("mes") is not None:
        query["mes"] = params.get("mes")
    if DASHBOARD_META.get(route, {}).get("supports_plate") and params.get("placa"):
        query["placa"] = params.get("placa")
    return query


def compare_bundle(
    current_route: str,
    current_data: dict,
    params: dict[str, object],
    selected_routes: list[str],
) -> list[tuple[str, dict]]:
    bundle = [(current_route, current_data)]
    for route in selected_routes:
        if route == current_route or route not in COMPARE_ALLOWED_ROUTES or route not in DASHBOARD_META:
            continue
        try:
            data = route_json(route, compare_query_params(route, params))
        except Exception:
            data = {}
        bundle.append((route, data))
    return bundle


def compare_chart_series(bundle: list[tuple[str, dict]], metric: str) -> list[dict]:
    series = []
    for route, data in bundle:
        config = COMPARE_SERIES.get(route, {}).get(metric)
        if not config:
            continue
        source_key, label_key, value_key = config
        source = data.get(source_key, {}) if isinstance(data, dict) else {}
        labels = list(source.get(label_key, []) or [])
        values = list(source.get(value_key, []) or [])
        series.append(
            {
                "route": route,
                "label": _series_label(route),
                "color": _series_color(route),
                "raw_labels": labels,
                "values": values,
            }
        )
    return series


def color_kpis(
    kpis: list[tuple],
    route: str,
    *,
    prefix: bool = False,
    exclude_ids: set[str] | None = None,
    section: str | None = None,
) -> list[tuple]:
    color = _series_color(route)
    route_label = _series_label(route)
    exclude_ids = exclude_ids or set()
    colored = []
    for item in kpis:
        item_id, label, value = item[0], item[1], item[2]
        if item_id in exclude_ids:
            continue
        display_label = f"{route_label} - {label}" if prefix else label
        if section:
            colored.append((item_id, display_label, value, color, section))
        else:
            colored.append((item_id, display_label, value, color))
    return colored


def _compare_total(route: str, data: dict) -> float:
    key = {
        "combustivel": "custo_total",
        "manutencao": "custo_total",
        "hoteis": "valor_total",
        "pedagio": "custo_total",
        "vex": "total_vex",
    }.get(route)
    if not key:
        return 0.0
    try:
        return float(data.get(key) or 0)
    except (TypeError, ValueError):
        return 0.0


def _compare_media(route: str, data: dict) -> float:
    key = {
        "combustivel": "media_mensal",
        "manutencao": "media_mensal",
        "hoteis": "media_mensal",
        "pedagio": "media_mensal",
    }.get(route)
    if not key:
        return 0.0
    try:
        return float(data.get(key) or 0)
    except (TypeError, ValueError):
        return 0.0


def compare_combined_kpis(bundle: list[tuple[str, dict]]) -> list[tuple]:
    total = sum(_compare_total(route, data) for route, data in bundle)
    media = sum(_compare_media(route, data) for route, data in bundle)
    return [
        ("compare_total_geral", "Total comparado (R$)", fmt_brl(total), JR_RED, "Comparativo geral"),
        ("compare_media_geral", "Média mensal comparada (R$)", fmt_brl(media), JR_RED, "Comparativo geral"),
    ]


def compare_kpi_cards(bundle: list[tuple[str, dict]]) -> list[tuple]:
    cards: list[tuple] = []
    for route, data in bundle:
        color = _series_color(route)
        label = _series_label(route)
        if route == "combustivel":
            route_cards = [
                ("total", "Total (R$)", fmt_brl(data.get("custo_total"))),
                ("media_mensal", "Média mensal (R$)", fmt_brl(data.get("media_mensal"))),
                ("km", "KM total", fmt_num(data.get("km_total"))),
                ("litros", "Total litros", fmt_num(data.get("litros_total"))),
                ("custo_km", "Custo médio por KM", fmt_brl(data.get("custo_por_km"))),
                ("kml", "Média KM/L", f"{fmt_num(data.get('km_por_litro'), 2)} km/L"),
                ("custo_litro", "Custo médio por litro", fmt_brl(data.get("custo_por_litro"))),
            ]
        elif route == "manutencao":
            route_cards = [
                ("custo_total", "Custo total (R$)", fmt_brl(data.get("custo_total"))),
                ("media_mensal", "Média mensal (R$)", fmt_brl(data.get("media_mensal"))),
                ("servicos", "Serviços", fmt_num(data.get("total_servicos"))),
                ("ticket", "Ticket médio", fmt_brl(data.get("media_servico"))),
            ]
        elif route == "hoteis":
            route_cards = [
                ("valor_total", "Valor total (R$)", fmt_brl(data.get("valor_total"))),
                ("media_mensal", "Média mensal", fmt_brl(data.get("media_mensal"))),
                ("reservas", "Reservas", fmt_num(data.get("reservas_total"))),
                ("nao_planejadas", "Reservas não planejadas", fmt_num(data.get("reservas_nao_planejadas"))),
                ("media_reserva", "Média por reserva", fmt_brl(data.get("valor_medio_reserva"))),
                ("sabados", "Gasto aos sábados (R$)", fmt_brl(data.get("valor_sabado"))),
                ("nao_planejado", "Gasto não planejado (R$)", fmt_brl(data.get("valor_nao_planejado"))),
            ]
        elif route == "pedagio":
            route_cards = [
                ("gasto_total", "Gasto total (R$)", fmt_brl(data.get("custo_total"))),
                ("media_mensal", "Média mensal (R$)", fmt_brl(data.get("media_mensal"))),
                ("pedagio", "Gasto pedágio", fmt_brl(data.get("gasto_pedagio"))),
                ("ipva", "Gasto IPVA", fmt_brl(data.get("gasto_ipva"))),
                ("seguro", "Gasto seguro", fmt_brl(data.get("gasto_seguro"))),
                ("media_valores", "Média de valores", fmt_brl(data.get("media_valores", data.get("ticket_medio")))),
            ]
        elif route == "vex":
            route_cards = [
                ("gasto_vex", "Gasto Vex total (R$)", fmt_brl(data.get("total_vex"))),
                ("combustivel", "Combustível Vex (R$)", fmt_brl(data.get("combustivel_total"))),
                ("km_total", "KM total", fmt_num(data.get("km_total"))),
                ("litros_total", "Total litros", fmt_num(data.get("litros_total"))),
                ("custo_km", "Custo médio por KM", fmt_brl(data.get("custo_por_km"))),
                ("media_kml", "Média KM/L", fmt_num(data.get("km_por_litro"), 2)),
                ("custo_litro", "Custo médio por litro", fmt_brl(data.get("custo_por_litro"))),
                ("manutencao", "Manutenção Vex (R$)", fmt_brl(data.get("manutencao_total"))),
                ("pedagio", "Pedágio/Seguro Vex (R$)", fmt_brl(data.get("pedagio_total"))),
            ]
        else:
            route_cards = []
        cards.extend((f"compare_{route}_{item_id}", card_label, value, color, label) for item_id, card_label, value in route_cards)
    return cards


def frota_filter_controls(seed: dict) -> dict[str, object]:
    years = seed.get("anos", []) or []
    year_options = ["Todos", *years]
    year_default = selected_or_default(years, st.session_state.get("rank_ano"), default_current_year=True)
    year_index = year_options.index(year_default) if year_default in year_options else 0

    with st.container(key="rank_filterbar"):
        cols = st.columns([0.75, 1.1, 1.0, 1.45, 1.05, 1.0, 0.72])
        with cols[0]:
            ano = st.selectbox(
                "Ano",
                year_options,
                index=year_index,
                key="rank_ano",
                format_func=select_all_label("Ano"),
                label_visibility="collapsed",
            )

        month_seed = seed if ano == "Todos" else route_json("frota", {"ano": ano, "mes": ["Todos"]})
        meses = unique_filter_options(month_seed.get("meses", []) or [])
        mes_options = ["Todos", *meses]
        mes_key = "rank_mes"
        mes_previous_key = f"{mes_key}__previous"
        mes_state_exists = mes_key in st.session_state
        current_meses = st.session_state.get(mes_key, ["Todos"])
        current_meses = [item for item in current_meses if item in mes_options] or ["Todos"]
        if mes_state_exists and st.session_state.get(mes_key) != current_meses:
            st.session_state[mes_key] = current_meses
            st.session_state[mes_previous_key] = current_meses
        if mes_previous_key not in st.session_state:
            st.session_state[mes_previous_key] = current_meses
        with cols[1]:
            mes_kwargs = {
                "key": mes_key,
                "on_change": sync_multiselect_selection,
                "args": (mes_key,),
                "format_func": month_filter_label,
                "label_visibility": "collapsed",
            }
            if not mes_state_exists:
                mes_kwargs["default"] = current_meses
            meses_selected = st.multiselect("Mês", mes_options, **mes_kwargs)
            meses_selected = normalize_multiselect(meses_selected, st.session_state.get(mes_previous_key, ["Todos"]))

        category_options = ["Todos", *unique_filter_options(seed.get("categorias", []) or [])]
        categoria_default = "Transporte" if "Transporte" in category_options else "Todos"
        categoria_current = st.session_state.get("rank_categoria", categoria_default)
        if categoria_current not in category_options:
            categoria_current = categoria_default
        with cols[2]:
            categoria = st.selectbox(
                "Categoria",
                category_options,
                index=category_options.index(categoria_current),
                key="rank_categoria",
                format_func=select_all_label("Categoria"),
                label_visibility="collapsed",
            )

        plate_seed = route_json(
            "frota",
            {
                "ano": None if ano == "Todos" else ano,
                "mes": query_mes(meses_selected),
                "categoria": None if categoria == "Todos" else categoria,
                "ordenar_por": "total",
            },
        )
        plate_options = ["Todos", *unique_filter_options(plate_seed.get("placas", []) or [])]
        plate_key = "rank_placa"
        plate_previous_key = f"{plate_key}__previous"
        plate_state_exists = plate_key in st.session_state
        current_plates = st.session_state.get(plate_key, ["Todos"])
        current_plates = [item for item in current_plates if item in plate_options] or ["Todos"]
        if plate_state_exists and st.session_state.get(plate_key) != current_plates:
            st.session_state[plate_key] = current_plates
            st.session_state[plate_previous_key] = current_plates
        if plate_previous_key not in st.session_state:
            st.session_state[plate_previous_key] = current_plates
        with cols[3]:
            plate_kwargs = {
                "key": plate_key,
                "on_change": sync_multiselect_selection,
                "args": (plate_key,),
                "label_visibility": "collapsed",
            }
            if not plate_state_exists:
                plate_kwargs["default"] = current_plates
            placas_selected = st.multiselect("Placa", plate_options, **plate_kwargs)
            placas_selected = normalize_multiselect(placas_selected, st.session_state.get(plate_previous_key, ["Todos"]))

        order_options = list(RANK_ORDER_OPTIONS)
        order_current = st.session_state.get("rank_ordenar_por", "total")
        if order_current not in order_options:
            order_current = "total"
        with cols[4]:
            ordenar_por = st.selectbox(
                "Ordenar por",
                order_options,
                index=order_options.index(order_current),
                key="rank_ordenar_por",
                format_func=lambda value: f"Ordenar: {RANK_ORDER_OPTIONS.get(value, value)}",
                label_visibility="collapsed",
            )

        with cols[5]:
            if st.button("Limpar filtros", key="rank_clear", width="stretch"):
                for state_key in list(st.session_state.keys()):
                    if state_key.startswith("rank_"):
                        del st.session_state[state_key]
                st.rerun()
        with cols[6]:
            st.markdown('<a class="filter-back" href="?page=home" target="_self">&larr; Voltar</a>', unsafe_allow_html=True)

    return {
        "ano": None if ano == "Todos" else ano,
        "mes": query_mes(meses_selected),
        "categoria": None if categoria == "Todos" else categoria,
        "placa": ["Todos"] if placas_selected == ["Todos"] else [str(item) for item in placas_selected],
        "ordenar_por": ordenar_por,
    }


def ranking_summary_html(row: dict) -> str:
    items = [
        ("Posição", f"#{int(row.get('rank') or 0):02d}"),
        ("Placa", row.get("placa") or "Sem placa"),
        ("Total", fmt_brl_big(row.get("total"))),
        ("Combustível", fmt_brl_big(row.get("combustivel"))),
        ("Manutenção", fmt_brl_big(row.get("manutencao"))),
        ("Pedágio/IPVA", fmt_brl_big(row.get("pedagio"))),
        ("Peso", fmt_peso(row.get("peso_total"))),
    ]
    return "".join(
        f'<div class="ranking-summary-item"><p class="ranking-summary-label">{h(label)}</p><p class="ranking-summary-value">{h(value)}</p></div>'
        for label, value in items
    )


def ranking_detail_html(row: dict) -> str:
    items = [
        ("Gasto total", fmt_brl_big(row.get("total"))),
        ("Combustível", fmt_brl_big(row.get("combustivel"))),
        ("Manutenção", fmt_brl_big(row.get("manutencao"))),
        ("Pedágio/IPVA", fmt_brl_big(row.get("pedagio"))),
        ("Peso", fmt_peso(row.get("peso_total"))),
        ("Valor entregas", fmt_brl_big(row.get("valor_peso"))),
        ("KM total", fmt_num(row.get("km_total"))),
        ("Litros", fmt_num(row.get("litros_total"))),
        ("Custo total por KM", fmt_brl(row.get("custo_por_km"))),
        ("Combustível por KM", fmt_brl(row.get("combustivel_por_km"))),
        ("Média KM/L", f"{fmt_num(row.get('km_por_litro'), 2)} km/L"),
        ("Custo por litro", fmt_brl(row.get("custo_por_litro"))),
        ("Abastecimentos", fmt_num(row.get("abastecimentos"))),
        ("Serviços", fmt_num(row.get("servicos"))),
        ("Pedágio/IPVA", fmt_num(row.get("despesas_pedagio"))),
    ]
    cards = "".join(
        f'<div class="ranking-detail-card"><p class="ranking-detail-label">{h(label)}</p><p class="ranking-detail-value">{h(value)}</p></div>'
        for label, value in items
    )
    return f'<div class="ranking-row-summary">{ranking_summary_html(row)}</div><div class="ranking-detail-grid">{cards}</div>'


def ranking_row_label(row: dict) -> str:
    def money(value: object) -> str:
        return fmt_brl_big(value).replace("$", r"\$")

    return (
        f"{int(row.get('rank') or 0):02d} | {row.get('placa') or 'Sem placa'} | "
        f"Total {money(row.get('total'))} | "
        f"Combustivel {money(row.get('combustivel'))} | "
        f"Manutencao {money(row.get('manutencao'))} | "
        f"Pedagio/IPVA {money(row.get('pedagio'))} | "
        f"Peso {fmt_peso(row.get('peso_total'))} | "
        f"KM {fmt_num(row.get('km_total'))} | "
        f"Litros {fmt_num(row.get('litros_total'))}"
    )


def _ranking_float(row: dict, key: str) -> float:
    try:
        return float(row.get(key) or 0)
    except (TypeError, ValueError):
        return 0.0


def ranking_difference_html(rows: list[dict]) -> str:
    if len(rows) < 2:
        return ""
    metrics = [
        ("Diferença total", "total", fmt_brl_big),
        ("Combustível", "combustivel", fmt_brl_big),
        ("Manutenção", "manutencao", fmt_brl_big),
        ("Pedágio/IPVA", "pedagio", fmt_brl_big),
        ("Peso", "peso_total", fmt_peso),
        ("KM total", "km_total", fmt_num),
        ("Litros", "litros_total", fmt_num),
        ("Custo/KM", "custo_por_km", fmt_brl),
        ("KM/L", "km_por_litro", lambda value: f"{fmt_num(value, 2)} km/L"),
    ]
    items = []
    for label, key, formatter in metrics:
        ordered = sorted(rows, key=lambda row: _ranking_float(row, key), reverse=True)
        high, low = ordered[0], ordered[-1]
        diff = abs(_ranking_float(high, key) - _ranking_float(low, key))
        note = f"{high.get('placa') or 'Sem placa'} acima de {low.get('placa') or 'Sem placa'}"
        items.append(
            f'<div class="ranking-difference-item"><p class="ranking-difference-label">{h(label)}</p>'
            f'<p class="ranking-difference-value">{h(formatter(diff))}</p>'
            f'<p class="ranking-difference-note">{h(note)}</p></div>'
        )
    return (
        '<section class="ranking-difference">'
        '<h3>Diferença entre as placas selecionadas</h3>'
        f'<div class="ranking-difference-grid">{"".join(items)}</div>'
        '</section>'
    )


def ranking_versus_html(rows: list[dict]) -> str:
    metrics = [
        ("Total", "total", fmt_brl_big),
        ("Combustível", "combustivel", fmt_brl_big),
        ("Manutenção", "manutencao", fmt_brl_big),
        ("Pedágio/IPVA", "pedagio", fmt_brl_big),
        ("Peso", "peso_total", fmt_peso),
        ("Valor entregas", "valor_peso", fmt_brl_big),
        ("KM total", "km_total", fmt_num),
        ("Litros", "litros_total", fmt_num),
        ("Custo/KM", "custo_por_km", fmt_brl),
        ("KM/L", "km_por_litro", lambda value: f"{fmt_num(value, 2)} km/L"),
    ]
    cards = []
    for row in rows:
        body = "".join(
            f'<div class="ranking-versus-metric"><span>{h(label)}</span><span>{h(formatter(row.get(key)))}</span></div>'
            for label, key, formatter in metrics
        )
        cards.append(f'<article class="ranking-versus-card"><h3>{h(row.get("placa") or "Sem placa")}</h3>{body}</article>')
    return f'<h2 class="ranking-versus-title">Versus entre placas</h2><section class="ranking-versus-grid">{"".join(cards)}</section>'


def selected_plotly_label(event, labels: list[str]) -> str | None:
    try:
        points = event.selection.points
    except AttributeError:
        points = (event or {}).get("selection", {}).get("points", []) if isinstance(event, dict) else []
    if not points:
        return None
    point = points[0]
    label = point.get("label") or point.get("x")
    if label:
        return str(label)
    point_index = point.get("point_index", point.get("pointIndex"))
    try:
        return labels[int(point_index)]
    except (TypeError, ValueError, IndexError):
        return None


def dominance_cities_html(placa: str, cidades: list[dict]) -> str:
    city_rows = [item for item in cidades if str(item.get("placa") or "") == placa]
    city_rows = sorted(city_rows, key=lambda item: float(item.get("peso") or 0), reverse=True)
    if not city_rows:
        return (
            '<div class="dominance-panel">'
            f'<p class="dominance-title">{h(placa)}</p>'
            '<p class="dominance-note">Essa placa aparece no peso total, mas ainda nao ha cidades dominadas por ela nos dados filtrados.</p>'
            '</div>'
        )

    rows_html = []
    for item in city_rows:
        rows_html.append(
            '<div class="dominance-city-row">'
            f'<span class="dominance-city-name">{h(item.get("cidade") or "Sem cidade")}</span>'
            f'<span class="dominance-city-metric">{h(fmt_peso(item.get("peso")))}</span>'
            f'<span class="dominance-city-metric">{h(fmt_num(item.get("participacao"), 2))}%</span>'
            f'<span class="dominance-city-metric">total {h(fmt_peso(item.get("peso_cidade")))}</span>'
            '</div>'
        )
    return (
        '<div class="dominance-panel">'
        f'<p class="dominance-title">Cidades dominadas por {h(placa)}</p>'
        f'<p class="dominance-note">{len(city_rows)} cidade(s), ordenadas pelo maior peso dominante.</p>'
        f'<div class="dominance-cities">{"".join(rows_html)}</div>'
        '</div>'
    )


def dominance_city_ranking_html(cidades: list[dict]) -> str:
    city_rows = sorted(cidades or [], key=lambda item: float(item.get("peso") or 0), reverse=True)
    if not city_rows:
        return (
            '<div class="dominance-panel">'
            '<p class="dominance-title">Ranking de peso por cidade e placa</p>'
            '<p class="dominance-note">Ainda nao ha cidades com peso para os filtros selecionados.</p>'
            '</div>'
        )

    rows_html = [
        '<div class="dominance-city-row dominance-city-row--head">'
        '<span class="dominance-city-name">Cidade</span>'
        '<span class="dominance-city-plate">Placa</span>'
        '<span class="dominance-city-metric">Peso</span>'
        '<span class="dominance-city-metric">Dominio</span>'
        '<span class="dominance-city-metric">Total cidade</span>'
        '</div>'
    ]
    for index, item in enumerate(city_rows, start=1):
        rows_html.append(
            '<div class="dominance-city-row">'
            f'<span class="dominance-city-name">{index:02d} - {h(item.get("cidade") or "Sem cidade")}</span>'
            f'<span class="dominance-city-plate">{h(item.get("placa") or "Sem placa")}</span>'
            f'<span class="dominance-city-metric">{h(fmt_peso(item.get("peso")))}</span>'
            f'<span class="dominance-city-metric">{h(fmt_num(item.get("participacao"), 2))}%</span>'
            f'<span class="dominance-city-metric">{h(fmt_peso(item.get("peso_cidade")))}</span>'
            '</div>'
        )
    return (
        '<div class="dominance-panel">'
        '<p class="dominance-title">Ranking de peso por cidade e placa</p>'
        '<p class="dominance-note">Cada cidade aparece com a placa que mais domina por peso nos filtros selecionados.</p>'
        f'<div class="dominance-cities">{"".join(rows_html)}</div>'
        '</div>'
    )


def peso_bar_chart(labels: list, values: list) -> go.Figure:
    rows = [(str(label), float(value or 0)) for label, value in zip(labels or [], values or [])]
    values_clean = [item[1] for item in rows]
    labels_clean = [item[0] for item in rows]
    text = [fmt_peso(value) for value in values_clean]
    fig = go.Figure(
        go.Bar(
            x=labels_clean,
            y=values_clean,
            marker={"color": JR_BLUE},
            text=text,
            textposition="outside",
            textfont={"size": 11, "color": JR_BLUE},
            cliponaxis=False,
            hovertemplate="<b>%{x}</b><br>%{text}<extra></extra>",
        )
    )
    max_value = max(values_clean) if values_clean else 0
    fig.update_xaxes(tickangle=-30, type="category")
    fig.update_yaxes(title="kg", range=[0, max_value * 1.16] if max_value else None, automargin=True)
    return apply_theme(fig, height=370, margin={"l": 60, "r": 48, "t": 58, "b": 70})


def dominance_city_summaries(cidades: list[dict], *, limit: int = 10) -> dict[str, str]:
    grouped: dict[str, list[dict]] = {}
    for item in cidades:
        placa = str(item.get("placa") or "")
        if placa:
            grouped.setdefault(placa, []).append(item)
    summaries = {}
    for placa, rows in grouped.items():
        ordered = sorted(rows, key=lambda item: float(item.get("peso") or 0), reverse=True)
        names = [str(item.get("cidade") or "Sem cidade") for item in ordered[:limit]]
        remaining = max(len(ordered) - len(names), 0)
        suffix = f" +{remaining}" if remaining else ""
        summaries[placa] = ", ".join(names) + suffix
    return summaries


def render_frota() -> None:
    topbar("JR DASHBOARD • Ranking de caminhões", back=False)
    seed = route_json("frota", {"ano": "Todos", "mes": ["Todos"], "categoria": "Transporte", "ordenar_por": "total"})
    params = frota_filter_controls(seed)
    data = route_json("frota", params)
    totais = data.get("totais", {}) or {}
    order_label = RANK_ORDER_OPTIONS.get(str(data.get("ordenar_por") or params.get("ordenar_por")), "Combustível")

    render_kpis(
        [
            ("Placas no ranking", fmt_num(totais.get("placas")), JR_BLUE),
            ("Gasto total", fmt_brl_big(totais.get("total")), JR_RED),
            ("Combustível", fmt_brl_big(totais.get("combustivel")), JR_BLUE),
            ("Manutenção", fmt_brl_big(totais.get("manutencao")), JR_RED),
            ("Pedágio/IPVA", fmt_brl_big(totais.get("pedagio")), "#D97706"),
            ("Peso total", fmt_peso(totais.get("peso_total")), JR_BLUE),
            ("Ordenado por", order_label, JR_BLUE),
        ]
    )

    ranking = data.get("ranking", []) or []
    if not ranking:
        st.markdown('<div class="ranking-empty">Nenhum caminhão encontrado para os filtros selecionados.</div>', unsafe_allow_html=True)
        footer("Ranking calculado com dados de combustível, manutenção e pedágio/IPVA do Neon. © JR")
        return

    selected_plates = [str(item) for item in (params.get("placa") or []) if item not in (None, "", "Todos")]
    if len(selected_plates) >= 2:
        selected_set = set(selected_plates)
        versus_rows = [row for row in ranking if row.get("placa") in selected_set]
        st.html(ranking_difference_html(versus_rows) + ranking_versus_html(versus_rows))

    include_year = params.get("ano") is None
    fallback_year = params.get("ano")
    total_labels, total_values = sorted_series(
        data.get("mensal_total", {}),
        "Mes",
        "Valor",
        include_year=include_year,
        fallback_year=fallback_year,
    )
    peso_labels, peso_values = sorted_series(
        data.get("peso_mensal", {}),
        "Mes",
        "Peso",
        include_year=include_year,
        fallback_year=fallback_year,
    )
    total_monthly_fig = (
        yearly_month_line_chart(data.get("mensal_total", {}), "Mes", "Valor", fallback_year=fallback_year)
        if include_year
        else line_chart(total_labels, total_values)
    )
    peso_monthly_fig = (
        yearly_month_bar_chart(data.get("peso_mensal", {}), "Mes", "Peso", fallback_year=fallback_year, currency=False)
        if include_year
        else peso_bar_chart(peso_labels, peso_values)
    )
    chart_grid([
        ("Gasto total por mês", total_monthly_fig),
        ("Peso por mês", peso_monthly_fig),
    ])

    dominancia = data.get("dominancia_peso", {}) or {}
    cidades_dominadas = dominancia.get("cidades", []) or []
    if cidades_dominadas:
        with st.container(border=True):
            st.html('<div class="chart-title">Ranking de peso por cidade e placa</div>')
            st.markdown(dominance_city_ranking_html(cidades_dominadas), unsafe_allow_html=True)

    st.markdown(
        """
        <div class="ranking-header">
          <span>#</span><span>Placa</span><span>Total</span><span>Combustivel</span><span>Manutencao</span><span>Pedagio/IPVA</span><span>Peso</span><span>KM</span><span>Litros</span>
        </div>
        """,
        unsafe_allow_html=True,
    )
    with st.container(key="frota_ranking_table"):
        for row in ranking:
            with st.expander(ranking_row_label(row), expanded=False):
                st.html(ranking_detail_html(row))
    footer("Ranking calculado com dados de combustível, manutenção e pedágio/IPVA do Neon. © JR")


def render_home() -> None:
    logo = logo_data_uri()
    last_update = h(last_update_label(APP_VERSION))
    st.markdown('<main class="home-wrapper">', unsafe_allow_html=True)
    st.markdown(
        f"""
        <header class="home-header">
          <div class="home-header-actions">
            <p class="home-last-update">&Uacute;ltima atualiza&ccedil;&atilde;o: {last_update}</p>
            <a class="home-admin-link" href="?page=cadastro" target="_self">Adicionar dados</a>
          </div>
          <div class="home-brand">
            <img src="{logo}" alt="JR" class="home-logo">
            <div>
              <p class="home-eyebrow">JR Ferragens &amp; Madeiras</p>
              <h1>Dashboards operacionais</h1>
              <p class="home-subtitle">Monitore combustível, manutenção, hospedagens e despesas de pedágio/seguro/IPVA em tempo real, com dados centralizados no Neon.</p>
            </div>
          </div>
          <a class="home-cta" href="#dashboards">Explorar dashboards</a>
        </header>
        """,
        unsafe_allow_html=True,
    )

    overview_all = route_json("overview")
    year_options = ["Todos", *(overview_all.get("anos_disponiveis", []) or [])]
    default_year = CURRENT_YEAR if CURRENT_YEAR in year_options else "Todos"
    ano_state = st.session_state.get("home_ano", default_year)
    ano = ano_state if ano_state in year_options else default_year
    months_seed = overview_all if ano == "Todos" else route_json("overview", {"ano": ano})
    month_options = ["Todos", *(months_seed.get("meses_disponiveis", []) or [])]
    home_mes_exists = "home_mes" in st.session_state
    home_mes_previous_key = "home_mes__previous"
    current_months = st.session_state.get("home_mes", ["Todos"])
    current_months = [item for item in current_months if item in month_options] or ["Todos"]
    if home_mes_exists and st.session_state.get("home_mes") != current_months:
        st.session_state["home_mes"] = current_months
        st.session_state[home_mes_previous_key] = current_months
    if home_mes_previous_key not in st.session_state:
        st.session_state[home_mes_previous_key] = current_months
    meses = normalize_multiselect(current_months, st.session_state.get(home_mes_previous_key, ["Todos"]))

    if meses == ["Todos"]:
        overview = months_seed
    else:
        overview = route_json(
            "overview",
            {
                "ano": None if ano == "Todos" else ano,
                "mes": [int(item) for item in meses],
            },
        )
    filter_text = ""
    if ano != "Todos":
        filter_text = f"Filtro aplicado: {ano}."
    if meses != ["Todos"]:
        month_text = ", ".join(MONTH_NAMES[int(item) - 1] for item in meses)
        filter_text = f"Filtro aplicado: {month_text}/{ano}." if ano != "Todos" else f"Filtro aplicado: {month_text}."
    suffix = filter_text or "Cálculo baseado nos dados mais recentes do banco."
    home_total_cards = [
        ("Gasto consolidado", fmt_brl(overview.get("total_geral")), suffix),
        ("Peso total", fmt_peso(overview.get("peso_total")), "Somatório dos pesos lançados nas entregas."),
        ("Gasto transporte", fmt_brl(overview.get("total_transporte")), 'Somatório das despesas marcadas como "Transporte".'),
        ("Gasto Vex", fmt_brl(overview.get("total_vex")), 'Somatório das despesas marcadas como "Vex".'),
    ]

    with st.container(key="home_total_section"):
        ready_key = "home_export_ready"
        ready_file = st.session_state.get(ready_key)
        if ready_file and ready_file.get("version") != APP_VERSION:
            st.session_state.pop(ready_key, None)

        export_cols = st.columns(2)
        export_options = [
            ("cards", "Exportar cards (PNG)", False),
            ("pagina", "Exportar página (PNG)", True),
        ]
        for col, (scope, label, include_header) in zip(export_cols, export_options):
            with col:
                if st.button(label, key=f"home_export_{scope}_png", width="stretch"):
                    image = compose_home_export_image(home_total_cards, include_header=include_header)
                    st.session_state[ready_key] = {
                        "data": image_to_png_bytes(image),
                        "file_name": export_file_name("home", scope, "png"),
                        "mime": "image/png",
                        "version": APP_VERSION,
                    }
        ready_file = st.session_state.get(ready_key)
        if ready_file:
            st.download_button(
                "Baixar arquivo gerado",
                data=ready_file["data"],
                file_name=ready_file["file_name"],
                mime=ready_file["mime"],
                key="home_download_ready",
                width="stretch",
            )

        render_home_totals(home_total_cards)
        st.markdown('<div class="home-filter-row"></div>', unsafe_allow_html=True)
        cols = st.columns([0.55, 0.55, 2.2])
        with cols[0]:
            st.selectbox("Ano", year_options, index=year_options.index(ano), key="home_ano")
        with cols[1]:
            home_mes_kwargs = {
                "format_func": month_label,
                "key": "home_mes",
                "on_change": sync_multiselect_selection,
                "args": ("home_mes",),
            }
            if not home_mes_exists:
                home_mes_kwargs["default"] = current_months
            st.multiselect("Mês", month_options, **home_mes_kwargs)
        with cols[2]:
            st.write("")
            st.write("")
            if st.button("Limpar filtro", key="home_clear", width="stretch"):
                for key in ("home_ano", "home_mes", "home_mes__previous"):
                    st.session_state.pop(key, None)
                st.rerun()

    st.markdown(
        """
        <section id="dashboards" class="home-grid">
          <a class="home-card" href="?page=frota" target="_self" aria-label="Abrir dashboard Ranking">
            <div><span class="home-chip">Ranking</span><h2>Ranking de gastos por caminhão</h2></div>
            <p class="home-card-text">Veja todos os caminhões em formato de tabela, do maior para o menor gasto, com detalhamento por placa.</p>
            <ul class="home-list"><li>Ordenação por combustível, manutenção, pedágio/IPVA ou total</li><li>Métricas individuais dentro da própria linha</li><li>Filtros por ano, mês e categoria</li></ul>
            <span class="home-link">Abrir dashboard &rarr;</span>
          </a>
          <a class="home-card" href="?page=combustivel" target="_self" aria-label="Abrir dashboard CombustÃ­vel">
            <div><span class="home-chip">Combustível</span><h2>Consumo, custo e eficiência da frota</h2></div>
            <p class="home-card-text">Filtros por mês, placa, posto e tipo de combustível com KPIs e gráficos de desempenho.</p>
            <ul class="home-list"><li>KPIs automáticos de custo, km e litros</li><li>Comparativo por posto e tipo de combustível</li><li>Histórico mensal de consumo e gastos</li></ul>
            <span class="home-link">Abrir dashboard &rarr;</span>
          </a>
          <a class="home-card" href="?page=manutencao" target="_self" aria-label="Abrir dashboard ManutenÃ§Ã£o">
            <div><span class="home-chip">Manutenção</span><h2>Gestão de oficinas e serviços</h2></div>
            <p class="home-card-text">Acompanhe gastos por placa, oficina e mês, com ticket médio atualizado.</p>
            <ul class="home-list"><li>Resumo financeiro com ticket médio</li><li>Distribuição por placa e oficina</li><li>Curva mensal de investimentos</li></ul>
            <span class="home-link">Abrir dashboard &rarr;</span>
          </a>
          <a class="home-card" href="?page=hoteis" target="_self" aria-label="Abrir dashboard HotÃ©is">
            <div><span class="home-chip">Hotéis</span><h2>Reservas e hospedagens da equipe</h2></div>
            <p class="home-card-text">Filtros por mês, cidade e hotel para entender os investimentos em hospedagem.</p>
            <ul class="home-list"><li>KPIs automáticos de valor total, reservas e médias</li><li>Ranking por cidade e hotel/pousada</li><li>Histórico mensal dos gastos com hospedagem</li></ul>
            <span class="home-link">Abrir dashboard &rarr;</span>
          </a>
          <a class="home-card" href="?page=pedagio" target="_self" aria-label="Abrir dashboard PedÃ¡gio, Seguro e IPVA">
            <div><span class="home-chip">Pedágio &amp; Seguros</span><h2>Pedágio, IPVA e seguros da frota</h2></div>
            <p class="home-card-text">Acompanhe quanto cada placa consome com pedágio, seguros e tributos, com KPIs dinâmicos.</p>
            <ul class="home-list"><li>Resumo mensal consolidado por tipo de despesa</li><li>Comparativo por placa e categoria</li><li>Filtros rápidos por mês, placa e tipo</li></ul>
            <span class="home-link">Abrir dashboard &rarr;</span>
          </a>
          <a class="home-card" href="?page=vex" target="_self" aria-label="Abrir dashboard Vex">
            <div><span class="home-chip">Vex</span><h2>Gastos exclusivos da categoria Vex</h2></div>
            <p class="home-card-text">Visão consolidada dos custos Vex, com filtros por ano e mês.</p>
            <ul class="home-list"><li>Totais Vex por área</li><li>Evolução mensal dos gastos</li><li>Resumo centralizado da categoria</li></ul>
            <span class="home-link">Abrir dashboard &rarr;</span>
          </a>
        </section>
        </main>
        """,
        unsafe_allow_html=True,
    )


def _entry_month(value: date) -> str:
    return f"{value.year}-{value.month:02d}"


def _entry_required_missing(row: dict, fields: list[str]) -> list[str]:
    missing = []
    for field in fields:
        value = row.get(field)
        if value is None:
            missing.append(field)
        elif isinstance(value, str) and not value.strip():
            missing.append(field)
    return missing


def _save_entry(
    dataset: str,
    row: dict,
    *,
    required: list[str],
    success: str,
    replace_keys: list[str] | None = None,
) -> bool:
    missing = _entry_required_missing(row, required)
    if missing:
        st.warning("Preencha os campos obrigatórios: " + ", ".join(missing))
        return False
    try:
        backend.save_dashboard_record(dataset, row, replace_keys=replace_keys)
    except Exception as exc:
        st.error("Não foi possível salvar no Neon.")
        st.exception(exc)
        return False
    clear_cached_reads()
    st.success(success)
    return True


def _registered_plate_map() -> dict[str, str]:
    try:
        df = backend.load_placas()
    except Exception:
        return {}
    if df.empty or "PLACA" not in df.columns:
        return {}
    mapping: dict[str, str] = {}
    for _, row in df.iterrows():
        placa = clean_text(row.get("PLACA")).strip().upper()
        categoria = clean_text(row.get("Categoria") or "Transporte").strip()
        if placa:
            mapping[placa] = "Vex" if categoria.lower() == "vex" else "Transporte"
    return dict(sorted(mapping.items()))


def _plate_editor_token(plate_map: dict[str, str]) -> str:
    raw = "|".join(f"{placa}:{categoria}" for placa, categoria in sorted(plate_map.items()))
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:12]


def _reset_plate_editor() -> None:
    st.session_state["cad_placas_editor_nonce"] = st.session_state.get("cad_placas_editor_nonce", 0) + 1


def _clear_stale_plate_editor_state(active_key: str) -> None:
    for key in list(st.session_state.keys()):
        if key == "cad_placas_editor" or (key.startswith("cad_placas_editor_") and key != active_key):
            del st.session_state[key]


def _registered_text_options(loader, column: str) -> list[str]:
    try:
        df = loader()
    except Exception:
        return []
    if df.empty or column not in df.columns:
        return []
    options = []
    for value in df[column].dropna().tolist():
        text = clean_text(value).strip()
        if text:
            options.append(text)
    return sorted(set(options))


def _select_or_create_text(
    label: str,
    options: list[str],
    prefix: str,
    create_label: str,
    manual_label: str,
    *,
    placeholder: str = "",
) -> str:
    clean_options = [option for option in options if option]
    if not clean_options:
        return st.text_input(label, placeholder=placeholder, key=f"{prefix}_manual_only")
    selected = st.selectbox(label, [*clean_options, create_label], key=f"{prefix}_select")
    if selected == create_label:
        return st.text_input(manual_label, placeholder=placeholder, key=f"{prefix}_manual")
    return selected


def _plate_fields(prefix: str, plate_map: dict[str, str] | None = None) -> tuple[str, str]:
    plate_map = plate_map if plate_map is not None else _registered_plate_map()
    options = ["Selecione uma placa", *plate_map.keys(), "Cadastrar nova placa"]
    selected = st.selectbox("Placa", options, key=f"{prefix}_placa_select")
    if selected == "Cadastrar nova placa":
        placa = st.text_input("Nova placa", placeholder="ABC1D23", key=f"{prefix}_placa_manual").upper()
        categoria = st.selectbox("Categoria da placa", ["Transporte", "Vex"], key=f"{prefix}_categoria_manual")
        return placa, categoria
    if selected == "Selecione uma placa":
        st.text_input("Categoria da placa", value="", disabled=True, key=f"{prefix}_categoria_empty")
        return "", "Transporte"
    categoria = plate_map.get(selected, "Transporte")
    st.text_input("Categoria da placa", value=categoria, disabled=True, key=f"{prefix}_categoria_locked")
    return selected, categoria


def _save_registered_plate(placa: str, categoria: str) -> bool:
    if not placa:
        st.warning("Selecione ou cadastre uma placa.")
        return False
    try:
        backend.save_dashboard_record("placas", {"PLACA": placa, "Categoria": categoria}, replace_keys=["PLACA"])
    except Exception as exc:
        st.error("Não foi possível salvar a placa no Neon.")
        st.exception(exc)
        return False
    clear_cached_reads()
    return True


def _save_text_registry(dataset: str, column: str, value: str, success: str) -> bool:
    text = clean_text(value).strip()
    if not text:
        st.warning("Informe um valor para cadastrar.")
        return False
    try:
        backend.save_dashboard_record(dataset, {column: text}, replace_keys=[column])
    except Exception as exc:
        st.error("Não foi possível salvar no Neon.")
        st.exception(exc)
        return False
    clear_cached_reads()
    st.success(success)
    return True


def _edit_registered_plate(old_plate: str, new_plate: str, categoria: str) -> bool:
    if not old_plate or not new_plate:
        st.warning("Selecione a placa e informe o nome correto.")
        return False
    try:
        backend.rename_plate(old_plate, new_plate, categoria)
    except Exception as exc:
        st.error("Não foi possível editar a placa no Neon.")
        st.exception(exc)
        return False
    clear_cached_reads()
    st.success("Placa atualizada nos cadastros e lançamentos.")
    return True


def _save_plate_sheet(original_map: dict[str, str], edited: pd.DataFrame) -> bool:
    if edited is None or edited.empty:
        st.warning("Informe pelo menos uma placa.")
        return False

    rows = []
    final_categories: dict[str, str] = {}
    for _, row in edited.iterrows():
        old_plate = clean_text(row.get("Placa atual")).strip().upper()
        new_plate = clean_text(row.get("Placa")).strip().upper()
        categoria = clean_text(row.get("Categoria") or "Transporte").strip()
        if not old_plate and not new_plate:
            continue
        if not new_plate:
            st.warning("Existe uma linha sem placa preenchida.")
            return False
        categoria_final = "Vex" if categoria.lower() == "vex" else "Transporte"
        final_categories[new_plate] = categoria_final
        rows.append((old_plate, new_plate))

    if not rows:
        st.warning("Informe pelo menos uma placa.")
        return False

    changed = False
    try:
        for old_plate, new_plate in rows:
            categoria = final_categories.get(new_plate, "Transporte")
            if old_plate:
                if old_plate == new_plate and original_map.get(old_plate, "Transporte") == categoria:
                    continue
                backend.rename_plate(old_plate, new_plate, categoria)
                changed = True
            else:
                backend.save_dashboard_record("placas", {"PLACA": new_plate, "Categoria": categoria}, replace_keys=["PLACA"])
                changed = True
    except Exception as exc:
        st.error("NÃ£o foi possÃ­vel salvar a tabela de placas no Neon.")
        st.exception(exc)
        return False

    _reset_plate_editor()
    if changed:
        clear_cached_reads()
        st.success("Tabela de placas salva.")
    else:
        st.info("Nenhuma alteraÃ§Ã£o para salvar.")
    st.rerun()
    return True


def _editor_empty_value(value: object) -> bool:
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except (TypeError, ValueError):
        pass
    if isinstance(value, str):
        return not value.strip()
    return False


ROW_ID_COLUMN = "Linha"


def _prepare_editor_rows(dataset: str, edited: pd.DataFrame, columns: list[str], required: list[str]) -> list[dict] | None:
    if edited is None:
        return []

    rows: list[dict] = []
    for idx, row in edited.iterrows():
        item = {column: row.get(column) for column in columns}
        if all(_editor_empty_value(value) for value in item.values()):
            continue
        if "Data" in item and not _editor_empty_value(item.get("Data")):
            item["Mes"] = _entry_month(item["Data"])
        missing = [column for column in required if _editor_empty_value(item.get(column))]
        if missing:
            st.warning(f"Linha {idx + 1}: preencha {', '.join(missing)}.")
            return None
        rows.append(item)
    return rows


def _prepare_editor_records(edited: pd.DataFrame, columns: list[str], required: list[str]) -> list[tuple[int | None, dict]] | None:
    if edited is None:
        return []

    records: list[tuple[int | None, dict]] = []
    for idx, row in edited.iterrows():
        item = {column: row.get(column) for column in columns}
        if all(_editor_empty_value(value) for value in item.values()):
            continue
        if "Data" in item and not _editor_empty_value(item.get("Data")):
            item["Mes"] = _entry_month(item["Data"])
        missing = [column for column in required if _editor_empty_value(item.get(column))]
        if missing:
            st.warning(f"Linha {idx + 1}: preencha {', '.join(missing)}.")
            return None

        row_id = row.get(ROW_ID_COLUMN)
        try:
            row_id = int(row_id) if not _editor_empty_value(row_id) else None
        except (TypeError, ValueError):
            row_id = None
        records.append((row_id, item))
    return records


def _merge_filtered_editor_rows(
    original_table: pd.DataFrame,
    edited: pd.DataFrame,
    columns: list[str],
    required: list[str],
    visible_row_ids: set[int],
) -> list[dict] | None:
    records = _prepare_editor_records(edited, columns, required)
    if records is None:
        return None

    edited_by_id = {row_id: item for row_id, item in records if row_id is not None}
    new_rows = [item for row_id, item in records if row_id is None]
    merged: list[dict] = []

    for idx, row in original_table.reset_index(drop=True).iterrows():
        row_id = idx + 1
        if row_id in visible_row_ids:
            if row_id in edited_by_id:
                merged.append(edited_by_id[row_id])
            continue
        merged.append({column: row.get(column) for column in columns})

    merged.extend(new_rows)
    return merged


def _reset_dataset_editor(key_prefix: str) -> None:
    key = f"{key_prefix}_editor_nonce"
    st.session_state[key] = st.session_state.get(key, 0) + 1


def _save_dataset_editor(
    dataset: str,
    edited: pd.DataFrame,
    columns: list[str],
    required: list[str],
    key_prefix: str,
    original_table: pd.DataFrame,
    visible_row_ids: set[int],
) -> bool:
    rows = _merge_filtered_editor_rows(original_table, edited, columns, required, visible_row_ids)
    if rows is None:
        return False
    try:
        backend.replace_dashboard_records(dataset, rows)
    except Exception as exc:
        st.error("Não foi possível salvar a tabela no Neon.")
        st.exception(exc)
        return False
    _reset_dataset_editor(key_prefix)
    clear_cached_reads()
    st.success("Tabela salva no Neon.")
    st.rerun()
    return True


def _filter_text_options(series: pd.Series) -> list[str]:
    if series is None or series.empty:
        return []
    values = series.dropna().astype("string").str.strip()
    values = values[(values != "") & (~values.str.lower().isin(["nan", "none", "nat", "<na>"]))]
    return sorted(values.unique().tolist())


def _clear_table_filter_state(key_prefix: str) -> None:
    for key in list(st.session_state.keys()):
        if key == f"{key_prefix}_search" or key.startswith(f"{key_prefix}_filter_"):
            del st.session_state[key]


def _apply_table_filters(table: pd.DataFrame, columns: list[str], key_prefix: str, filter_columns: list[str]) -> pd.DataFrame:
    filtered = table.copy()
    with st.expander("Filtros da tabela", expanded=False):
        search = st.text_input("Buscar", key=f"{key_prefix}_search", placeholder="Digite para buscar em qualquer coluna")
        if search.strip():
            needle = search.strip().lower()
            search_frame = filtered[columns].astype("string").fillna("").apply(lambda col: col.str.lower())
            mask = search_frame.apply(lambda row: any(needle in value for value in row), axis=1)
            filtered = filtered.loc[mask].copy()

        if filter_columns:
            filter_cols = st.columns(min(4, len(filter_columns)))
            for idx, column in enumerate(filter_columns):
                if column not in table.columns:
                    continue
                options = _filter_text_options(table[column])
                with filter_cols[idx % len(filter_cols)]:
                    selected = st.multiselect(column, options, key=f"{key_prefix}_filter_{column}")
                if selected:
                    values = filtered[column].astype("string").fillna("").str.strip()
                    filtered = filtered.loc[values.isin(selected)].copy()

        st.button("Limpar filtros", key=f"{key_prefix}_clear_filters", on_click=_clear_table_filter_state, args=(key_prefix,))
    return filtered


def _render_dataset_editor(
    dataset: str,
    loader,
    columns: list[str],
    required: list[str],
    key_prefix: str,
    column_config: dict,
    filter_columns: list[str],
) -> None:
    try:
        df = loader()
    except Exception as exc:
        st.error("Não foi possível carregar a tabela do Neon.")
        st.exception(exc)
        return

    table = df[[column for column in columns if column in df.columns]].copy() if not df.empty else pd.DataFrame(columns=columns)
    for column in columns:
        if column not in table.columns:
            table[column] = pd.NA
    table = table[columns]
    table = table.reset_index(drop=True)
    table.insert(0, ROW_ID_COLUMN, range(1, len(table) + 1))

    st.markdown("#### Tabela cadastrada")
    filtered_table = _apply_table_filters(table, columns, key_prefix, filter_columns)
    visible_row_ids = set(pd.to_numeric(filtered_table[ROW_ID_COLUMN], errors="coerce").dropna().astype(int).tolist())
    nonce = st.session_state.get(f"{key_prefix}_editor_nonce", 0)
    editor_config = {ROW_ID_COLUMN: st.column_config.NumberColumn("Linha")}
    editor_config.update(column_config)
    edited = st.data_editor(
        filtered_table,
        width="stretch",
        height=420,
        hide_index=True,
        num_rows="dynamic",
        disabled=[ROW_ID_COLUMN],
        column_config=editor_config,
        key=f"{key_prefix}_sheet_editor_{nonce}",
    )
    if st.button("Salvar tabela", type="primary", width="stretch", key=f"{key_prefix}_sheet_save_{nonce}"):
        _save_dataset_editor(dataset, edited, columns, required, key_prefix, table[columns], visible_row_ids)


def _date_col(label: str = "Data"):
    return st.column_config.DateColumn(label, format="DD/MM/YYYY")


def _money_col(label: str):
    return st.column_config.NumberColumn(label, min_value=0.0, step=10.0, format="R$ %.2f")


def _number_col(label: str):
    return st.column_config.NumberColumn(label, min_value=0.0, step=1.0)


def _normalize_sheet_header(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    return "".join(ch for ch in text.upper() if ch.isalnum())


def _parse_brl_number(value: object) -> float | None:
    if _editor_empty_value(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value)
    text = text.replace("R$", "").replace("\u00a0", " ").strip()
    text = "".join(ch for ch in text if ch.isdigit() or ch in ",.-")
    if not text:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _parse_sheet_date(value: object) -> tuple[date, str] | None:
    if _editor_empty_value(value):
        return None
    parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")
    if pd.notna(parsed):
        parsed_date = date(int(parsed.year), int(parsed.month), int(parsed.day))
        return parsed_date, f"{parsed_date.year}-{parsed_date.month:02d}"

    return None


def _parse_sheet_month(value: object) -> tuple[str, date] | None:
    parsed_date = _parse_sheet_date(value)
    if parsed_date:
        data, mes = parsed_date
        return mes, date(data.year, data.month, 1)
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii").lower().strip()
    month = None
    month_names = {
        "jan": 1,
        "janeiro": 1,
        "fev": 2,
        "fevereiro": 2,
        "mar": 3,
        "marco": 3,
        "abr": 4,
        "abril": 4,
        "mai": 5,
        "maio": 5,
        "jun": 6,
        "junho": 6,
        "jul": 7,
        "julho": 7,
        "ago": 8,
        "agosto": 8,
        "set": 9,
        "setembro": 9,
        "out": 10,
        "outubro": 10,
        "nov": 11,
        "novembro": 11,
        "dez": 12,
        "dezembro": 12,
    }
    for label, number in month_names.items():
        if label in text:
            month = number
            break

    digits = [int(part) for part in "".join(ch if ch.isdigit() else " " for ch in text).split()]
    year = None
    if month is not None and digits:
        year = digits[-1]
    elif len(digits) >= 2:
        if digits[0] > 31:
            year, month = digits[0], digits[1]
        else:
            month, year = digits[0], digits[-1]
    if year is not None and year < 100:
        year += 2000
    if not year or not month or not (1 <= month <= 12):
        return None
    return f"{year}-{month:02d}", date(year, month, 1)


PEDAGIO_SHEET_ALIASES = {
    "PLACA": ["PLACA", "PLACAS"],
    "TIPO": ["TIPO"],
    "CUSTO": ["CUSTO", "VALOR"],
    "MES": ["MES", "MS"],
}

HOTEIS_SHEET_ALIASES = {
    "DATA": ["DATA", "DT"],
    "CIDADE": ["CIDADE"],
    "VALOR": ["VALOR", "CUSTO"],
    "HOTEL": ["HOTELPOUSADA", "HOTEL", "POUSADA"],
}

PESO_SHEET_ALIASES = {
    "DATA": ["DATA", "DT"],
    "CIDADE": ["CIDADE"],
    "PESO": ["PESO"],
    "PLACA": ["PLACA", "PLACAS"],
}


def _read_uploaded_sheet(uploaded_file, aliases: dict[str, list[str]]) -> pd.DataFrame:
    name = clean_text(getattr(uploaded_file, "name", "")).lower()
    if name.endswith(".csv"):
        raw = pd.read_csv(uploaded_file, sep=None, engine="python", header=None)
        return _detect_sheet_header(raw, aliases)
    raw = _read_uploaded_excel_rows(uploaded_file, aliases)
    return _detect_sheet_header(raw, aliases)


def _read_uploaded_excel_rows(uploaded_file, aliases: dict[str, list[str]]) -> pd.DataFrame:
    from openpyxl import load_workbook

    try:
        uploaded_file.seek(0)
    except Exception:
        pass

    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    worksheet = workbook.active
    rows: list[list[object]] = []
    header_found = False
    blank_after_header = 0
    for row in worksheet.iter_rows(values_only=True):
        values = list(row)
        has_value = any(not _editor_empty_value(value) for value in values)
        rows.append(values)

        if not header_found:
            normalized = [_normalize_sheet_header(value) for value in values]
            header_found = all(any(alias in normalized for alias in field_aliases) for field_aliases in aliases.values())
            continue

        if has_value:
            blank_after_header = 0
        else:
            blank_after_header += 1
            if blank_after_header >= 50:
                break

    workbook.close()
    return pd.DataFrame(rows)


def _detect_sheet_header(raw: pd.DataFrame, aliases: dict[str, list[str]]) -> pd.DataFrame:
    for idx, row in raw.iterrows():
        normalized = [_normalize_sheet_header(value) for value in row.tolist()]
        if all(any(alias in normalized for alias in field_aliases) for field_aliases in aliases.values()):
            df = raw.iloc[idx + 1 :].copy()
            df.columns = [clean_text(value).strip() or f"Coluna {pos + 1}" for pos, value in enumerate(row.tolist())]
            return df.dropna(how="all").reset_index(drop=True)

    raw = raw.copy()
    raw.columns = [clean_text(value).strip() or f"Coluna {idx + 1}" for idx, value in enumerate(raw.iloc[0].tolist())]
    return raw.iloc[1:].dropna(how="all").reset_index(drop=True)


def _pedagio_rows_from_sheet(df: pd.DataFrame, plate_map: dict[str, str]) -> tuple[list[dict], list[str]]:
    header_map = {_normalize_sheet_header(column): column for column in df.columns}
    aliases = {
        "PLACA": ["PLACA", "PLACAS"],
        "TIPO": ["TIPO"],
        "CUSTO": ["CUSTO", "VALOR"],
        "MES": ["MES", "MS"],
    }
    resolved = {field: next((header_map[key] for key in keys if key in header_map), None) for field, keys in aliases.items()}
    labels = {"PLACA": "PLACA", "TIPO": "TIPO", "CUSTO": "Custo", "MES": "MES"}
    missing = [labels[field] for field, column in resolved.items() if column is None]
    if missing:
        return [], [f"Colunas faltando: {', '.join(missing)}."]

    rows: list[dict] = []
    errors: list[str] = []
    for idx, row in df.iterrows():
        placa = clean_text(row.get(resolved["PLACA"])).strip().upper()
        tipo = clean_text(row.get(resolved["TIPO"])).strip()
        custo = _parse_brl_number(row.get(resolved["CUSTO"]))
        mes_info = _parse_sheet_month(row.get(resolved["MES"]))

        if not placa and not tipo and custo is None and mes_info is None:
            continue
        missing_row = []
        if not placa:
            missing_row.append("PLACA")
        if not tipo:
            missing_row.append("TIPO")
        if custo is None:
            missing_row.append("Custo")
        if mes_info is None:
            missing_row.append("MES")
        if missing_row:
            errors.append(f"Linha {idx + 2}: preencher {', '.join(missing_row)}.")
            continue

        mes, data = mes_info
        rows.append(
            {
                "Data": data,
                "Mes": mes,
                "PLACA": placa,
                "Tipo": tipo,
                "Custo": custo,
                "Categoria": plate_map.get(placa, "Transporte"),
            }
        )
    return rows, errors


def _sheet_text(row: pd.Series, column: str | None, *, upper: bool = False) -> str:
    if not column:
        return ""
    value = row.get(column)
    if pd.isna(value):
        return ""
    text = clean_text(value).strip()
    return text.upper() if upper else text


def _hoteis_rows_from_sheet(df: pd.DataFrame) -> tuple[list[dict], list[str]]:
    header_map = {_normalize_sheet_header(column): column for column in df.columns}
    aliases = {
        "DATA": ["DATA", "DT"],
        "MOTORISTA": ["MOTORISTA"],
        "AJUDANTE": ["AJUDANTE"],
        "CIDADE": ["CIDADE"],
        "VALOR": ["VALOR", "CUSTO"],
        "HOTEL": ["HOTELPOUSADA", "HOTEL", "POUSADA"],
        "DIAS": ["DIAS", "DIA"],
        "TIPO": ["TIPO"],
    }
    resolved = {field: next((header_map[key] for key in keys if key in header_map), None) for field, keys in aliases.items()}
    labels = {"DATA": "DATA", "CIDADE": "CIDADE", "VALOR": "VALOR", "HOTEL": "HOTEL/POUSADA"}
    missing = [labels[field] for field in labels if resolved.get(field) is None]
    if missing:
        return [], [f"Colunas faltando: {', '.join(missing)}."]

    rows: list[dict] = []
    errors: list[str] = []
    for idx, row in df.iterrows():
        data_info = _parse_sheet_date(row.get(resolved["DATA"]))
        motorista = _sheet_text(row, resolved.get("MOTORISTA"), upper=True)
        ajudante = _sheet_text(row, resolved.get("AJUDANTE"), upper=True)
        cidade = _sheet_text(row, resolved.get("CIDADE"), upper=True)
        hotel = _sheet_text(row, resolved.get("HOTEL"), upper=True)
        valor = _parse_brl_number(row.get(resolved["VALOR"]))
        dias = _parse_brl_number(row.get(resolved.get("DIAS"))) if resolved.get("DIAS") else 1.0
        tipo = _sheet_text(row, resolved.get("TIPO"), upper=True) or "Hospedagem"

        if not any([data_info, motorista, ajudante, cidade, hotel, valor is not None]):
            continue
        missing_row = []
        if data_info is None:
            missing_row.append("DATA")
        if not cidade:
            missing_row.append("CIDADE")
        if valor is None:
            missing_row.append("VALOR")
        if not hotel:
            missing_row.append("HOTEL/POUSADA")
        if missing_row:
            errors.append(f"Linha {idx + 2}: preencher {', '.join(missing_row)}.")
            continue

        data, mes = data_info
        rows.append(
            {
                "Data": data,
                "Mes": mes,
                "Valor": valor,
                "Dias": dias if dias is not None else 1.0,
                "Motorista": motorista,
                "Ajudante": ajudante,
                "Cidade": cidade,
                "Hotel": hotel,
                "Tipo": tipo,
                "Categoria": "Transporte",
            }
        )
    return rows, errors


def _peso_rows_from_sheet(df: pd.DataFrame, plate_map: dict[str, str]) -> tuple[list[dict], list[str]]:
    header_map = {_normalize_sheet_header(column): column for column in df.columns}
    aliases = {
        "DATA": ["DATA", "DT"],
        "CIDADE": ["CIDADE"],
        "PESO": ["PESO"],
        "VALOR": ["VALOR", "CUSTO"],
        "PLACA": ["PLACA", "PLACAS"],
    }
    resolved = {field: next((header_map[key] for key in keys if key in header_map), None) for field, keys in aliases.items()}
    labels = {"DATA": "DATA", "PESO": "PESO", "PLACA": "PLACA"}
    missing = [labels[field] for field in labels if resolved.get(field) is None]
    if missing:
        return [], [f"Colunas faltando: {', '.join(missing)}."]

    rows: list[dict] = []
    errors: list[str] = []
    for idx, row in df.iterrows():
        data_info = _parse_sheet_date(row.get(resolved["DATA"]))
        cidade = _sheet_text(row, resolved.get("CIDADE"), upper=True)
        peso = _parse_brl_number(row.get(resolved["PESO"]))
        valor = _parse_brl_number(row.get(resolved.get("VALOR"))) if resolved.get("VALOR") else 0.0
        placa_raw = _sheet_text(row, resolved.get("PLACA"), upper=True)
        placa_normalizada = backend._normalize_plate_value(placa_raw)
        placa = "" if pd.isna(placa_normalizada) else str(placa_normalizada)

        if not any([data_info, cidade, peso is not None, valor not in (None, 0.0), placa_raw]):
            continue
        missing_row = []
        if data_info is None:
            missing_row.append("DATA")
        if peso is None:
            missing_row.append("PESO")
        if not placa:
            missing_row.append("PLACA")
        if missing_row:
            errors.append(f"Linha {idx + 2}: preencher {', '.join(missing_row)}.")
            continue

        data, mes = data_info
        rows.append(
            {
                "Data": data,
                "Mes": mes,
                "Cidade": cidade,
                "Peso": peso,
                "Valor": valor if valor is not None else 0.0,
                "PLACA": placa,
                "Categoria": plate_map.get(placa, "Transporte"),
            }
        )
    return rows, errors


def _clear_hoteis_last_import() -> None:
    st.session_state.pop("cad_hotel_last_import_rows", None)
    st.session_state.pop("cad_hotel_last_import_count", None)


def _undo_hoteis_last_import() -> None:
    rows = st.session_state.get("cad_hotel_last_import_rows") or []
    if not rows:
        st.warning("Nao ha importacao recente para apagar.")
        return
    try:
        deleted = backend.delete_matching_dashboard_records("hoteis", rows)
    except Exception as exc:
        st.error("Nao foi possivel apagar a ultima importacao.")
        st.exception(exc)
        return
    _clear_hoteis_last_import()
    _reset_dataset_editor("cad_hotel_table")
    clear_cached_reads()
    st.success(f"{deleted} hospedagem(ns) apagada(s).")
    st.rerun()


def _render_hoteis_sheet_import() -> None:
    last_rows = st.session_state.get("cad_hotel_last_import_rows") or []
    if last_rows:
        last_count = st.session_state.get("cad_hotel_last_import_count", len(last_rows))
        st.warning(f"Ultima importacao por planilha: {last_count} hospedagem(ns).")
        undo_col, clear_col = st.columns([1, 1])
        with undo_col:
            if st.button("Apagar ultima importacao", type="primary", width="stretch", key="cad_hotel_undo_import"):
                _undo_hoteis_last_import()
        with clear_col:
            if st.button("Manter importacao", width="stretch", key="cad_hotel_keep_import"):
                _clear_hoteis_last_import()
                st.rerun()

    with st.expander("Adicionar hoteis por planilha", expanded=False):
        uploaded = st.file_uploader("Enviar planilha", type=["xlsx", "csv"], key="cad_hotel_upload")
        if uploaded is None:
            return

        try:
            raw_df = _read_uploaded_sheet(uploaded, HOTEIS_SHEET_ALIASES)
        except Exception as exc:
            st.error("Nao foi possivel ler a planilha. Envie um arquivo .xlsx ou .csv.")
            st.exception(exc)
            return

        rows, errors = _hoteis_rows_from_sheet(raw_df)
        if errors:
            st.warning("Revise a planilha antes de importar.")
            for error in errors[:8]:
                st.write(error)
            if len(errors) > 8:
                st.write(f"...mais {len(errors) - 8} erro(s).")
            return
        if not rows:
            st.warning("Nenhuma linha valida encontrada na planilha.")
            return

        preview = pd.DataFrame(rows)
        st.dataframe(preview[["Data", "Mes", "Cidade", "Hotel", "Motorista", "Ajudante", "Valor"]], width="stretch", hide_index=True)
        if st.button(f"Importar {len(rows)} hospedagem(ns)", type="primary", width="stretch", key="cad_hotel_import_sheet"):
            imported_rows: list[dict] = []
            try:
                imported_rows = _append_records_in_batches("hoteis", rows, batch_size=100)
            except Exception as exc:
                if imported_rows:
                    st.session_state["cad_hotel_last_import_rows"] = imported_rows
                    st.session_state["cad_hotel_last_import_count"] = len(imported_rows)
                    st.error(f"O envio parou depois de {len(imported_rows)} hospedagem(ns). Voce pode apagar essa importacao parcial pelo botao acima.")
                else:
                    st.error("Nao foi possivel importar a planilha para o Neon.")
                st.exception(exc)
                return
            st.session_state["cad_hotel_last_import_rows"] = imported_rows
            st.session_state["cad_hotel_last_import_count"] = len(imported_rows)
            _reset_dataset_editor("cad_hotel_table")
            st.success(f"{len(imported_rows)} hospedagem(ns) importada(s).")
            st.rerun()


def _clear_peso_last_import() -> None:
    st.session_state.pop("cad_peso_last_import_rows", None)
    st.session_state.pop("cad_peso_last_import_count", None)


def _undo_peso_last_import() -> None:
    rows = st.session_state.get("cad_peso_last_import_rows") or []
    if not rows:
        st.warning("Nao ha importacao recente para apagar.")
        return
    try:
        deleted = backend.delete_matching_dashboard_records("peso", rows)
    except Exception as exc:
        st.error("Nao foi possivel apagar a ultima importacao.")
        st.exception(exc)
        return
    _clear_peso_last_import()
    _reset_dataset_editor("cad_peso_table")
    clear_cached_reads()
    st.success(f"{deleted} entrega(s) apagada(s).")
    st.rerun()


def _render_peso_sheet_import(plate_map: dict[str, str]) -> None:
    last_rows = st.session_state.get("cad_peso_last_import_rows") or []
    if last_rows:
        last_count = st.session_state.get("cad_peso_last_import_count", len(last_rows))
        st.warning(f"Ultima importacao por planilha: {last_count} entrega(s).")
        undo_col, clear_col = st.columns([1, 1])
        with undo_col:
            if st.button("Apagar ultima importacao", type="primary", width="stretch", key="cad_peso_undo_import"):
                _undo_peso_last_import()
        with clear_col:
            if st.button("Manter importacao", width="stretch", key="cad_peso_keep_import"):
                _clear_peso_last_import()
                st.rerun()

    with st.expander("Adicionar peso por planilha", expanded=False):
        uploaded = st.file_uploader("Enviar planilha", type=["xlsx", "csv"], key="cad_peso_upload")
        if uploaded is None:
            return

        try:
            raw_df = _read_uploaded_sheet(uploaded, PESO_SHEET_ALIASES)
        except Exception as exc:
            st.error("Nao foi possivel ler a planilha. Envie um arquivo .xlsx ou .csv.")
            st.exception(exc)
            return

        rows, errors = _peso_rows_from_sheet(raw_df, plate_map)
        if errors:
            st.warning("Revise a planilha antes de importar.")
            for error in errors[:8]:
                st.write(error)
            if len(errors) > 8:
                st.write(f"...mais {len(errors) - 8} erro(s).")
            return
        if not rows:
            st.warning("Nenhuma linha valida encontrada na planilha.")
            return

        preview = pd.DataFrame(rows)
        st.dataframe(preview[["Data", "Mes", "Cidade", "PLACA", "Peso", "Valor", "Categoria"]], width="stretch", hide_index=True)
        if st.button(f"Importar {len(rows)} entrega(s)", type="primary", width="stretch", key="cad_peso_import_sheet"):
            imported_rows: list[dict] = []
            try:
                imported_rows = _append_records_in_batches("peso", rows, batch_size=100)
            except Exception as exc:
                if imported_rows:
                    st.session_state["cad_peso_last_import_rows"] = imported_rows
                    st.session_state["cad_peso_last_import_count"] = len(imported_rows)
                    st.error(f"O envio parou depois de {len(imported_rows)} entrega(s). Voce pode apagar essa importacao parcial pelo botao acima.")
                else:
                    st.error("Nao foi possivel importar a planilha para o Neon.")
                st.exception(exc)
                return
            st.session_state["cad_peso_last_import_rows"] = imported_rows
            st.session_state["cad_peso_last_import_count"] = len(imported_rows)
            _reset_dataset_editor("cad_peso_table")
            st.success(f"{len(imported_rows)} entrega(s) importada(s).")
            st.rerun()


def _render_peso_month_reset() -> None:
    with st.expander("Zerar peso por mes", expanded=False):
        st.warning("Essa acao apaga todos os lancamentos de peso do mes escolhido.")

        anos = [CURRENT_YEAR]
        df_preview = pd.DataFrame()
        try:
            df_preview = backend.load_peso()
            if not df_preview.empty and "Mes" in df_preview.columns:
                parsed_years = pd.to_datetime(df_preview["Mes"], errors="coerce").dt.year.dropna().astype(int).unique().tolist()
                anos = sorted(set(anos) | set(parsed_years), reverse=True)
        except Exception as exc:
            st.info("Nao foi possivel carregar a tabela de peso para prever os totais, mas a exclusao ainda pode ser tentada.")
            st.caption(str(exc))

        c1, c2 = st.columns(2)
        with c1:
            ano = st.selectbox("Ano", anos, index=0, key="cad_peso_reset_ano")
        with c2:
            mes_num = st.selectbox(
                "Mes",
                list(range(1, 13)),
                index=max(date.today().month - 1, 0),
                format_func=lambda value: f"{value:02d}",
                key="cad_peso_reset_mes",
            )

        mes_key = f"{int(ano)}-{int(mes_num):02d}"
        if not df_preview.empty and "Mes" in df_preview.columns:
            mask = df_preview["Mes"].astype("string").str.strip().eq(mes_key)
            linhas = int(mask.sum())
            peso_total = float(pd.to_numeric(df_preview.loc[mask, "Peso"], errors="coerce").sum()) if "Peso" in df_preview.columns else 0.0
            st.caption(f"{mes_key}: {linhas} lancamento(s), {fmt_peso(peso_total)}.")
        else:
            st.caption(f"Mes selecionado: {mes_key}.")

        confirmar = st.checkbox(f"Confirmo que quero zerar o peso de {mes_key}.", key="cad_peso_reset_confirm")
        if st.button("Zerar mes de peso", type="primary", width="stretch", disabled=not confirmar, key="cad_peso_reset_button"):
            try:
                delete_month = getattr(backend, "delete_dashboard_month", None)
                if callable(delete_month):
                    deleted = delete_month("peso", mes_key)
                else:
                    if df_preview.empty:
                        raise RuntimeError("A tabela de peso nao foi carregada para apagar pelo modo compativel.")
                    fallback = df_preview.copy()
                    month_mask = fallback["Mes"].astype("string").str.strip().eq(mes_key) if "Mes" in fallback.columns else pd.Series(False, index=fallback.index)
                    if "Data" in fallback.columns:
                        data_period = pd.to_datetime(fallback["Data"], errors="coerce").dt.to_period("M").astype("string")
                        month_mask = month_mask | data_period.eq(mes_key)
                    rows = fallback.loc[month_mask].to_dict("records")
                    deleted = backend.delete_matching_dashboard_records("peso", rows) if rows else 0
            except Exception as exc:
                st.error("Nao foi possivel zerar esse mes no Neon.")
                st.exception(exc)
                return
            _reset_dataset_editor("cad_peso_table")
            clear_cached_reads()
            st.success(f"{deleted} lancamento(s) de peso apagado(s) em {mes_key}.")
            st.rerun()


def _clear_pedagio_last_import() -> None:
    st.session_state.pop("cad_ped_last_import_rows", None)
    st.session_state.pop("cad_ped_last_import_count", None)


def _undo_pedagio_last_import() -> None:
    rows = st.session_state.get("cad_ped_last_import_rows") or []
    if not rows:
        st.warning("Nao ha importacao recente para apagar.")
        return
    try:
        deleted = backend.delete_matching_dashboard_records("pedagio", rows)
    except Exception as exc:
        st.error("Nao foi possivel apagar a ultima importacao.")
        st.exception(exc)
        return
    _clear_pedagio_last_import()
    _reset_dataset_editor("cad_ped_table")
    clear_cached_reads()
    st.success(f"{deleted} lancamento(s) apagado(s).")
    st.rerun()


def _append_records_in_batches(dataset: str, rows: list[dict], *, batch_size: int = 100) -> list[dict]:
    imported: list[dict] = []
    total = len(rows)
    progress = st.progress(0, text="Preparando envio...")
    status = st.empty()

    for start in range(0, total, batch_size):
        batch = rows[start : start + batch_size]
        batch_number = (start // batch_size) + 1
        batch_total = (total + batch_size - 1) // batch_size
        status.info(f"Enviando lote {batch_number}/{batch_total} ({start + 1}-{min(start + len(batch), total)} de {total})...")
        backend.append_dashboard_records(dataset, batch, update_plate_registry=False)
        imported.extend(batch)
        progress.progress(min(len(imported) / total, 1.0), text=f"{len(imported)} de {total} lancamentos enviados")

    status.empty()
    progress.empty()
    clear_cached_reads()
    return imported


def _render_pedagio_sheet_import(plate_map: dict[str, str]) -> None:
    last_rows = st.session_state.get("cad_ped_last_import_rows") or []
    if last_rows:
        last_count = st.session_state.get("cad_ped_last_import_count", len(last_rows))
        st.warning(f"Ultima importacao por planilha: {last_count} lancamento(s).")
        undo_col, clear_col = st.columns([1, 1])
        with undo_col:
            if st.button("Apagar ultima importacao", type="primary", width="stretch", key="cad_ped_undo_import"):
                _undo_pedagio_last_import()
        with clear_col:
            if st.button("Manter importacao", width="stretch", key="cad_ped_keep_import"):
                _clear_pedagio_last_import()
                st.rerun()

    with st.expander("Adicionar pedagio/IPVA por planilha", expanded=False):
        uploaded = st.file_uploader("Enviar planilha", type=["xlsx", "csv"], key="cad_ped_upload")
        if uploaded is None:
            return

        try:
            raw_df = _read_uploaded_sheet(uploaded, PEDAGIO_SHEET_ALIASES)
        except Exception as exc:
            st.error("Nao foi possivel ler a planilha. Envie um arquivo .xlsx ou .csv.")
            st.exception(exc)
            return

        rows, errors = _pedagio_rows_from_sheet(raw_df, plate_map)
        if errors:
            st.warning("Revise a planilha antes de importar.")
            for error in errors[:8]:
                st.write(error)
            if len(errors) > 8:
                st.write(f"...mais {len(errors) - 8} erro(s).")
            return
        if not rows:
            st.warning("Nenhuma linha válida encontrada na planilha.")
            return

        preview = pd.DataFrame(rows)
        st.dataframe(preview[["Mes", "PLACA", "Categoria", "Tipo", "Custo"]], width="stretch", hide_index=True)
        if st.button(f"Importar {len(rows)} lancamentos", type="primary", width="stretch", key="cad_ped_import_sheet"):
            imported_rows: list[dict] = []
            try:
                imported_rows = _append_records_in_batches("pedagio", rows, batch_size=100)
            except Exception as exc:
                if imported_rows:
                    st.session_state["cad_ped_last_import_rows"] = imported_rows
                    st.session_state["cad_ped_last_import_count"] = len(imported_rows)
                    st.error(f"O envio parou depois de {len(imported_rows)} lancamento(s). Voce pode apagar essa importacao parcial pelo botao acima.")
                else:
                    st.error("Nao foi possivel importar a planilha para o Neon.")
                st.exception(exc)
                return
            st.session_state["cad_ped_last_import_rows"] = imported_rows
            st.session_state["cad_ped_last_import_count"] = len(imported_rows)
            _reset_dataset_editor("cad_ped_table")
            st.success(f"{len(imported_rows)} lancamentos importados.")
            st.rerun()


def render_cadastro() -> None:
    topbar("JR DASHBOARD • Adicionar dados", back=True)
    with st.container(key="cadastro_shell"):
        tabs = st.tabs(["Placas", "Combustível", "KM mensal", "Manutenção", "Hotéis", "Peso", "Pedágio/IPVA"])

        with tabs[0]:
            with st.form("form_placas", clear_on_submit=True):
                c1, c2, c3 = st.columns([1.2, 1.0, 1.0])
                with c1:
                    placa = st.text_input("Placa", placeholder="ABC1D23", key="cad_placa_nome").upper()
                with c2:
                    categoria = st.selectbox("Categoria", ["Transporte", "Vex"], key="cad_placa_categoria")
                with c3:
                    st.write("")
                    submitted = st.form_submit_button("Cadastrar placa", type="primary", width="stretch")
                if submitted:
                    _save_entry(
                        "placas",
                        {"PLACA": placa, "Categoria": categoria},
                        required=["PLACA", "Categoria"],
                        replace_keys=["PLACA"],
                        success="Placa cadastrada/atualizada.",
                    )

            plate_map = _registered_plate_map()
            if plate_map:
                editor_token = _plate_editor_token(plate_map)
                editor_nonce = st.session_state.get("cad_placas_editor_nonce", 0)
                editor_key = f"cad_placas_editor_{editor_token}_{editor_nonce}"
                _clear_stale_plate_editor_state(editor_key)
                table = pd.DataFrame(
                    [
                        {"Placa atual": placa, "Placa": placa, "Categoria": categoria}
                        for placa, categoria in plate_map.items()
                    ]
                )
                edited_plates = st.data_editor(
                    table,
                    width="stretch",
                    hide_index=True,
                    num_rows="dynamic",
                    disabled=["Placa atual"],
                    column_config={
                        "Placa atual": st.column_config.TextColumn("Placa atual"),
                        "Placa": st.column_config.TextColumn("Placa"),
                        "Categoria": st.column_config.SelectboxColumn(
                            "Categoria",
                            options=["Transporte", "Vex"],
                            required=True,
                        ),
                    },
                    key=editor_key,
                )
                save_key = f"cad_placas_sheet_save_{editor_token}_{editor_nonce}"
                if st.button("Salvar tabela de placas", type="primary", width="stretch", key=save_key):
                    _save_plate_sheet(plate_map, edited_plates)
            else:
                st.info("Cadastre a primeira placa para liberar a edição.")

        with tabs[1]:
            combustivel_options = _registered_text_options(backend.load_combustiveis, "Combustivel")
            posto_options = _registered_text_options(backend.load_postos, "POSTOS")

            r1, r2 = st.columns(2)
            with r1:
                with st.form("form_cadastrar_combustivel_tipo", clear_on_submit=True):
                    novo_combustivel = st.text_input("Cadastrar combustível", placeholder="Diesel S10", key="cad_tipo_combustivel")
                    submit_combustivel = st.form_submit_button("Cadastrar combustível", type="primary", width="stretch")
                    if submit_combustivel:
                        if _save_text_registry("combustiveis", "Combustivel", novo_combustivel, "Combustível cadastrado."):
                            combustivel_options = _registered_text_options(backend.load_combustiveis, "Combustivel")
            with r2:
                with st.form("form_cadastrar_posto", clear_on_submit=True):
                    novo_posto = st.text_input("Cadastrar posto", placeholder="Posto JR", key="cad_posto_nome")
                    submit_posto = st.form_submit_button("Cadastrar posto", type="primary", width="stretch")
                    if submit_posto:
                        if _save_text_registry("postos", "POSTOS", novo_posto, "Posto cadastrado."):
                            posto_options = _registered_text_options(backend.load_postos, "POSTOS")

            if combustivel_options or posto_options:
                with st.expander("Ver combustíveis e postos cadastrados"):
                    l1, l2 = st.columns(2)
                    with l1:
                        st.dataframe(pd.DataFrame({"Combustível": combustivel_options}), width="stretch", hide_index=True)
                    with l2:
                        st.dataframe(pd.DataFrame({"Posto": posto_options}), width="stretch", hide_index=True)

            with st.form("form_combustivel", clear_on_submit=True):
                c1, c2, c3 = st.columns(3)
                with c1:
                    data = st.date_input("Data", value=date.today(), key="cad_comb_data")
                    placa, categoria = _plate_fields("cad_comb", plate_map)
                with c2:
                    combustivel = _select_or_create_text(
                        "Combustível",
                        combustivel_options,
                        "cad_comb_combustivel",
                        "Cadastrar novo combustível",
                        "Novo combustível",
                        placeholder="Diesel S10",
                    )
                    posto = _select_or_create_text(
                        "Posto",
                        posto_options,
                        "cad_comb_posto",
                        "Cadastrar novo posto",
                        "Novo posto",
                        placeholder="Posto JR",
                    )
                    km = st.number_input("KM rodados", min_value=0.0, step=1.0, key="cad_comb_km")
                with c3:
                    litros = st.number_input("Litros", min_value=0.0, step=1.0, key="cad_comb_litros")
                    custo = st.number_input("Custo total", min_value=0.0, step=10.0, format="%.2f", key="cad_comb_custo")
                    submitted = st.form_submit_button("Salvar combustível", type="primary", width="stretch")
                if submitted:
                    _save_entry(
                        "combustivel",
                        {
                            "Data": data,
                            "Mes": _entry_month(data),
                            "Km Rodados": km,
                            "Litros": litros,
                            "Custo": custo,
                            "Combustivel": combustivel,
                            "POSTOS": posto,
                            "PLACA": placa,
                            "Categoria": categoria,
                        },
                        required=["Data", "PLACA", "Combustivel", "POSTOS"],
                        success="Lançamento de combustível salvo.",
                    )

            _render_dataset_editor(
                "combustivel",
                backend.load_combustivel,
                ["Data", "Mes", "PLACA", "Categoria", "Combustivel", "POSTOS", "Km Rodados", "Litros", "Custo"],
                ["Data", "PLACA", "Combustivel", "POSTOS"],
                "cad_comb_table",
                {
                    "Data": _date_col(),
                    "Mes": st.column_config.TextColumn("Mes"),
                    "PLACA": st.column_config.TextColumn("Placa"),
                    "Categoria": st.column_config.SelectboxColumn("Categoria", options=["Transporte", "Vex"], required=True),
                    "Combustivel": st.column_config.TextColumn("Combustivel"),
                    "POSTOS": st.column_config.TextColumn("Posto"),
                    "Km Rodados": _number_col("KM rodados"),
                    "Litros": _number_col("Litros"),
                    "Custo": _money_col("Custo"),
                },
                ["Mes", "PLACA", "Categoria", "Combustivel", "POSTOS"],
            )

        with tabs[2]:
            with st.form("form_km_mensal", clear_on_submit=True):
                c1, c2, c3 = st.columns(3)
                with c1:
                    ano = st.number_input("Ano", min_value=2020, max_value=2100, value=CURRENT_YEAR, step=1, key="cad_km_ano")
                with c2:
                    mes = st.selectbox("Mês", list(range(1, 13)), index=date.today().month - 1, format_func=month_label, key="cad_km_mes")
                    placa, categoria = _plate_fields("cad_km", plate_map)
                with c3:
                    km = st.number_input("KM do mês", min_value=0.0, step=1.0, key="cad_km_total")
                    substituir = st.checkbox("Substituir se já existir", value=True, key="cad_km_replace")
                    submitted = st.form_submit_button("Salvar KM mensal", type="primary", width="stretch")
                if submitted:
                    if _save_registered_plate(placa, categoria):
                        _save_entry(
                            "combustivel_km",
                            {"Mes": f"{int(ano)}-{int(mes):02d}", "PLACA": placa, "Km Rodados": km},
                            required=["Mes", "PLACA"],
                            replace_keys=["Mes", "PLACA"] if substituir else None,
                            success="KM mensal salvo.",
                        )

            _render_dataset_editor(
                "combustivel_km",
                backend.load_combustivel_km,
                ["Mes", "PLACA", "Km Rodados"],
                ["Mes", "PLACA"],
                "cad_km_table",
                {
                    "Mes": st.column_config.TextColumn("Mes"),
                    "PLACA": st.column_config.TextColumn("Placa"),
                    "Km Rodados": _number_col("KM rodados"),
                },
                ["Mes", "PLACA"],
            )

        with tabs[3]:
            with st.form("form_manutencao", clear_on_submit=True):
                c1, c2, c3 = st.columns(3)
                with c1:
                    data = st.date_input("Data", value=date.today(), key="cad_manu_data")
                    placa, categoria = _plate_fields("cad_manu", plate_map)
                with c2:
                    oficina = st.text_input("Oficina", key="cad_manu_oficina")
                with c3:
                    custo = st.number_input("Custo", min_value=0.0, step=10.0, format="%.2f", key="cad_manu_custo")
                    submitted = st.form_submit_button("Salvar manutenção", type="primary", width="stretch")
                if submitted:
                    _save_entry(
                        "manutencao",
                        {
                            "Data": data,
                            "Mes": _entry_month(data),
                            "Custo": custo,
                            "PLACA": placa,
                            "OFICINA": oficina,
                            "Categoria": categoria,
                        },
                        required=["Data", "PLACA", "OFICINA"],
                        success="Lançamento de manutenção salvo.",
                    )

            _render_dataset_editor(
                "manutencao",
                backend.load_manutencao,
                ["Data", "Mes", "PLACA", "Categoria", "OFICINA", "Custo"],
                ["Data", "PLACA", "OFICINA"],
                "cad_manu_table",
                {
                    "Data": _date_col(),
                    "Mes": st.column_config.TextColumn("Mes"),
                    "PLACA": st.column_config.TextColumn("Placa"),
                    "Categoria": st.column_config.SelectboxColumn("Categoria", options=["Transporte", "Vex"], required=True),
                    "OFICINA": st.column_config.TextColumn("Oficina"),
                    "Custo": _money_col("Custo"),
                },
                ["Mes", "PLACA", "Categoria", "OFICINA"],
            )

        with tabs[4]:
            _render_hoteis_sheet_import()

            with st.form("form_hoteis", clear_on_submit=True):
                c1, c2, c3 = st.columns(3)
                with c1:
                    data = st.date_input("Data", value=date.today(), key="cad_hotel_data")
                    cidade = st.text_input("Cidade", key="cad_hotel_cidade")
                    hotel = st.text_input("Hotel/Pousada", key="cad_hotel_nome")
                with c2:
                    motorista = st.text_input("Motorista", key="cad_hotel_motorista")
                    ajudante = st.text_input("Ajudante", key="cad_hotel_ajudante")
                    tipo = st.text_input("Tipo", placeholder="Hospedagem", key="cad_hotel_tipo")
                with c3:
                    dias = st.number_input("Dias", min_value=0.0, step=1.0, key="cad_hotel_dias")
                    valor = st.number_input("Valor", min_value=0.0, step=10.0, format="%.2f", key="cad_hotel_valor")
                    submitted = st.form_submit_button("Salvar hotel", type="primary", width="stretch")
                if submitted:
                    _save_entry(
                        "hoteis",
                        {
                            "Data": data,
                            "Mes": _entry_month(data),
                            "Valor": valor,
                            "Dias": dias,
                            "Motorista": motorista,
                            "Ajudante": ajudante,
                            "Cidade": cidade,
                            "Hotel": hotel,
                            "Tipo": tipo,
                            "Categoria": "Transporte",
                        },
                        required=["Data", "Cidade", "Hotel"],
                        success="Reserva/hospedagem salva.",
                    )

            _render_dataset_editor(
                "hoteis",
                backend.load_hoteis,
                ["Data", "Mes", "Cidade", "Hotel", "Tipo", "Motorista", "Ajudante", "Dias", "Valor", "Categoria"],
                ["Data", "Cidade", "Hotel"],
                "cad_hotel_table",
                {
                    "Data": _date_col(),
                    "Mes": st.column_config.TextColumn("Mes"),
                    "Cidade": st.column_config.TextColumn("Cidade"),
                    "Hotel": st.column_config.TextColumn("Hotel/Pousada"),
                    "Tipo": st.column_config.TextColumn("Tipo"),
                    "Motorista": st.column_config.TextColumn("Motorista"),
                    "Ajudante": st.column_config.TextColumn("Ajudante"),
                    "Dias": _number_col("Dias"),
                    "Valor": _money_col("Valor"),
                    "Categoria": st.column_config.SelectboxColumn("Categoria", options=["Transporte", "Vex"], required=True),
                },
                ["Mes", "Cidade", "Hotel", "Tipo", "Motorista", "Categoria"],
            )

        with tabs[5]:
            _render_peso_sheet_import(plate_map)
            _render_peso_month_reset()

            with st.form("form_peso", clear_on_submit=True):
                c1, c2, c3 = st.columns(3)
                with c1:
                    data = st.date_input("Data", value=date.today(), key="cad_peso_data")
                    placa, categoria = _plate_fields("cad_peso", plate_map)
                with c2:
                    cidade = st.text_input("Cidade", key="cad_peso_cidade")
                with c3:
                    peso = st.number_input("Peso", min_value=0.0, step=1.0, format="%.3f", key="cad_peso_peso")
                    valor = st.number_input("Valor", min_value=0.0, step=10.0, format="%.2f", key="cad_peso_valor")
                    submitted = st.form_submit_button("Salvar peso", type="primary", width="stretch")
                if submitted:
                    _save_entry(
                        "peso",
                        {
                            "Data": data,
                            "Mes": _entry_month(data),
                            "Cidade": cidade,
                            "Peso": peso,
                            "Valor": valor,
                            "PLACA": placa,
                            "Categoria": categoria,
                        },
                        required=["Data", "PLACA", "Cidade", "Peso"],
                        success="Lancamento de peso salvo.",
                    )

            _render_dataset_editor(
                "peso",
                backend.load_peso,
                ["Data", "Mes", "Cidade", "PLACA", "Categoria", "Peso", "Valor"],
                ["Data", "PLACA", "Cidade"],
                "cad_peso_table",
                {
                    "Data": _date_col(),
                    "Mes": st.column_config.TextColumn("Mes"),
                    "Cidade": st.column_config.TextColumn("Cidade"),
                    "PLACA": st.column_config.TextColumn("Placa"),
                    "Categoria": st.column_config.SelectboxColumn("Categoria", options=["Transporte", "Vex"], required=True),
                    "Peso": _number_col("Peso"),
                    "Valor": _money_col("Valor"),
                },
                ["Mes", "Cidade", "PLACA", "Categoria"],
            )

        with tabs[6]:
            _render_pedagio_sheet_import(plate_map)

            with st.form("form_pedagio", clear_on_submit=True):
                c1, c2, c3 = st.columns(3)
                with c1:
                    data = st.date_input("Data", value=date.today(), key="cad_ped_data")
                    placa, categoria = _plate_fields("cad_ped", plate_map)
                with c2:
                    tipo = st.selectbox("Tipo", ["Pedagio", "IPVA", "Seguro", "Licenciamento", "DPVAT", "Outros"], key="cad_ped_tipo")
                with c3:
                    custo = st.number_input("Custo", min_value=0.0, step=10.0, format="%.2f", key="cad_ped_custo")
                    submitted = st.form_submit_button("Salvar pedágio/IPVA", type="primary", width="stretch")
                if submitted:
                    _save_entry(
                        "pedagio",
                        {
                            "Data": data,
                            "Mes": _entry_month(data),
                            "PLACA": placa,
                            "Tipo": tipo,
                            "Custo": custo,
                            "Categoria": categoria,
                        },
                        required=["Data", "PLACA", "Tipo"],
                        success="Lançamento de pedágio/IPVA salvo.",
                    )
            _render_dataset_editor(
                "pedagio",
                backend.load_pedagio,
                ["Data", "Mes", "PLACA", "Categoria", "Tipo", "Custo"],
                ["Data", "PLACA", "Tipo"],
                "cad_ped_table",
                {
                    "Data": _date_col(),
                    "Mes": st.column_config.TextColumn("Mes"),
                    "PLACA": st.column_config.TextColumn("Placa"),
                    "Categoria": st.column_config.SelectboxColumn("Categoria", options=["Transporte", "Vex"], required=True),
                    "Tipo": st.column_config.SelectboxColumn("Tipo", options=["Pedagio", "IPVA", "Seguro", "Licenciamento", "DPVAT", "Outros"], required=True),
                    "Custo": _money_col("Custo"),
                },
                ["Mes", "PLACA", "Categoria", "Tipo"],
            )


def render_combustivel() -> None:
    topbar("JR DASHBOARD • Combustível", back=False)
    seed = route_json("combustivel", {"ano": "Todos", "mes": ["Todos"]})
    params, filter_state = filter_controls(
        "combustivel",
        extra_filters=[
            ("placa", "Placa", seed.get("placas", []) or []),
            ("posto", "Posto", seed.get("postos", []) or []),
            ("combustivel", "Combustível", seed.get("combustiveis", []) or []),
        ],
        key_prefix="comb",
        all_data=seed,
    )
    data = route_json("combustivel", params)
    kpis = [
        ("total", "Total (R$)", fmt_brl(data.get("custo_total"))),
        ("media_mensal", "Média mensal (R$)", fmt_brl(data.get("media_mensal"))),
        ("km_total", "KM total", fmt_num(data.get("km_total"))),
        ("litros_total", "Total litros", fmt_num(data.get("litros_total"))),
        ("custo_km", "Custo médio por KM", fmt_brl(data.get("custo_por_km"))),
        ("media_kml", "Média KM/L", f"{fmt_num(data.get('km_por_litro'), 2)} km/L"),
        ("custo_litro", "Custo médio por litro", fmt_brl(data.get("custo_por_litro"))),
    ]
    include_year = params.get("ano") is None
    fallback_year = params.get("ano")
    mensal_labels, mensal_values = sorted_series(data.get("custo_mensal", {}), "Mes", "Custo", include_year=include_year, fallback_year=fallback_year)
    km_labels, km_values = sorted_series(data.get("km_mensal", {}), "Mes", "Km Rodados", include_year=include_year, fallback_year=fallback_year)
    litros_labels, litros_values = sorted_series(data.get("litros_mensal", {}), "Mes", "Litros", include_year=include_year, fallback_year=fallback_year)
    compare_selected = filter_state.get("_compare", [])
    compare_data = compare_bundle("combustivel", data, params, compare_selected)
    if compare_selected:
        kpis = compare_combined_kpis(compare_data) + color_kpis(kpis, "combustivel", section=_series_label("combustivel")) + compare_kpi_cards(compare_data[1:])
    monthly_fig = (
        multi_line_chart(compare_chart_series(compare_data, "monthly"), include_year=include_year, fallback_year=fallback_year)
        if compare_selected
        else yearly_month_line_chart(data.get("custo_mensal", {}), "Mes", "Custo", fallback_year=fallback_year)
        if include_year
        else line_chart(mensal_labels, mensal_values)
    )
    weekly_fig = (
        multi_bar_chart(compare_chart_series(compare_data, "weekly"))
        if compare_selected
        else bar_chart(data.get("gasto_semana", {}).get("Dia", []), data.get("gasto_semana", {}).get("Custo", []))
    )
    plate_fig = (
        multi_bar_chart(compare_chart_series(compare_data, "plate"), horizontal=True, sort_desc=True)
        if compare_selected
        else bar_chart(data.get("gasto_por_placa", {}).get("PLACA", []), data.get("gasto_por_placa", {}).get("Custo", []), horizontal=True, sort_desc=True, show_text=True)
    )
    charts = [
        ("gasto_mes", "Gasto total por mês", monthly_fig),
        ("gasto_semana", "Gasto semanal (últimos 7 dias)", weekly_fig),
        ("gasto_posto", "Gasto por posto", pie_chart(data.get("gasto_por_posto", {}).get("POSTOS", []), data.get("gasto_por_posto", {}).get("Custo", []))),
        ("gasto_combustivel", "Gasto por combustível", pie_chart(data.get("gasto_por_combustivel", {}).get("Combustivel", []), data.get("gasto_por_combustivel", {}).get("Custo", []))),
        ("gasto_placa", "Gasto por placa", plate_fig),
        (
            "km_mes",
            "KM por mês",
            yearly_month_bar_chart(data.get("km_mensal", {}), "Mes", "Km Rodados", fallback_year=fallback_year, currency=False)
            if include_year
            else bar_chart(km_labels, km_values, currency=False, show_text=True),
        ),
        (
            "litros_mes",
            "Litros por mês",
            yearly_month_bar_chart(data.get("litros_mensal", {}), "Mes", "Litros", fallback_year=fallback_year, currency=False)
            if include_year
            else bar_chart(litros_labels, litros_values, currency=False, show_text=True),
        ),
    ]
    render_controlled_dashboard("comb", title="JR Dashboard - Combustível", kpis=kpis, charts=charts)
    footer("Dados atualizados pelo Neon. © JR")


def render_manutencao() -> None:
    topbar("JR DASHBOARD • Manutenção", back=False)
    seed = route_json("manutencao", {"ano": "Todos", "mes": ["Todos"]})
    params, filter_state = filter_controls(
        "manutencao",
        extra_filters=[
            ("placa", "Placa", seed.get("placas", []) or []),
            ("oficina", "Oficina", seed.get("oficinas", []) or []),
            ("segmento", "Categoria", seed.get("segmentos", []) or []),
        ],
        key_prefix="manu",
        all_data=seed,
    )
    data = route_json("manutencao", params)
    kpis = [
        ("custo_total", "Custo total (R$)", fmt_brl(data.get("custo_total"))),
        ("servicos", "Serviços", fmt_num(data.get("total_servicos"))),
        ("ticket_medio", "Ticket médio", fmt_brl(data.get("media_servico"))),
        ("media_mensal", "Média mensal (R$)", fmt_brl(data.get("media_mensal"))),
    ]
    include_year = params.get("ano") is None
    mensal_labels, mensal_values = sorted_series(data.get("custo_mensal", {}), "Mes", "Custo", include_year=include_year, fallback_year=params.get("ano"))
    compare_selected = filter_state.get("_compare", [])
    compare_data = compare_bundle("manutencao", data, params, compare_selected)
    if compare_selected:
        kpis = compare_combined_kpis(compare_data) + color_kpis(kpis, "manutencao", section=_series_label("manutencao")) + compare_kpi_cards(compare_data[1:])
    monthly_fig = (
        multi_line_chart(compare_chart_series(compare_data, "monthly"), include_year=include_year, fallback_year=params.get("ano"))
        if compare_selected
        else yearly_month_line_chart(data.get("custo_mensal", {}), "Mes", "Custo", fallback_year=params.get("ano"))
        if include_year
        else line_chart(mensal_labels, mensal_values)
    )
    weekly_fig = (
        multi_bar_chart(compare_chart_series(compare_data, "weekly"))
        if compare_selected
        else bar_chart(data.get("custo_semana", {}).get("Dia", []), data.get("custo_semana", {}).get("Custo", []))
    )
    plate_fig = (
        multi_bar_chart(compare_chart_series(compare_data, "plate"), horizontal=True, sort_desc=True)
        if compare_selected
        else bar_chart(data.get("gasto_por_placa", {}).get("PLACA", []), data.get("gasto_por_placa", {}).get("Custo", []), horizontal=True, sort_desc=True, show_text=True)
    )
    charts = [
        ("gasto_placa", "Gasto por placa", plate_fig),
        ("gasto_oficina", "Gasto por oficina", bar_chart(data.get("gasto_por_oficina", {}).get("OFICINA", []), data.get("gasto_por_oficina", {}).get("Custo", []), horizontal=True, sort_desc=True, show_text=True)),
        ("gasto_mensal", "Gasto mensal", monthly_fig),
        ("gasto_semana", "Gasto semanal (últimos 7 dias)", weekly_fig),
    ]
    render_controlled_dashboard("manu", title="JR Dashboard - Manutenção", kpis=kpis, charts=charts)
    footer("Dados atualizados pelo Neon. © JR")


def render_hoteis() -> None:
    topbar("JR DASHBOARD • Reservas de Hotéis", back=False)
    seed = route_json("hoteis", {"ano": "Todos", "mes": ["Todos"]})
    params, filter_state = filter_controls(
        "hoteis",
        extra_filters=[
            ("cidade", "Cidade", seed.get("cidades", []) or []),
            ("hotel", "Hotel/Pousada", seed.get("hoteis", []) or []),
        ],
        key_prefix="hotel",
        all_data=seed,
    )
    data = route_json("hoteis", params)
    kpis = [
        ("valor_total", "Valor total (R$)", fmt_brl(data.get("valor_total"))),
        ("reservas", "Reservas", fmt_num(data.get("reservas_total"))),
        ("reservas_nao_planejadas", "Reservas não planejadas", fmt_num(data.get("reservas_nao_planejadas"))),
        ("media_reserva", "Média por reserva", fmt_brl(data.get("valor_medio_reserva"))),
        ("media_mensal", "Média mensal", fmt_brl(data.get("media_mensal"))),
        ("sabados", "Gasto aos sábados (R$)", fmt_brl(data.get("valor_sabado"))),
        ("nao_planejado", "Gasto não planejado (R$)", fmt_brl(data.get("valor_nao_planejado"))),
    ]
    include_year = params.get("ano") is None
    mensal_labels, mensal_values = sorted_series(data.get("valor_mensal", {}), "Mes", "Valor", include_year=include_year, fallback_year=params.get("ano"))
    week = data.get("valor_semana", {})
    week_colors = []
    for is_unplanned in week.get("NaoPlanejada", []) or []:
        week_colors.append("#F59E0B" if is_unplanned else JR_BLUE)
    compare_selected = filter_state.get("_compare", [])
    compare_data = compare_bundle("hoteis", data, params, compare_selected)
    if compare_selected:
        kpis = compare_combined_kpis(compare_data) + color_kpis(kpis, "hoteis", section=_series_label("hoteis")) + compare_kpi_cards(compare_data[1:])
    monthly_fig = (
        multi_line_chart(compare_chart_series(compare_data, "monthly"), include_year=include_year, fallback_year=params.get("ano"))
        if compare_selected
        else yearly_month_line_chart(data.get("valor_mensal", {}), "Mes", "Valor", fallback_year=params.get("ano"))
        if include_year
        else line_chart(mensal_labels, mensal_values)
    )
    weekly_fig = (
        multi_bar_chart(compare_chart_series(compare_data, "weekly"))
        if compare_selected
        else bar_chart(week.get("Dia", []), week.get("Valor", []), marker_colors=week_colors or None)
    )
    charts = [
        ("valor_mensal", "Valor mensal", monthly_fig),
        ("valor_semanal", "Valor semanal (últimos 7 dias)", weekly_fig),
        ("valor_cidade", "Valor por cidade", bar_chart(data.get("valor_por_cidade", {}).get("Cidade", []), data.get("valor_por_cidade", {}).get("Valor", []), horizontal=True, sort_desc=True, show_text=True)),
        ("valor_hotel", "Valor por hotel/pousada", bar_chart(data.get("valor_por_hotel", {}).get("Hotel", []), data.get("valor_por_hotel", {}).get("Valor", []), horizontal=True, sort_desc=True, show_text=True)),
    ]
    render_controlled_dashboard("hotel", title="JR Dashboard - Reservas de Hotéis", kpis=kpis, charts=charts)
    footer("Dados atualizados pelo Neon. © JR")


def render_pedagio() -> None:
    topbar("JR DASHBOARD • Pedágio, Seguro e IPVA", back=False)
    seed = route_json("pedagio", {"ano": "Todos", "mes": ["Todos"]})
    params, filter_state = filter_controls(
        "pedagio",
        extra_filters=[
            ("placa", "Placa", seed.get("placas", []) or []),
            ("tipo", "Tipo", seed.get("tipos", []) or []),
            ("segmento", "Segmento", seed.get("segmentos", []) or []),
        ],
        key_prefix="ped",
        all_data=seed,
    )
    data = route_json("pedagio", params)
    kpis = [
        ("gasto_total", "Gasto total (R$)", fmt_brl(data.get("custo_total"))),
        ("media_mensal", "Média mensal (R$)", fmt_brl(data.get("media_mensal"))),
        ("gasto_pedagio", "Gasto pedágio", fmt_brl(data.get("gasto_pedagio"))),
        ("gasto_ipva", "Gasto IPVA", fmt_brl(data.get("gasto_ipva"))),
        ("gasto_seguro", "Gasto seguro", fmt_brl(data.get("gasto_seguro"))),
        ("media_valores", "Média de valores", fmt_brl(data.get("media_valores", data.get("ticket_medio")))),
    ]
    include_year = params.get("ano") is None
    mensal_labels, mensal_values = sorted_series(data.get("custo_mensal", {}), "Mes", "Custo", include_year=include_year, fallback_year=params.get("ano"))
    compare_selected = filter_state.get("_compare", [])
    compare_data = compare_bundle("pedagio", data, params, compare_selected)
    if compare_selected:
        kpis = compare_combined_kpis(compare_data) + color_kpis(kpis, "pedagio", section=_series_label("pedagio")) + compare_kpi_cards(compare_data[1:])
    monthly_fig = (
        multi_line_chart(compare_chart_series(compare_data, "monthly"), include_year=include_year, fallback_year=params.get("ano"))
        if compare_selected
        else yearly_month_line_chart(data.get("custo_mensal", {}), "Mes", "Custo", fallback_year=params.get("ano"))
        if include_year
        else line_chart(mensal_labels, mensal_values)
    )
    weekly_fig = (
        multi_bar_chart(compare_chart_series(compare_data, "weekly"))
        if compare_selected
        else bar_chart(data.get("custo_semana", {}).get("Dia", []), data.get("custo_semana", {}).get("Custo", []))
    )
    plate_fig = (
        multi_bar_chart(compare_chart_series(compare_data, "plate"), horizontal=True, sort_desc=True)
        if compare_selected
        else bar_chart(data.get("gasto_por_placa", {}).get("PLACA", []), data.get("gasto_por_placa", {}).get("Custo", []), horizontal=True, sort_desc=True, show_text=True)
    )
    charts = [
        ("gasto_mensal", "Gasto mensal", monthly_fig),
        ("gasto_semana", "Gasto semanal (últimos 7 dias)", weekly_fig),
        ("gasto_tipo", "Gasto por tipo", pie_chart(data.get("gasto_por_tipo", {}).get("Tipo", []), data.get("gasto_por_tipo", {}).get("Custo", []))),
        ("gasto_placa", "Gasto por placa", plate_fig),
        ("gasto_segmento", "Gasto por segmento", bar_chart(data.get("gasto_por_categoria", {}).get("Categoria", []), data.get("gasto_por_categoria", {}).get("Custo", []))),
    ]
    render_controlled_dashboard("ped", title="JR Dashboard - Pedágio, Seguro e IPVA", kpis=kpis, charts=charts)
    footer("Dados atualizados pelo Neon. © JR")


def render_vex() -> None:
    topbar("JR DASHBOARD • Vex", back=False)
    seed = route_json("vex", {"ano": "Todos", "mes": ["Todos"]})
    params, filter_state = filter_controls(
        "vex",
        extra_filters=[("placa", "Placa", seed.get("placas", []) or [])],
        key_prefix="vex",
        all_data=seed,
    )
    data = route_json("vex", params)
    kpis = [
        ("gasto_vex", "Gasto Vex total (R$)", fmt_brl(data.get("total_vex"))),
        ("combustivel_vex", "Combustível Vex (R$)", fmt_brl(data.get("combustivel_total"))),
        ("km_total", "KM total", fmt_num(data.get("km_total"))),
        ("litros_total", "Total litros", fmt_num(data.get("litros_total"))),
        ("custo_km", "Custo médio por KM", fmt_brl(data.get("custo_por_km"))),
        ("media_kml", "Média KM/L", fmt_num(data.get("km_por_litro"), 2)),
        ("custo_litro", "Custo médio por litro", fmt_brl(data.get("custo_por_litro"))),
        ("manutencao_vex", "Manutenção Vex (R$)", fmt_brl(data.get("manutencao_total"))),
        ("pedagio_vex", "Pedágio/Seguro Vex (R$)", fmt_brl(data.get("pedagio_total"))),
    ]
    include_year = params.get("ano") is None
    mensal_labels, mensal_values = sorted_series(data.get("mensal_total", {}), "Mes", "Valor", include_year=include_year, fallback_year=params.get("ano"))
    compare_selected = filter_state.get("_compare", [])
    compare_data = compare_bundle("vex", data, params, compare_selected)
    if compare_selected:
        kpis = compare_combined_kpis(compare_data) + color_kpis(kpis, "vex", section=_series_label("vex")) + compare_kpi_cards(compare_data[1:])
    monthly_fig = (
        multi_line_chart(compare_chart_series(compare_data, "monthly"), include_year=include_year, fallback_year=params.get("ano"))
        if compare_selected
        else yearly_month_line_chart(data.get("mensal_total", {}), "Mes", "Valor", fallback_year=params.get("ano"))
        if include_year
        else line_chart(mensal_labels, mensal_values)
    )
    plate_fig = (
        multi_bar_chart(compare_chart_series(compare_data, "plate"), horizontal=True, sort_desc=True)
        if compare_selected
        else bar_chart(data.get("gasto_por_placa", {}).get("PLACA", []), data.get("gasto_por_placa", {}).get("Valor", []), horizontal=True, sort_desc=True, show_text=True)
    )
    charts = [
        ("gasto_mensal", "Gasto Vex mensal", monthly_fig),
        ("gasto_area", "Gasto Vex por área", bar_chart(data.get("por_area", {}).get("Area", []), data.get("por_area", {}).get("Valor", []))),
        ("gasto_placa", "Gasto Vex por placa", plate_fig),
    ]
    render_controlled_dashboard("vex", title="JR Dashboard - Vex", kpis=kpis, charts=charts)
    footer("Dados Vex consolidados pelo Neon. © JR")


def footer(text: str) -> None:
    st.markdown(f'<div class="footer-note">{h(text)}</div>', unsafe_allow_html=True)


def main() -> None:
    st.set_page_config(page_title="JR Dashboard", page_icon=str(LOGO_PATH), layout="wide")
    inject_css()
    page = page_param()
    try:
        if page == "combustivel":
            render_combustivel()
        elif page == "manutencao":
            render_manutencao()
        elif page == "hoteis":
            render_hoteis()
        elif page == "pedagio":
            render_pedagio()
        elif page == "vex":
            render_vex()
        elif page in {"frota", "ranking"}:
            render_frota()
        elif page in {"cadastro", "dados"}:
            render_cadastro()
        else:
            render_home()
    except Exception as exc:
        st.error("Não foi possível carregar este dashboard. Configure o DATABASE_URL do Neon nos Secrets do Streamlit e tente novamente.")
        st.exception(exc)


if __name__ == "__main__":
    main()
