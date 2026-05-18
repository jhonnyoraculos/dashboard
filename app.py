from flask import Flask, render_template, jsonify, request
import io
import os
import re
import threading
import zipfile
import concurrent.futures
import pandas as pd
import unicodedata
from collections import defaultdict
from pathlib import Path


STREAMLIT_APP_MODULE = "streamlit_app"


def _running_as_streamlit_entrypoint() -> bool:
    if __name__ != "__main__":
        return False
    try:
        from streamlit.runtime.scriptrunner import get_script_run_ctx
    except Exception:
        return False
    return get_script_run_ctx() is not None


if _running_as_streamlit_entrypoint():
    os.environ.setdefault("JR_SKIP_WARM_CACHE", "1")
    # Streamlit Cloud may be configured to run app.py. Keep this shim pointing
    # at the real Streamlit interface so changes in streamlit_app.py go live.
    from streamlit_app import main as streamlit_main
    import streamlit as st

    streamlit_main()
    st.stop()


app = Flask(__name__)

BASE_PATH = Path(__file__).parent / "data"
DATA_COMB = BASE_PATH / "combustivel.xlsx"
DATA_MANU = BASE_PATH / "manutencao.xlsx"
DATA_HOTEIS = BASE_PATH / "reserva de hoteis.xlsx"
DATA_PEDAGIO = BASE_PATH / "pedagio seguro e ipva.xlsx"
_PEDAGIO_CACHE = {"mtime": None, "df": None, "lock": threading.Lock()}
_COMBUSTIVEL_CACHE = {"mtime": None, "df": None, "lock": threading.Lock()}
_MANUTENCAO_CACHE = {"mtime": None, "df": None, "lock": threading.Lock()}
_HOTEIS_CACHE = {"mtime": None, "df": None, "lock": threading.Lock()}
_OVERVIEW_CACHE = {"mtimes": None, "dados": None}
_CACHE_MAP = {
    "combustivel": _COMBUSTIVEL_CACHE,
    "manutencao": _MANUTENCAO_CACHE,
    "hoteis": _HOTEIS_CACHE,
    "pedagio": _PEDAGIO_CACHE,
}
DB_TABLES = {
    "combustivel": "dashboard_combustivel",
    "combustivel_km": "dashboard_combustivel_km",
    "manutencao": "dashboard_manutencao",
    "hoteis": "dashboard_hoteis",
    "pedagio": "dashboard_pedagio",
}
DB_METADATA_TABLE = "dashboard_metadata"
_DB_ENGINE = None


def _database_url() -> str | None:
    for key in ("DATABASE_URL", "NEON_DATABASE_URL"):
        value = os.environ.get(key)
        if value:
            return value.strip()
    try:
        import streamlit as st

        for key in ("DATABASE_URL", "NEON_DATABASE_URL"):
            value = st.secrets.get(key)
            if value:
                return str(value).strip()
    except Exception:
        pass
    return None


def _normalize_database_url(url: str) -> str:
    if url.startswith("postgresql+"):
        return url
    if url.startswith("postgresql://"):
        return f"postgresql+psycopg://{url[len('postgresql://'):]}"
    if url.startswith("postgres://"):
        return f"postgresql+psycopg://{url[len('postgres://'):]}"
    return url


def _db_engine():
    global _DB_ENGINE
    url = _database_url()
    if not url:
        raise RuntimeError("DATABASE_URL nao configurada para ler o banco Neon.")
    normalized_url = _normalize_database_url(url)
    if _DB_ENGINE is None:
        from sqlalchemy import create_engine

        _DB_ENGINE = create_engine(normalized_url, pool_pre_ping=True)
    return _DB_ENGINE


def _data_source_mode() -> str:
    value = os.environ.get("JR_DATA_SOURCE")
    if value:
        return value.strip().lower()
    try:
        import streamlit as st

        value = st.secrets.get("JR_DATA_SOURCE")
        if value:
            return str(value).strip().lower()
    except Exception:
        pass
    return "auto"


def _excel_files_available() -> bool:
    return all(path.exists() for path in (DATA_COMB, DATA_MANU, DATA_HOTEIS, DATA_PEDAGIO))


def _should_use_database() -> bool:
    mode = _data_source_mode()
    if mode in {"excel", "xlsx", "planilha", "file", "files"}:
        return False
    if mode in {"database", "db", "postgres", "postgresql", "neon"}:
        if not _database_url():
            raise RuntimeError("JR_DATA_SOURCE=database exige DATABASE_URL/NEON_DATABASE_URL.")
        return True
    if _database_url():
        return True
    if _excel_files_available():
        return False
    raise RuntimeError(
        "DATABASE_URL/NEON_DATABASE_URL nao configurada. "
        "As planilhas foram removidas do repositorio, entao o app precisa ler do Neon."
    )


def _db_metadata(key: str, default=None):
    try:
        from sqlalchemy import text

        query = text(f'SELECT value_json FROM "{DB_METADATA_TABLE}" WHERE "key" = :key')
        rows = pd.read_sql_query(query, _db_engine(), params={"key": key})
    except Exception:
        return default
    if rows.empty:
        return default
    try:
        import json

        return json.loads(rows.iloc[0]["value_json"])
    except Exception:
        return default


def _db_version(dataset: str):
    return _db_metadata(f"{dataset}.version", _db_metadata("import.version", "database"))


def _read_database_table(dataset: str, columns: list[str], *, date_columns: list[str] | None = None) -> pd.DataFrame:
    table = DB_TABLES[dataset]
    try:
        from sqlalchemy import text

        df = pd.read_sql_query(text(f'SELECT * FROM "{table}"'), _db_engine())
    except Exception as exc:
        raise RuntimeError(f'Tabela "{table}" nao encontrada no Neon. Rode scripts/import_to_neon.py antes de abrir o app.') from exc
    for column in columns:
        if column not in df.columns:
            df[column] = pd.NA
    if date_columns:
        for column in date_columns:
            if column in df.columns:
                df[column] = pd.to_datetime(df[column], errors="coerce")
    df = df[columns + [column for column in df.columns if column not in columns]]
    df.attrs["anos_sheets"] = _db_metadata(f"{dataset}.anos_sheets", [])
    return df.copy()


def _clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Remove colunas artificiais e normaliza nomes em ASCII."""
    df = df.loc[:, ~df.columns.astype(str).str.startswith("Unnamed")]
    rename_map = {}
    keep_cols = []
    for col in df.columns:
        normal = unicodedata.normalize("NFKD", str(col)).encode("ascii", "ignore").decode("ascii")
        cleaned = normal.strip()
        if cleaned == "" or cleaned.lower() == "nan":
            continue
        keep_cols.append(col)
        rename_map[col] = cleaned
    df = df.loc[:, keep_cols]
    return df.rename(columns=rename_map)


def _group_sum(
    df: pd.DataFrame,
    group_col: str,
    value_col: str = "Custo",
    *,
    sort_by: str = "value",
) -> dict:
    if df.empty or group_col not in df.columns or value_col not in df.columns:
        return {group_col: [], value_col: []}

    grouped = (
        df.dropna(subset=[group_col])
        .groupby(group_col, as_index=False)[value_col]
        .sum()
    )

    if sort_by == "group" and group_col in grouped.columns:
        grouped = grouped.sort_values(group_col)
    else:
        grouped = grouped.sort_values(value_col, ascending=False)

    return grouped.to_dict(orient="list")


def _unique_sorted(df: pd.DataFrame, column: str) -> list:
    if column not in df.columns:
        return []
    series = df[column].dropna()
    if series.empty:
        return []
    series = series.astype("string").str.strip()
    series = series[(series != "") & (~series.str.lower().isin(["nan", "none", "nat", "<na>"]))]
    return sorted(series.unique().tolist())


def _unique_years(df: pd.DataFrame) -> list[int]:
    if "Mes" not in df.columns:
        return []
    periodos = pd.to_datetime(df["Mes"], errors="coerce")
    anos = sorted({int(ano) for ano in periodos.dt.year.dropna().unique()})
    return anos


def _years_from_sheet_names(sheet_names) -> list[int]:
    years = set()
    for name in sheet_names or []:
        for match in re.findall(r"(20\d{2})", str(name)):
            years.add(int(match))
    return sorted(years)


def _sheet_year(name: str) -> int | None:
    match = re.search(r"(20\d{2})", str(name))
    if match:
        return int(match.group(1))
    return None


def _apply_sheet_year(
    df: pd.DataFrame,
    *,
    year_col: str = "_SheetYear",
    date_col: str | None = "Data",
    mes_col: str | None = "Mes",
) -> pd.DataFrame:
    if year_col not in df.columns:
        return df
    year_series = pd.to_numeric(df[year_col], errors="coerce")
    if year_series.isna().all():
        return df

    if date_col and date_col in df.columns:
        dt = pd.to_datetime(df[date_col], errors="coerce")
        valid = dt.notna() & year_series.notna()
        if valid.any():
            parts = pd.DataFrame({
                "year": year_series[valid].astype(int),
                "month": dt.loc[valid].dt.month,
                "day": dt.loc[valid].dt.day,
            })
            new_dt = pd.to_datetime(parts, errors="coerce")
            dt = dt.copy()
            dt.loc[valid] = new_dt
            df[date_col] = dt

    if mes_col and mes_col in df.columns:
        mes_dt = pd.to_datetime(df[mes_col], errors="coerce")
        month_source = mes_dt.dt.month
        if date_col and date_col in df.columns:
            dt = pd.to_datetime(df[date_col], errors="coerce")
            month_source = month_source.fillna(dt.dt.month)
        valid_month = month_source.notna() & year_series.notna()
        if valid_month.any():
            parts = pd.DataFrame({
                "year": year_series[valid_month].astype(int),
                "month": month_source[valid_month].astype(int),
                "day": 1,
            })
            new_mes_dt = pd.to_datetime(parts, errors="coerce")
            mes_dt = mes_dt.copy()
            mes_dt.loc[valid_month] = new_mes_dt
            df[mes_col] = mes_dt.dt.to_period("M").astype(str)

    return df


def _parse_int(value, *, min_value: int | None = None, max_value: int | None = None) -> int | None:
    if value in (None, "", "Todos"):
        return None
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return None
    if min_value is not None and parsed < min_value:
        return None
    if max_value is not None and parsed > max_value:
        return None
    return parsed


def _normalize_categoria(series: pd.Series) -> pd.Series:
    if series is None:
        return pd.Series(dtype="string")
    return (
        series.astype("string")
        .fillna("")
        .str.strip()
        .str.lower()
    )


def _exclude_vex(df: pd.DataFrame) -> pd.DataFrame:
    if "Categoria" not in df.columns:
        return df
    mask = _normalize_categoria(df["Categoria"]) != "vex"
    return df.loc[mask].copy()


def _only_vex(df: pd.DataFrame) -> pd.DataFrame:
    if "Categoria" not in df.columns:
        return df.iloc[0:0].copy()
    mask = _normalize_categoria(df["Categoria"]) == "vex"
    return df.loc[mask].copy()


def _filter_by_period(
    df: pd.DataFrame,
    *,
    ano: int | None = None,
    mes: int | None = None,
    meses: list[int] | None = None,
) -> pd.DataFrame:
    meses = meses or []
    if df.empty or "Mes" not in df.columns or (ano is None and mes is None and not meses):
        return df
    periodos = pd.to_datetime(df["Mes"], errors="coerce")
    mask = periodos.notna()
    if ano is not None:
        mask &= periodos.dt.year == ano
    if mes is not None:
        mask &= periodos.dt.month == mes
    if meses:
        mask &= periodos.dt.month.isin(meses)
    return df.loc[mask].copy()


def _parse_mes_list(raw: list[str] | str | None) -> list[str]:
    if raw is None:
        return []
    if isinstance(raw, str):
        raw_values = [raw]
    else:
        raw_values = list(raw)

    meses: list[str] = []
    for value in raw_values:
        if value in (None, "", "Todos"):
            continue
        parts = str(value).split(",")
        for part in parts:
            mes = part.strip()
            if mes and mes.lower() != "todos":
                meses.append(mes)
    return meses


def _parse_mes_int_list(raw: list[str] | None) -> list[int]:
    if raw is None:
        return []
    meses: list[int] = []
    for value in raw:
        if value in (None, "", "Todos"):
            continue
        try:
            num = int(value)
        except (TypeError, ValueError):
            continue
        if 1 <= num <= 12:
            meses.append(num)
    return meses


def _weekly_series(df: pd.DataFrame, date_col: str, value_col: str, label: str) -> dict:
    today = pd.Timestamp.today().normalize()
    start_default = today - pd.Timedelta(days=6)
    default_index = pd.date_range(start_default, today, freq="D")
    template = {
        "Dia": default_index.strftime("%d/%m").tolist(),
        "DiaISO": default_index.strftime("%Y-%m-%d").tolist(),
        label: [0.0] * len(default_index),
    }

    if df.empty or date_col not in df.columns or value_col not in df.columns:
        return template

    dates = pd.to_datetime(df[date_col], errors="coerce")
    values = pd.to_numeric(df[value_col], errors="coerce")
    valid = dates.notna() & values.notna()
    if not valid.any():
        return template

    data = pd.DataFrame(
        {
            "data": dates.loc[valid].dt.normalize(),
            "valor": values.loc[valid].astype("float64"),
        }
    )
    data = data.dropna()
    if data.empty:
        return template

    start = start_default
    end = today
    index = default_index

    window_mask = data["data"].between(start, end)
    if not window_mask.any():
        end = data["data"].max()
        if pd.isna(end):
            return template
        start = end - pd.Timedelta(days=6)
        index = pd.date_range(start, end, freq="D")
        template = {
            "Dia": index.strftime("%d/%m").tolist(),
            label: [0.0] * len(index),
        }
        window_mask = data["data"].between(start, end)
        if not window_mask.any():
            return template

    grouped = data.loc[window_mask].groupby("data")["valor"].sum()
    grouped = grouped.reindex(index, fill_value=0.0).astype("float64")
    return {
        "Dia": index.strftime("%d/%m").tolist(),
        "DiaISO": index.strftime("%Y-%m-%d").tolist(),
        label: grouped.round(2).tolist(),
    }


def _to_numeric_currency(series: pd.Series) -> pd.Series:
    if series is None:
        return pd.Series(dtype="float64")
    if pd.api.types.is_numeric_dtype(series):
        return pd.to_numeric(series, errors="coerce")

    cleaned = series.astype("string").str.strip()
    cleaned = cleaned.replace({"": pd.NA, "nan": pd.NA, "None": pd.NA})
    cleaned = cleaned.str.replace("R$", "", regex=False).str.replace("\u00a0", "", regex=False)
    cleaned = cleaned.str.replace(" ", "", regex=False)

    mask_comma = cleaned.str.contains(",", regex=False, na=False)
    cleaned.loc[mask_comma] = (
        cleaned.loc[mask_comma]
        .str.replace(".", "", regex=False)
        .str.replace(",", ".", regex=False)
    )
    cleaned.loc[~mask_comma] = cleaned.loc[~mask_comma].str.replace(",", ".", regex=False)

    return pd.to_numeric(cleaned, errors="coerce")


def _normalize_ascii(value):
    if pd.isna(value):
        return value
    return unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode("ascii").strip()


def _canonical_plate(value):
    if pd.isna(value):
        return pd.NA
    text = _normalize_ascii(value).upper()
    text = re.sub(r"\s+", " ", text).strip()
    if not text or text in {"NAN", "NONE", "NAT", "<NA>"}:
        return pd.NA
    if "SEM" in text and "PLACA" in text:
        return "SEM PLACA"
    compact = re.sub(r"[^A-Z0-9]", "", text)
    match = re.search(r"[A-Z]{3}[0-9][A-Z0-9][0-9]{2}|[A-Z]{3}[0-9]{4}", compact)
    if match:
        return match.group(0)
    return text


def _normalize_plate_series(series: pd.Series) -> pd.Series:
    if series is None:
        return pd.Series(dtype="string")
    return series.apply(_canonical_plate).astype("string")


def _canonical_tipo(value):
    if pd.isna(value):
        return None
    text = _normalize_ascii(value).upper()
    if not text:
        return None
    if "PEDAG" in text:
        return "Pedagio"
    if "IPVA" in text:
        return "IPVA"
    if "SEGUR" in text or "APOLI" in text or "APOLICE" in text:
        return "Seguro"
    if "LICENCI" in text:
        return "Licenciamento"
    if "DPVAT" in text:
        return "DPVAT"
    return text.title()


def load_combustivel() -> pd.DataFrame:
    def _empty() -> pd.DataFrame:
        return pd.DataFrame(columns=[
            "Data",
            "Mes",
            "Km Rodados",
            "Litros",
            "Custo",
            "Combustivel",
            "POSTOS",
            "PLACA",
            "Categoria",
        ])

    if _should_use_database():
        cache = _COMBUSTIVEL_CACHE
        lock = cache["lock"]
        with lock:
            version = _db_version("combustivel")
            cached = cache.get("df")
            if cached is not None and cache.get("mtime") == version:
                return cached.copy()
            df = _read_database_table(
                "combustivel",
                ["Data", "Mes", "Km Rodados", "Litros", "Custo", "Combustivel", "POSTOS", "PLACA", "Categoria"],
                date_columns=["Data"],
            )
            try:
                cache["km_rodados_mensal"] = _read_database_table(
                    "combustivel_km",
                    ["Mes", "PLACA", "Km Rodados"],
                )
            except RuntimeError:
                cache["km_rodados_mensal"] = pd.DataFrame(columns=["Mes", "PLACA", "Km Rodados"])
            cache["mtime"] = version
            cache["df"] = df.copy()
            return df.copy()

    def _read_km_rodados(path: Path) -> pd.DataFrame:
        try:
            raw = pd.read_excel(path, sheet_name="KM RODADOS", header=None, dtype=str)
        except Exception:
            return pd.DataFrame(columns=["Mes", "PLACA", "Km Rodados"])

        header_idx = None
        for idx in range(min(len(raw), 10)):
            row = raw.iloc[idx]
            normalized = set()
            for value in row:
                if pd.isna(value):
                    continue
                normalized.add(_normalize_ascii(value).upper())
            has_mes = "MES" in normalized or "MÊS" in normalized
            has_placa = "PLACA" in normalized or "PLACAS" in normalized
            has_km = any("KM" in value for value in normalized)
            if has_mes and has_placa and has_km:
                header_idx = idx
                break

        if header_idx is None:
            return pd.DataFrame(columns=["Mes", "PLACA", "Km Rodados"])

        df_km = raw.iloc[header_idx + 1 :].copy()
        df_km.columns = raw.iloc[header_idx]
        df_km = _clean_columns(df_km)
        if df_km.empty:
            return pd.DataFrame(columns=["Mes", "PLACA", "Km Rodados"])

        rename_map = {}
        for col in df_km.columns:
            col_norm = _normalize_ascii(col).upper()
            if col_norm in {"MES", "MÊS"}:
                rename_map[col] = "Mes"
            elif col_norm in {"PLACA", "PLACAS"}:
                rename_map[col] = "PLACA"
            elif "KM" in col_norm:
                rename_map[col] = "Km Rodados"
        if rename_map:
            df_km = df_km.rename(columns=rename_map)

        if "Mes" not in df_km.columns or "Km Rodados" not in df_km.columns:
            return pd.DataFrame(columns=["Mes", "PLACA", "Km Rodados"])

        df_km["PLACA"] = _normalize_plate_series(df_km.get("PLACA"))
        df_km["Km Rodados"] = pd.to_numeric(df_km.get("Km Rodados"), errors="coerce")

        mes_raw = df_km["Mes"]
        mes_dt = pd.to_datetime(mes_raw, errors="coerce")
        if mes_dt.isna().any():
            mes_norm = (
                mes_raw.astype("string")
                .str.strip()
                .str.replace(".", "/", regex=False)
                .str.replace("-", "/", regex=False)
                .str.replace(" ", "", regex=False)
                .str.lower()
            )
            mes_dt_alt = pd.to_datetime(mes_norm, errors="coerce", format="%b/%y")
            mes_dt = mes_dt.combine_first(mes_dt_alt)

        df_km["Mes"] = mes_dt.dt.to_period("M").astype(str)
        df_km = df_km.dropna(subset=["Mes", "Km Rodados"])
        return df_km[["Mes", "PLACA", "Km Rodados"]].copy()

    cache = _COMBUSTIVEL_CACHE
    lock = cache["lock"]

    with lock:
        try:
            mtime = DATA_COMB.stat().st_mtime
        except FileNotFoundError:
            cache["mtime"] = None
            cache["df"] = None
            return _empty()
        except PermissionError:
            print("Aviso: sem permissao para ler planilha de combustivel. Verifique se o arquivo esta aberto.")
            cached = cache.get("df")
            if cached is not None:
                return cached.copy()
            return _empty()

        cached = cache.get("df")
        if cached is not None and cache.get("mtime") == mtime:
            return cached.copy()

        try:
            sheets = pd.read_excel(DATA_COMB, sheet_name=None, header=1)
            sheet_years = _years_from_sheet_names(sheets.keys())
        except PermissionError:
            print("Aviso: sem permissao para ler planilha de combustivel. Verifique se o arquivo esta aberto.")
            if cached is not None:
                return cached.copy()
            return _empty()
        except Exception as exc:
            print(f"Aviso: falha ao ler planilha de combustivel: {exc}")
            if cached is not None:
                return cached.copy()
            return _empty()

        frames = []
        for sheet_name, sheet_df in sheets.items():
            if sheet_df is None or sheet_df.empty:
                continue
            cleaned = _clean_columns(sheet_df)
            if cleaned.empty:
                continue
            if "Data" not in cleaned.columns:
                continue
            sheet_year = _sheet_year(sheet_name)
            if sheet_year is not None:
                cleaned["_SheetYear"] = sheet_year
            frames.append(cleaned)
        if not frames:
            return _empty()
        df = pd.concat(frames, ignore_index=True)

        df = df.rename(columns={
            "COMBUSTIVEL": "Combustivel",
            "MES": "Mes",
        })

        df["Data"] = pd.to_datetime(df.get("Data"), errors="coerce")
        df = _apply_sheet_year(df, date_col="Data", mes_col=None)
        for col in ["Km Rodados", "Litros", "Custo"]:
            df[col] = pd.to_numeric(df.get(col), errors="coerce")

        df = df.dropna(subset=["Data"]).copy()
        df["Mes"] = df["Data"].dt.to_period("M").astype(str)

        for col in ["Combustivel", "POSTOS"]:
            if col in df.columns:
                df[col] = df[col].astype("string").str.strip()
        if "PLACA" in df.columns:
            df["PLACA"] = _normalize_plate_series(df["PLACA"])

        vex_col = next((col for col in df.columns if col.lower() == "vex"), None)
        if vex_col:
            df[vex_col] = df[vex_col].astype("string").str.strip()
            df["Categoria"] = df[vex_col].apply(lambda value: "Vex" if pd.notna(value) and value != "" else "Transporte")
        else:
            df["Categoria"] = "Transporte"

        df.attrs["anos_sheets"] = sheet_years
        cache["km_rodados_mensal"] = _read_km_rodados(DATA_COMB)
        cache["mtime"] = mtime
        cache["df"] = df.copy()
        return df


def agg_combustivel(df: pd.DataFrame, *, km_override: pd.DataFrame | None = None) -> dict:
    custo_total = float(df["Custo"].sum()) if "Custo" in df else 0.0
    if km_override is not None and "Km Rodados" in km_override.columns:
        km_total = float(km_override["Km Rodados"].sum())
    else:
        km_total = float(df["Km Rodados"].sum()) if "Km Rodados" in df else 0.0
    litros_total = float(df["Litros"].sum()) if "Litros" in df else 0.0
    custo_por_km = (custo_total / km_total) if km_total else 0.0
    km_por_litro = (km_total / litros_total) if litros_total else 0.0
    custo_por_litro = (custo_total / litros_total) if litros_total else 0.0
    meses_distintos = df["Mes"].dropna().unique() if "Mes" in df else []
    media_mensal = float(custo_total / len(meses_distintos)) if len(meses_distintos) else 0.0

    return {
        "km_total": km_total,
        "litros_total": litros_total,
        "custo_total": custo_total,
        "media_mensal": media_mensal,
        "custo_por_km": custo_por_km,
        "km_por_litro": km_por_litro,
        "custo_por_litro": custo_por_litro,
        "custo_mensal": _group_sum(df, "Mes", sort_by="group"),
        "km_mensal": _group_sum(km_override, "Mes", "Km Rodados", sort_by="group") if km_override is not None else _group_sum(df, "Mes", "Km Rodados", sort_by="group"),
        "litros_mensal": _group_sum(df, "Mes", "Litros", sort_by="group"),
        "gasto_por_posto": _group_sum(df, "POSTOS"),
        "gasto_por_combustivel": _group_sum(df, "Combustivel"),
        "gasto_por_placa": _group_sum(df, "PLACA"),
        "placas": _unique_sorted(df, "PLACA"),
        "postos": _unique_sorted(df, "POSTOS"),
        "combustiveis": _unique_sorted(df, "Combustivel"),
        "meses": _unique_sorted(df, "Mes"),
        "segmentos": _unique_sorted(df, "Categoria"),
        "gasto_semana": _weekly_series(df, "Data", "Custo", "Custo"),
    }


def load_manutencao() -> pd.DataFrame:
    def _empty() -> pd.DataFrame:
        return pd.DataFrame(columns=[
            "Data",
            "Mes",
            "Custo",
            "PLACA",
            "OFICINA",
            "Categoria",
        ])

    if _should_use_database():
        cache = _MANUTENCAO_CACHE
        lock = cache["lock"]
        with lock:
            version = _db_version("manutencao")
            cached = cache.get("df")
            if cached is not None and cache.get("mtime") == version:
                return cached.copy()
            df = _read_database_table(
                "manutencao",
                ["Data", "Mes", "Custo", "PLACA", "OFICINA", "Categoria"],
                date_columns=["Data"],
            )
            cache["mtime"] = version
            cache["df"] = df.copy()
            return df.copy()

    cache = _MANUTENCAO_CACHE
    lock = cache["lock"]

    with lock:
        try:
            mtime = DATA_MANU.stat().st_mtime
        except FileNotFoundError:
            cache["mtime"] = None
            cache["df"] = None
            return _empty()
        except PermissionError:
            print("Aviso: sem permissao para ler planilha de manutencao. Verifique se o arquivo esta aberto.")
            cached = cache.get("df")
            if cached is not None:
                return cached.copy()
            return _empty()

        cached = cache.get("df")
        if cached is not None and cache.get("mtime") == mtime:
            return cached.copy()

        try:
            sheets = pd.read_excel(DATA_MANU, sheet_name=None, header=1)
            sheet_years = _years_from_sheet_names(sheets.keys())
        except PermissionError:
            print("Aviso: sem permissao para ler planilha de manutencao. Verifique se o arquivo esta aberto.")
            if cached is not None:
                return cached.copy()
            return _empty()
        except Exception as exc:
            print(f"Aviso: falha ao ler planilha de manutencao: {exc}")
            if cached is not None:
                return cached.copy()
            return _empty()

        frames = []
        for sheet_name, sheet_df in sheets.items():
            if sheet_df is None or sheet_df.empty:
                continue
            cleaned = _clean_columns(sheet_df)
            if cleaned.empty:
                continue
            sheet_year = _sheet_year(sheet_name)
            if sheet_year is not None:
                cleaned["_SheetYear"] = sheet_year
            frames.append(cleaned)
        if not frames:
            return _empty()
        df = pd.concat(frames, ignore_index=True)

        df = df.rename(columns={
            "PLACAS": "PLACA",
            "MES": "Mes",
            "DATA": "Data",
        })

        df["Data"] = pd.to_datetime(df.get("Data"), errors="coerce")
        if "Mes" in df.columns:
            mes_raw = df["Mes"]
            mes_dt = pd.to_datetime(mes_raw, errors="coerce")
            if mes_dt.isna().any():
                mes_norm = (
                    mes_raw.astype("string")
                    .str.strip()
                    .str.lower()
                    .str.replace(".", "/", regex=False)
                    .str.replace("-", "/", regex=False)
                    .str.replace(" ", "", regex=False)
                )
                month_map = {
                    "jan": "01",
                    "fev": "02",
                    "mar": "03",
                    "abr": "04",
                    "mai": "05",
                    "jun": "06",
                    "jul": "07",
                    "ago": "08",
                    "set": "09",
                    "out": "10",
                    "nov": "11",
                    "dez": "12",
                }
                for abbr, num in month_map.items():
                    mes_norm = mes_norm.str.replace(abbr, num, regex=False)
                mes_dt_alt = pd.to_datetime(mes_norm, errors="coerce", format="%m/%y")
                mes_dt_alt = mes_dt_alt.combine_first(
                    pd.to_datetime(mes_norm, errors="coerce", format="%m/%Y")
                )
                mes_dt = mes_dt.combine_first(mes_dt_alt)
            df["Data"] = df["Data"].combine_first(mes_dt)

        df = _apply_sheet_year(df, date_col="Data", mes_col=None)
        df["Mes"] = df["Data"].dt.to_period("M").astype(str)
        df["Custo"] = _to_numeric_currency(df.get("Custo"))

        df = df.dropna(subset=["Mes", "Custo"]).copy()
        if "PLACA" in df.columns:
            df["PLACA"] = _normalize_plate_series(df["PLACA"])
        if "OFICINA" in df.columns:
            df["OFICINA"] = df["OFICINA"].astype("string").str.strip()

        vex_col = next((col for col in df.columns if col.lower() == "vex"), None)
        if vex_col:
            df[vex_col] = df[vex_col].astype("string").str.strip()
            df["Categoria"] = df[vex_col].apply(lambda value: "Vex" if pd.notna(value) and value != "" else "Transporte")
        else:
            df["Categoria"] = "Transporte"

        df.attrs["anos_sheets"] = sheet_years
        cache["mtime"] = mtime
        cache["df"] = df.copy()
        return df


def agg_manutencao(df: pd.DataFrame) -> dict:
    custo_total = float(df["Custo"].sum()) if "Custo" in df else 0.0
    total_servicos = int(len(df))
    media_servico = float(custo_total / total_servicos) if total_servicos else 0.0
    meses_distintos = df["Mes"].dropna().unique() if "Mes" in df else []
    media_mensal = float(custo_total / len(meses_distintos)) if len(meses_distintos) else 0.0

    return {
        "custo_total": custo_total,
        "total_servicos": total_servicos,
        "media_servico": media_servico,
        "media_mensal": media_mensal,
        "custo_mensal": _group_sum(df, "Mes", sort_by="group"),
        "gasto_por_placa": _group_sum(df, "PLACA"),
        "gasto_por_oficina": _group_sum(df, "OFICINA"),
        "placas": _unique_sorted(df, "PLACA"),
        "oficinas": _unique_sorted(df, "OFICINA"),
        "meses": _unique_sorted(df, "Mes"),
        "segmentos": _unique_sorted(df, "Categoria"),
        "custo_semana": _weekly_series(df, "Data", "Custo", "Custo"),
    }


def load_hoteis() -> pd.DataFrame:
    def _empty() -> pd.DataFrame:
        return pd.DataFrame(columns=[
            "Data",
            "Valor",
            "Dias",
            "Mes",
            "Motorista",
            "Ajudante",
            "Cidade",
            "Hotel",
            "Tipo",
        ])

    if _should_use_database():
        cache = _HOTEIS_CACHE
        lock = cache["lock"]
        with lock:
            version = _db_version("hoteis")
            cached = cache.get("df")
            if cached is not None and cache.get("mtime") == version:
                return cached.copy()
            df = _read_database_table(
                "hoteis",
                ["Data", "Valor", "Dias", "Mes", "Motorista", "Ajudante", "Cidade", "Hotel", "Tipo", "Categoria"],
                date_columns=["Data"],
            )
            cache["mtime"] = version
            cache["df"] = df.copy()
            return df.copy()

    cache = _HOTEIS_CACHE
    lock = cache["lock"]

    def _strip_autofilter(path: Path) -> io.BytesIO | None:
        try:
            with zipfile.ZipFile(path, "r") as zin:
                mem = io.BytesIO()
                with zipfile.ZipFile(mem, "w") as zout:
                    for name in zin.namelist():
                        data = zin.read(name)
                        if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                            try:
                                text = data.decode("utf-8")
                                text = re.sub(r"<autoFilter[^>]*?>.*?</autoFilter>", "", text, flags=re.DOTALL)
                                text = re.sub(r"<autoFilter[^>]*/>", "", text, flags=re.DOTALL)
                                data = text.encode("utf-8")
                            except Exception:
                                pass
                        zout.writestr(name, data)
                mem.seek(0)
                return mem
        except Exception as exc:  # pragma: no cover
            print(f"Aviso: nao foi possivel limpar autoFilter da planilha de hoteis: {exc}")
            return None

    def _concat_hoteis_sheets(sheets: dict) -> pd.DataFrame | None:
        frames = []
        for sheet_name, sheet_df in sheets.items():
            if sheet_df is None or sheet_df.empty:
                continue
            normalized = {
                _normalize_ascii(col).upper()
                for col in sheet_df.columns
                if col is not None
            }
            if not {"DATA", "VALOR"}.issubset(normalized):
                continue
            sheet_year = _sheet_year(sheet_name)
            if sheet_year is not None:
                sheet_df = sheet_df.copy()
                sheet_df["_SheetYear"] = sheet_year
            frames.append(sheet_df)
        if not frames:
            return None
        return pd.concat(frames, ignore_index=True)

    def _read_hoteis_fallback(path: Path) -> pd.DataFrame | None:
        try:
            from openpyxl import load_workbook
        except Exception as exc:  # pragma: no cover - dependencia externa
            print(f"Aviso: fallback de hoteis indisponivel ({exc})")
            return None

        try:
            wb = load_workbook(path, data_only=True, read_only=True)
        except Exception as exc:  # pragma: no cover - leitura quebrada
            print(f"Aviso: falha ao ler planilha de hoteis (fallback): {exc}")
            return None

        sheet_names = [
            name for name in wb.sheetnames
            if "HOTEIS" in _normalize_ascii(name).upper()
        ]
        if not sheet_names:
            sheet_names = [wb.sheetnames[0]]

        frames = []
        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            header_idx = None
            header_row = None
            try:
                all_rows = list(ws.iter_rows(values_only=True))
            except Exception as exc:  # pragma: no cover
                print(f"Aviso: falha ao percorrer planilha de hoteis (fallback): {exc}")
                continue

            for idx, row in enumerate(all_rows):
                normalized = {_normalize_ascii(value).upper() for value in row if value is not None}
                if {"DATA", "VALOR"}.issubset(normalized):
                    header_idx = idx
                    header_row = list(row)
                    break
                if idx >= 20:
                    break

            if header_row is None or header_idx is None:
                continue

            max_len = len(header_row)
            data_rows = []
            for row in all_rows[header_idx + 1 :]:
                if all(cell is None for cell in row):
                    continue
                data_rows.append(list(row[:max_len]))

            if data_rows:
                frames.append(pd.DataFrame(data_rows, columns=header_row))

        if not frames:
            print("Aviso: cabecalho de hoteis nao encontrado no fallback.")
            return None

        return pd.concat(frames, ignore_index=True)

    with lock:
        try:
            mtime = DATA_HOTEIS.stat().st_mtime
        except FileNotFoundError:
            cache["mtime"] = None
            cache["df"] = None
            return _empty()
        except PermissionError:
            print("Aviso: sem permissao para ler planilha de hoteis. Verifique se o arquivo esta aberto.")
            cached = cache.get("df")
            if cached is not None:
                return cached.copy()
            return _empty()

        cached = cache.get("df")
        if cached is not None and cache.get("mtime") == mtime:
            return cached.copy()

        sheet_years: list[int] = []
        try:
            sheets = pd.read_excel(DATA_HOTEIS, sheet_name=None, header=4)
            sheet_years = _years_from_sheet_names(sheets.keys())
            df = _concat_hoteis_sheets(sheets)
        except PermissionError:
            print("Aviso: sem permissao para ler planilha de hoteis. Verifique se o arquivo esta aberto.")
            if cached is not None:
                return cached.copy()
            return _empty()
        except Exception as exc:
            print(f"Aviso: falha ao ler planilha de hoteis: {exc}")
            df = None
            buffer_clean = _strip_autofilter(DATA_HOTEIS)
            if buffer_clean is not None:
                try:
                    sheets = pd.read_excel(buffer_clean, sheet_name=None, header=4)
                    sheet_years = _years_from_sheet_names(sheets.keys())
                    df = _concat_hoteis_sheets(sheets)
                except Exception as exc_clean:  # pragma: no cover - leitura fallback
                    print(f"Aviso: falha ao ler planilha de hoteis (limpa): {exc_clean}")
            if df is None:
                df = _read_hoteis_fallback(DATA_HOTEIS)
            if df is None:
                if cached is not None:
                    return cached.copy()
                return _empty()

        if df is None:
            df = _read_hoteis_fallback(DATA_HOTEIS)
            if df is None:
                if cached is not None:
                    return cached.copy()
                return _empty()
        if not sheet_years:
            try:
                sheet_years = _years_from_sheet_names(pd.ExcelFile(DATA_HOTEIS).sheet_names)
            except Exception:
                sheet_years = []

        df = _clean_columns(df)

        df = df.rename(columns={
            "DATA": "Data",
            "VALOR": "Valor",
            "HOTEL/POUSADA": "Hotel",
            "MOTORISTA": "Motorista",
            "AJUDANTE": "Ajudante",
            "CIDADE": "Cidade",
            "TIPO": "Tipo",
            "DIAS": "Dias",
            "CATEGORIA": "Categoria",
        })

        df["Data"] = pd.to_datetime(df.get("Data"), errors="coerce")
        df = _apply_sheet_year(df, date_col="Data", mes_col=None)
        df["Valor"] = pd.to_numeric(df.get("Valor"), errors="coerce")
        df["Dias"] = pd.to_numeric(df.get("Dias"), errors="coerce")

        period = df["Data"].dt.to_period("M")
        df["Mes"] = period.astype(str)
        df.loc[period.isna(), "Mes"] = None

        for col in ["Motorista", "Ajudante", "Cidade", "Hotel", "Tipo", "Categoria"]:
            if col in df.columns:
                df[col] = df[col].astype("string").str.strip()
        # Hotéis não usam categoria; sempre considerar como Transporte.
        df["Categoria"] = "Transporte"

        df.attrs["anos_sheets"] = sheet_years
        cache["mtime"] = mtime
        cache["df"] = df.copy()
        return df.copy()


def agg_hoteis(df: pd.DataFrame) -> dict:
    reservas = df[df["Data"].notna()].copy() if "Data" in df.columns else df.copy()
    if "Data" in reservas.columns:
        reservas["Data"] = pd.to_datetime(reservas["Data"], errors="coerce")
    valor_total = float(reservas["Valor"].fillna(0).sum()) if "Valor" in reservas else 0.0
    reservas_total = int(reservas.shape[0])
    meses_distintos = reservas["Mes"].dropna().unique() if "Mes" in reservas else []
    media_mensal = float(valor_total / len(meses_distintos)) if len(meses_distintos) else 0.0
    valor_medio_reserva = float(valor_total / reservas_total) if reservas_total else 0.0
    col_nao_planejada = next(
        (
            col
            for col in reservas.columns
            if str(col).strip().upper().replace(" ", "") in {"NAOPLANEJADA", "NAOPLANEJADAS"}
        ),
        None,
    )
    if col_nao_planejada:
        flag_series = pd.to_numeric(reservas[col_nao_planejada], errors="coerce").fillna(0)
        reservas = reservas.assign(_NaoPlanejada=flag_series.astype("int64"))
    else:
        reservas = reservas.assign(_NaoPlanejada=0)
    reservas_nao_planejadas = 0
    if "Data" in reservas.columns and "Valor" in reservas.columns:
        mask_sabado = reservas["Data"].dt.dayofweek.isin([4, 5]).fillna(False)
        mask_nao_planejada = reservas["_NaoPlanejada"] == 1
        mask_nao_planejada_total = mask_nao_planejada | mask_sabado

        valor_sabado = float(reservas.loc[mask_sabado, "Valor"].fillna(0).sum())
        valor_nao_planejado = float(reservas.loc[mask_nao_planejada_total, "Valor"].fillna(0).sum())
        reservas_nao_planejadas = int(mask_nao_planejada_total.sum())
    else:
        valor_sabado = 0.0
        valor_nao_planejado = 0.0

    semanal = _weekly_series(reservas, "Data", "Valor", "Valor")
    if "DiaISO" in semanal:
        nao_planejada_flags = []
        if "Data" in reservas.columns:
            nao_planejada_por_dia = (
                reservas.loc[reservas["_NaoPlanejada"] == 1, "Data"]
                .dt.normalize()
                .value_counts()
            )
        else:
            nao_planejada_por_dia = pd.Series(dtype="int64")
        for iso in semanal["DiaISO"]:
            dt = pd.to_datetime(iso, errors="coerce")
            if pd.isna(dt):
                nao_planejada_flags.append(False)
            else:
                nao_planejada_flags.append(bool(nao_planejada_por_dia.get(dt.normalize(), 0)))
        semanal["NaoPlanejada"] = nao_planejada_flags
    else:
        semanal["NaoPlanejada"] = []

    return {
        "valor_total": valor_total,
        "reservas_total": reservas_total,
        "media_mensal": media_mensal,
        "valor_medio_reserva": valor_medio_reserva,
        "valor_mensal": _group_sum(reservas, "Mes", "Valor", sort_by="group"),
        "valor_por_cidade": _group_sum(reservas, "Cidade", "Valor"),
        "valor_por_hotel": _group_sum(reservas, "Hotel", "Valor"),
        "meses": _unique_sorted(reservas, "Mes"),
        "cidades": _unique_sorted(reservas, "Cidade"),
        "hoteis": _unique_sorted(reservas, "Hotel"),
        "valor_semana": semanal,
        "valor_sabado": valor_sabado,
        "valor_nao_planejado": valor_nao_planejado,
        "reservas_nao_planejadas": reservas_nao_planejadas,
    }


def load_pedagio() -> pd.DataFrame:
    def _empty() -> pd.DataFrame:
        return pd.DataFrame(columns=['PLACA', 'Tipo', 'Custo', 'Mes', 'Data', 'Categoria'])

    if _should_use_database():
        cache = _PEDAGIO_CACHE
        lock = cache['lock']
        with lock:
            version = _db_version("pedagio")
            cached_df = cache.get('df')
            if cached_df is not None and cache.get('mtime') == version:
                return cached_df.copy(deep=False)
            df = _read_database_table(
                "pedagio",
                ["PLACA", "Tipo", "Custo", "Mes", "Data", "Categoria"],
                date_columns=["Data"],
            )
            cache['mtime'] = version
            cache['df'] = df.copy()
            return df.copy(deep=False)

    cache = _PEDAGIO_CACHE
    lock = cache['lock']

    with lock:
        if not DATA_PEDAGIO.exists():
            cache['mtime'] = None
            cache['df'] = None
            return _empty()

        try:
            mtime = DATA_PEDAGIO.stat().st_mtime
        except PermissionError:
            print('Aviso: sem permissao para ler planilha de pedagio/seguro/IPVA. Verifique se o arquivo esta aberto.')
            cached_df = cache.get('df')
            if cached_df is not None:
                return cached_df.copy(deep=False)
            return _empty()

        cached_df = cache.get('df')
        if cached_df is not None and cache.get('mtime') == mtime:
            return cached_df.copy(deep=False)

        try:
            raw_sheets = pd.read_excel(DATA_PEDAGIO, sheet_name=None, header=None, engine='openpyxl')
            sheet_years = _years_from_sheet_names(raw_sheets.keys())
        except PermissionError:
            print('Aviso: sem permissao para ler planilha de pedagio/seguro/IPVA. Verifique se o arquivo esta aberto.')
            cached_df = cache.get('df')
            if cached_df is not None:
                return cached_df.copy(deep=False)
            return _empty()
        except Exception as exc:
            print(f'Aviso: falha ao ler planilha de pedagio/seguro/IPVA: {exc}')
            cached_df = cache.get('df')
            if cached_df is not None:
                return cached_df.copy(deep=False)
            return _empty()

        expected_core = {'TIPO', 'CUSTO'}
        frames = []
        for sheet_name, raw in raw_sheets.items():
            header_idx = None
            for idx in range(min(len(raw), 10)):
                row = raw.iloc[idx]
                normalized = set()
                for value in row.tolist():
                    if pd.isna(value):
                        continue
                    normalized.add(_normalize_ascii(value).upper())
                has_placa = any(label in normalized for label in ('PLACA', 'PLACAS'))
                if has_placa and expected_core.issubset(normalized):
                    header_idx = idx
                    break

            if header_idx is None:
                continue

            df_sheet = raw.iloc[header_idx + 1 :].copy()
            df_sheet.columns = raw.iloc[header_idx]
            df_sheet = df_sheet.dropna(how='all').reset_index(drop=True)
            df_sheet = _clean_columns(df_sheet)
            if df_sheet.empty:
                continue
            sheet_year = _sheet_year(sheet_name)
            if sheet_year is not None:
                df_sheet["_SheetYear"] = sheet_year
            frames.append(df_sheet)

        if not frames:
            print('Aviso: cabecalho da planilha de pedagio/seguro/IPVA nao encontrado.')
            return _empty()

        df = pd.concat(frames, ignore_index=True)

        rename_map = {}
        for col in df.columns:
            col_norm = _normalize_ascii(col).upper()
            if col_norm in {'PLACA', 'PLACAS'}:
                rename_map[col] = 'PLACA'
            elif col_norm == 'TIPO':
                rename_map[col] = 'Tipo'
            elif col_norm in {'CUSTO', 'VALOR', 'VALORES'}:
                rename_map[col] = 'Custo'
            elif col_norm == 'MES':
                rename_map[col] = 'Mes'
            elif col_norm == 'DATA':
                rename_map[col] = 'Data'
            elif col_norm == 'VEX':
                rename_map[col] = 'Vex'
            elif col_norm in {'DESCRICAO', 'DESCRICAO'}:
                rename_map[col] = 'Descricao'
            elif col_norm in {'SEGURADORA', 'FORNECEDOR'}:
                rename_map[col] = 'Fornecedor'

        if rename_map:
            df = df.rename(columns=rename_map)

        if 'Custo' in df.columns:
            df['Custo'] = _to_numeric_currency(df['Custo'])
        else:
            df['Custo'] = pd.Series(pd.NA, index=df.index, dtype='float64')

        if 'Tipo' in df.columns:
            df['Tipo'] = df['Tipo'].apply(_canonical_tipo).astype('string')
        else:
            df['Tipo'] = pd.Series(pd.NA, index=df.index, dtype='string')

        if 'PLACA' in df.columns:
            df['PLACA'] = _normalize_plate_series(df['PLACA'])
        else:
            df['PLACA'] = pd.Series(pd.NA, index=df.index, dtype='string')

        if 'Mes' in df.columns:
            mes_raw = df['Mes']
            mes_dt = pd.to_datetime(mes_raw, errors='coerce')
            df['Mes'] = mes_dt.dt.to_period('M').astype('string')
            missing_mes = df['Mes'].isna() | (df['Mes'] == '')
            if missing_mes.any():
                alt = (
                    mes_raw.astype('string')
                    .str.strip()
                    .str.replace(' ', '', regex=False)
                    .str.replace('.', '/', regex=False)
                    .str.replace('-', '/', regex=False)
                )
                alt_dt = pd.to_datetime(alt, errors='coerce')
                valid_alt = missing_mes & alt_dt.notna()
                df.loc[valid_alt, 'Mes'] = alt_dt[valid_alt].dt.to_period('M').astype('string')
                df.loc[missing_mes & ~valid_alt, 'Mes'] = alt.loc[missing_mes & ~valid_alt]
        else:
            df['Mes'] = pd.Series(pd.NA, index=df.index, dtype='string')

        if 'Data' in df.columns:
            df['Data'] = pd.to_datetime(df['Data'], errors='coerce')
            empty_mes = df['Mes'].isna() | (df['Mes'] == '')
            df.loc[empty_mes, 'Mes'] = df.loc[empty_mes, 'Data'].dt.to_period('M').astype('string')
        df = _apply_sheet_year(df, date_col="Data", mes_col="Mes")

        vex_col = next((col for col in df.columns if col.lower() == 'vex'), None)
        if vex_col:
            df[vex_col] = df[vex_col].astype('string').str.strip()
            df['Categoria'] = df[vex_col].apply(lambda value: 'Vex' if pd.notna(value) and value != '' else 'Transporte')
        elif 'Categoria' not in df.columns:
            df['Categoria'] = 'Transporte'

        df = df[df['Custo'].notna()].copy()
        df['Tipo'] = df['Tipo'].fillna('Outros')

        df.attrs["anos_sheets"] = sheet_years
        result = df.copy()
        cache['mtime'] = mtime
        cache['df'] = result
        return result.copy(deep=False)



def agg_pedagio(df: pd.DataFrame) -> dict:
    registros = df.shape[0]
    custo_total = float(df["Custo"].sum()) if "Custo" in df else 0.0
    meses_distintos = df["Mes"].dropna().unique() if "Mes" in df else []
    media_mensal = float(custo_total / len(meses_distintos)) if len(meses_distintos) else 0.0
    media_valores = float(custo_total / registros) if registros else 0.0

    if "Tipo" in df.columns and not df.empty:
        tipo_totais = df.groupby("Tipo", dropna=False)["Custo"].sum()
    else:
        tipo_totais = pd.Series(dtype="float64")

    gasto_pedagio = float(tipo_totais.get("Pedagio", 0.0))
    gasto_ipva = float(tipo_totais.get("IPVA", 0.0))
    gasto_seguro = float(tipo_totais.get("Seguro", 0.0))

    resultado = {
        "custo_total": custo_total,
        "total_lancamentos": registros,
        "media_mensal": media_mensal,
        "ticket_medio": media_valores,
        "media_valores": media_valores,
        "gasto_pedagio": gasto_pedagio,
        "gasto_ipva": gasto_ipva,
        "gasto_seguro": gasto_seguro,
        "custo_mensal": _group_sum(df, "Mes", "Custo", sort_by="group"),
        "gasto_por_tipo": _group_sum(df, "Tipo", "Custo"),
        "gasto_por_placa": _group_sum(df, "PLACA", "Custo"),
        "meses": _unique_sorted(df, "Mes"),
        "tipos": _unique_sorted(df, "Tipo"),
        "placas": _unique_sorted(df, "PLACA"),
    }

    if "Categoria" in df.columns:
        resultado["segmentos"] = _unique_sorted(df, "Categoria")
        resultado["gasto_por_categoria"] = _group_sum(df, "Categoria", "Custo")
    else:
        resultado["segmentos"] = []
        resultado["gasto_por_categoria"] = {"Categoria": [], "Custo": []}

    resultado["custo_semana"] = _weekly_series(df, "Data", "Custo", "Custo")
    return resultado





@app.route("/")
def home():
    return render_template("home.html")


@app.route("/combustivel")
def comb_page():
    return render_template("combustivel.html")


@app.route("/manutencao")
def manut_page():
    return render_template("manutencao.html")


@app.route("/hoteis")
def hoteis_page():
    return render_template("hoteis.html")


@app.route("/pedagio")
def pedagio_page():
    return render_template("pedagio.html")


@app.route("/vex")
def vex_page():
    return render_template("vex.html")


@app.route("/data/combustivel")
def data_comb():
    df = load_combustivel()
    df = _exclude_vex(df)
    km_rodados = _COMBUSTIVEL_CACHE.get("km_rodados_mensal")

    ano = _parse_int(request.args.get("ano"))
    meses = _parse_mes_list(request.args.getlist("mes"))
    placa = request.args.get("placa")
    posto = request.args.get("posto")
    combustivel = request.args.get("combustivel")
    segmento = request.args.get("segmento")

    if placa and placa != "Todos":
        df = df[df["PLACA"] == placa]
    if posto and posto != "Todos":
        df = df[df["POSTOS"] == posto]
    if combustivel and combustivel != "Todos":
        df = df[df["Combustivel"] == combustivel]
    if segmento and segmento != "Todos" and "Categoria" in df.columns:
        df = df[df["Categoria"] == segmento]

    anos_disponiveis = _unique_years(df)
    sheet_years = df.attrs.get("anos_sheets", [])
    if sheet_years:
        anos_disponiveis = sorted({*anos_disponiveis, *sheet_years})
    df_meses = _filter_by_period(df, ano=ano) if ano is not None else df
    meses_disponiveis = _unique_sorted(df_meses, "Mes")

    if ano is not None:
        df = _filter_by_period(df, ano=ano)
    if meses:
        df = df[df["Mes"].isin(meses)]

    km_override = None
    if isinstance(km_rodados, pd.DataFrame) and not km_rodados.empty:
        km_override = km_rodados.copy()
        if placa and placa != "Todos":
            km_override = km_override[km_override["PLACA"] == placa]
        if ano is not None:
            km_override = km_override[pd.to_datetime(km_override["Mes"], errors="coerce").dt.year == ano]
        if meses:
            km_override = km_override[km_override["Mes"].isin(meses)]
        if not df.empty and "Mes" in df.columns and "PLACA" in df.columns:
            allowed = df[["Mes", "PLACA"]].dropna().drop_duplicates()
            if not allowed.empty:
                km_override = km_override.merge(allowed, on=["Mes", "PLACA"], how="inner")
        if km_override.empty:
            km_override = None

    resultado = agg_combustivel(df, km_override=km_override if ano == 2026 else None)
    resultado["anos"] = anos_disponiveis
    resultado["meses"] = meses_disponiveis
    return jsonify(resultado)


@app.route("/data/manutencao")
def data_manu():
    df = load_manutencao()
    df = _exclude_vex(df)

    ano = _parse_int(request.args.get("ano"))
    meses = _parse_mes_list(request.args.getlist("mes"))
    placa = request.args.get("placa")
    oficina = request.args.get("oficina")
    segmento = request.args.get("segmento")

    if placa and placa != "Todos":
        df = df[df["PLACA"] == placa]
    if oficina and oficina != "Todos":
        df = df[df["OFICINA"] == oficina]
    if segmento and segmento != "Todos" and "Categoria" in df.columns:
        df = df[df["Categoria"] == segmento]

    anos_disponiveis = _unique_years(df)
    sheet_years = df.attrs.get("anos_sheets", [])
    if sheet_years:
        anos_disponiveis = sorted({*anos_disponiveis, *sheet_years})
    df_meses = _filter_by_period(df, ano=ano) if ano is not None else df
    meses_disponiveis = _unique_sorted(df_meses, "Mes")

    if ano is not None:
        df = _filter_by_period(df, ano=ano)
    if meses:
        df = df[df["Mes"].isin(meses)]

    resultado = agg_manutencao(df)
    resultado["anos"] = anos_disponiveis
    resultado["meses"] = meses_disponiveis
    return jsonify(resultado)


@app.route("/data/hoteis")
def data_hoteis():
    df_total = load_hoteis()
    df_total = _exclude_vex(df_total)
    totais_gerais = agg_hoteis(df_total)
    df = df_total.copy()

    ano = _parse_int(request.args.get("ano"))
    meses = _parse_mes_list(request.args.getlist("mes"))
    cidade = request.args.get("cidade")
    hotel = request.args.get("hotel")

    if cidade and cidade != "Todos":
        df = df[df["Cidade"] == cidade]
    if hotel and hotel != "Todos":
        df = df[df["Hotel"] == hotel]

    anos_disponiveis = _unique_years(df)
    sheet_years = df.attrs.get("anos_sheets", [])
    if sheet_years:
        anos_disponiveis = sorted({*anos_disponiveis, *sheet_years})
    df_meses = _filter_by_period(df, ano=ano) if ano is not None else df
    meses_disponiveis = _unique_sorted(df_meses, "Mes")

    if ano is not None:
        df = _filter_by_period(df, ano=ano)
    if meses:
        df = df[df["Mes"].isin(meses)]

    resultado = agg_hoteis(df)
    resultado["valor_sabado_total"] = totais_gerais.get("valor_sabado", 0.0)
    resultado["valor_nao_planejado_total"] = totais_gerais.get("valor_nao_planejado", 0.0)
    resultado["anos"] = anos_disponiveis
    resultado["meses"] = meses_disponiveis
    return jsonify(resultado)


@app.route("/data/pedagio")
def data_pedagio():
    df = load_pedagio()
    df = _exclude_vex(df)

    ano = _parse_int(request.args.get("ano"))
    meses = _parse_mes_list(request.args.getlist("mes"))
    placa = request.args.get("placa")
    tipo = request.args.get("tipo")
    segmento = request.args.get("segmento")

    if placa and placa != "Todos":
        df = df[df["PLACA"] == placa]
    if tipo and tipo != "Todos":
        df = df[df["Tipo"] == tipo]
    if segmento and segmento != "Todos" and "Categoria" in df.columns:
        df = df[df["Categoria"] == segmento]

    anos_disponiveis = _unique_years(df)
    sheet_years = df.attrs.get("anos_sheets", [])
    if sheet_years:
        anos_disponiveis = sorted({*anos_disponiveis, *sheet_years})
    df_meses = _filter_by_period(df, ano=ano) if ano is not None else df
    meses_disponiveis = _unique_sorted(df_meses, "Mes")

    if ano is not None:
        df = _filter_by_period(df, ano=ano)
    if meses:
        df = df[df["Mes"].isin(meses)]

    resultado = agg_pedagio(df)
    resultado["anos"] = anos_disponiveis
    resultado["meses"] = meses_disponiveis
    return jsonify(resultado)


@app.route("/data/vex")
def data_vex():
    ano = _parse_int(request.args.get("ano"))
    meses = _parse_mes_list(request.args.getlist("mes"))
    placa = request.args.get("placa")

    df_comb = _only_vex(load_combustivel())
    df_manu = _only_vex(load_manutencao())
    df_hoteis = load_hoteis().iloc[0:0].copy()
    df_ped = _only_vex(load_pedagio())
    km_rodados = _COMBUSTIVEL_CACHE.get("km_rodados_mensal")

    anos_disponiveis: set[int] = set()
    for df_src in (df_comb, df_manu, df_hoteis, df_ped):
        anos_disponiveis.update(_unique_years(df_src))
        anos_disponiveis.update(df_src.attrs.get("anos_sheets", []))
    anos_list = sorted(anos_disponiveis)

    def _meses_disponiveis(*dfs: pd.DataFrame) -> list[str]:
        frames = [df for df in dfs if not df.empty and "Mes" in df.columns]
        if not frames:
            return []
        merged = pd.concat(frames, ignore_index=True)
        return _unique_sorted(merged, "Mes")

    df_meses_base = [df for df in (df_comb, df_manu, df_hoteis, df_ped)]
    if ano is not None:
        df_meses_base = [_filter_by_period(df, ano=ano) for df in df_meses_base]
    meses_disponiveis = _meses_disponiveis(*df_meses_base)

    df_placas_base = [df for df in (df_comb, df_manu, df_ped)]
    if ano is not None:
        df_placas_base = [_filter_by_period(df, ano=ano) for df in df_placas_base]
    if meses:
        df_placas_base = [df[df["Mes"].isin(meses)] for df in df_placas_base]
    placas_disponiveis: list[str] = []
    if df_placas_base:
        frames = [df[["PLACA"]] for df in df_placas_base if "PLACA" in df.columns]
        if frames:
            placas_disponiveis = _unique_sorted(pd.concat(frames, ignore_index=True), "PLACA")

    def _apply_filters(df: pd.DataFrame) -> pd.DataFrame:
        if ano is not None:
            df = _filter_by_period(df, ano=ano)
        if meses:
            df = df[df["Mes"].isin(meses)]
        if placa and placa != "Todos" and "PLACA" in df.columns:
            df = df[df["PLACA"] == placa]
        return df

    df_comb = _apply_filters(df_comb)
    df_manu = _apply_filters(df_manu)
    df_hoteis = _apply_filters(df_hoteis)
    df_ped = _apply_filters(df_ped)

    km_override = None
    if ano == 2026 and isinstance(km_rodados, pd.DataFrame) and not km_rodados.empty:
        km_override = km_rodados.copy()
        if ano is not None:
            km_override = km_override[pd.to_datetime(km_override["Mes"], errors="coerce").dt.year == ano]
        if meses:
            km_override = km_override[km_override["Mes"].isin(meses)]
        if placa and placa != "Todos":
            km_override = km_override[km_override["PLACA"] == placa]
        if not df_comb.empty and "Mes" in df_comb.columns and "PLACA" in df_comb.columns:
            allowed = df_comb[["Mes", "PLACA"]].dropna().drop_duplicates()
            if not allowed.empty:
                km_override = km_override.merge(allowed, on=["Mes", "PLACA"], how="inner")

    total_comb = float(df_comb["Custo"].sum()) if "Custo" in df_comb else 0.0
    if km_override is not None and "Km Rodados" in km_override.columns:
        km_total = float(km_override["Km Rodados"].sum())
    else:
        km_total = float(df_comb["Km Rodados"].sum()) if "Km Rodados" in df_comb else 0.0
    litros_total = float(df_comb["Litros"].sum()) if "Litros" in df_comb else 0.0
    km_por_litro = (km_total / litros_total) if litros_total else 0.0
    custo_por_km = (total_comb / km_total) if km_total else 0.0
    custo_por_litro = (total_comb / litros_total) if litros_total else 0.0
    total_manu = float(df_manu["Custo"].sum()) if "Custo" in df_manu else 0.0
    total_hoteis = 0.0
    total_ped = float(df_ped["Custo"].sum()) if "Custo" in df_ped else 0.0
    total_vex = total_comb + total_manu + total_hoteis + total_ped

    mensal_comb = _group_sum(df_comb, "Mes", "Custo", sort_by="group")
    mensal_manu = _group_sum(df_manu, "Mes", "Custo", sort_by="group")
    mensal_hoteis = {"Mes": [], "Valor": []}
    mensal_ped = _group_sum(df_ped, "Mes", "Custo", sort_by="group")

    monthly_map: dict[str, float] = {}
    for src, key in (
        (mensal_comb, "Custo"),
        (mensal_manu, "Custo"),
        (mensal_ped, "Custo"),
    ):
        meses_src = src.get("Mes", [])
        valores_src = src.get(key, [])
        for mes, valor in zip(meses_src, valores_src):
            monthly_map[mes] = monthly_map.get(mes, 0.0) + float(valor or 0)

    meses_sorted = sorted(monthly_map.keys())
    mensal_total = {
        "Mes": meses_sorted,
        "Valor": [round(monthly_map[mes], 2) for mes in meses_sorted],
    }

    por_area = {
        "Area": ["Combustivel", "Manutencao", "Pedagio"],
        "Valor": [round(total_comb, 2), round(total_manu, 2), round(total_ped, 2)],
    }

    placa_totais: dict[str, float] = {}
    for df_src, col_valor in ((df_comb, "Custo"), (df_manu, "Custo"), (df_ped, "Custo")):
        if df_src.empty or "PLACA" not in df_src.columns or col_valor not in df_src.columns:
            continue
        df_val = df_src.dropna(subset=["PLACA"]).copy()
        df_val[col_valor] = pd.to_numeric(df_val[col_valor], errors="coerce")
        for placa_val, total in df_val.groupby("PLACA")[col_valor].sum().items():
            key = str(placa_val).strip()
            placa_totais[key] = placa_totais.get(key, 0.0) + float(total or 0.0)

    placas_ordenadas = sorted(placa_totais.items(), key=lambda item: item[1], reverse=True)
    gasto_por_placa = {
        "PLACA": [item[0] for item in placas_ordenadas],
        "Valor": [round(item[1], 2) for item in placas_ordenadas],
    }

    return jsonify({
        "anos": anos_list,
        "meses": meses_disponiveis,
        "placas": placas_disponiveis,
        "total_vex": round(total_vex, 2),
        "combustivel_total": round(total_comb, 2),
        "manutencao_total": round(total_manu, 2),
        "hoteis_total": round(total_hoteis, 2),
        "pedagio_total": round(total_ped, 2),
        "km_total": round(km_total, 2),
        "litros_total": round(litros_total, 2),
        "km_por_litro": round(km_por_litro, 3),
        "custo_por_km": round(custo_por_km, 4),
        "custo_por_litro": round(custo_por_litro, 4),
        "mensal_total": mensal_total,
        "por_area": por_area,
        "gasto_por_placa": gasto_por_placa,
    })


def _warm_data_caches(*, blocking: bool = False) -> None:
    loaders = (
        (load_combustivel, "combustivel"),
        (load_manutencao, "manutencao"),
        (load_hoteis, "hoteis"),
        (load_pedagio, "pedagio/seguro/IPVA"),
    )

    def _run() -> None:
        with concurrent.futures.ThreadPoolExecutor(max_workers=len(loaders)) as pool:
            future_map = {
                pool.submit(loader): (loader, label)
                for loader, label in loaders
            }
            for future, (loader, label) in future_map.items():
                try:
                    future.result()
                except Exception as exc:  # pragma: no cover
                    print(f"Aviso: nao foi possivel pre-carregar {label} ({exc})")

    if blocking:
        _run()
    else:
        threading.Thread(target=_run, daemon=True).start()


if os.environ.get("JR_SKIP_WARM_CACHE", "").strip().lower() not in {"1", "true", "yes"}:
    _warm_data_caches(
        blocking=os.environ.get("WARM_CACHE_SYNC", "").strip().lower() in {"1", "true", "yes"}
    )


def _safe_total(
    loader,
    aggregator,
    key: str,
    value_col: str | None = None,
    *,
    ano: int | None = None,
    mes: int | None = None,
    meses: list[int] | None = None,
) -> dict:
    try:
        df = loader()
    except PermissionError:
        return {"status": "erro", "motivo": "permissao", "valor": None, "categorias": None, "anos_sheets": []}
    except FileNotFoundError:
        return {"status": "erro", "motivo": "arquivo_nao_encontrado", "valor": None, "categorias": None, "anos_sheets": []}
    except Exception as exc:  # pragma: no cover
        return {"status": "erro", "motivo": str(exc), "valor": None, "categorias": None, "anos_sheets": []}

    periodos_disponiveis: list[str] = []
    anos_sheets = df.attrs.get("anos_sheets", [])
    if "Mes" in df.columns:
        period_series = pd.to_datetime(df["Mes"], errors="coerce").dt.to_period("M")
        valores_periodo = {str(periodo) for periodo in period_series.dropna().unique()}
        periodos_disponiveis = sorted(valores_periodo)

    df = _filter_by_period(df, ano=ano, mes=mes, meses=meses or [])

    try:
        resumo = aggregator(df)
    except Exception as exc:  # pragma: no cover
        return {"status": "erro", "motivo": str(exc), "valor": None, "categorias": None, "anos_sheets": anos_sheets}

    valor = float(resumo.get(key, 0.0)) if resumo else 0.0
    categorias = None
    if value_col and value_col in df.columns and "Categoria" in df.columns:
        df_categoria = df.copy()
        df_categoria[value_col] = pd.to_numeric(df_categoria[value_col], errors="coerce")
        grupos = (
            df_categoria.groupby(
                df_categoria["Categoria"]
                .astype("string")
                .str.strip()
                .str.title()
                .replace({"": "Outros"})
                .fillna("Outros")
            )[value_col]
            .sum()
        )
        if not grupos.empty:
            categorias = {categoria: float(valor_cat) for categoria, valor_cat in grupos.items() if pd.notna(valor_cat)}

    return {
        "status": "ok",
        "motivo": None,
        "valor": valor,
        "categorias": categorias,
        "periodos": periodos_disponiveis,
        "anos_sheets": anos_sheets,
    }


def compute_overview_totals(*, ano: int | None = None, mes: int | None = None, meses_lista: list[int] | None = None) -> dict:
    ano = _parse_int(ano)
    mes = _parse_int(mes, min_value=1, max_value=12)
    meses_lista = list(meses_lista or [])
    if mes is not None and mes not in meses_lista:
        meses_lista.append(mes)
    areas = {
        "combustivel": (load_combustivel, agg_combustivel, "custo_total", "Custo"),
        "manutencao": (load_manutencao, agg_manutencao, "custo_total", "Custo"),
        "hoteis": (load_hoteis, agg_hoteis, "valor_total", "Valor"),
        "pedagio": (load_pedagio, agg_pedagio, "custo_total", "Custo"),
    }

    chave_cache = tuple(_CACHE_MAP[nome]["mtime"] for nome in ("combustivel", "manutencao", "hoteis", "pedagio"))
    use_cache = ano is None and mes is None and not meses_lista
    if (
        use_cache
        and _OVERVIEW_CACHE["mtimes"] == chave_cache
        and _OVERVIEW_CACHE["dados"] is not None
    ):
        return _OVERVIEW_CACHE["dados"]

    detalhes = {}
    total_geral = 0.0
    segmento_totais = defaultdict(float)
    periodos_unicos: set[str] = set()
    anos_extra: set[int] = set()

    use_threads = ano is None and mes is None
    if use_threads:
        pool = concurrent.futures.ThreadPoolExecutor(max_workers=len(areas))
    else:
        pool = concurrent.futures.ThreadPoolExecutor(max_workers=1)

    with pool:
        future_map = {
            pool.submit(
                _safe_total,
                loader,
                aggregator,
                chave,
                valor_col,
                ano=ano,
                mes=mes,
                meses=meses_lista,
            ): nome
            for nome, (loader, aggregator, chave, valor_col) in areas.items()
        }
        for future in concurrent.futures.as_completed(future_map):
            nome = future_map[future]
            resultado = future.result()
            detalhes[nome] = resultado
            if resultado["valor"] is not None:
                total_geral += resultado["valor"]
            categorias = resultado.get("categorias") or {}
            for categoria, valor in categorias.items():
                segmento_totais[categoria] += valor
            for periodo in resultado.get("periodos") or []:
                if periodo:
                    periodos_unicos.add(periodo)
            for ano_sheet in resultado.get("anos_sheets") or []:
                try:
                    anos_extra.add(int(ano_sheet))
                except (TypeError, ValueError):
                    continue

    segmentos_dict = {categoria: float(valor) for categoria, valor in segmento_totais.items()}
    periodos_ordenados = sorted(periodos_unicos)
    anos_disponiveis = sorted({int(p.split("-")[0]) for p in periodos_ordenados if "-" in p})
    if anos_extra:
        anos_disponiveis = sorted(set(anos_disponiveis) | anos_extra)
    periodos_base = periodos_ordenados
    if ano is not None:
        prefix = f"{ano}-"
        periodos_base = [periodo for periodo in periodos_ordenados if periodo.startswith(prefix)]
    meses_disponiveis = sorted({int(p.split("-")[1]) for p in periodos_base if "-" in p})

    detalhes["total_geral"] = float(total_geral)
    detalhes["segmentos"] = segmentos_dict
    detalhes["total_transporte"] = segmentos_dict.get("Transporte", 0.0)
    detalhes["total_vex"] = segmentos_dict.get("Vex", 0.0)
    detalhes["periodos_disponiveis"] = periodos_ordenados
    detalhes["anos_disponiveis"] = anos_disponiveis
    detalhes["meses_disponiveis"] = meses_disponiveis
    detalhes["filtro"] = {"ano": ano, "mes": mes, "meses": meses_lista}

    if use_cache:
        _OVERVIEW_CACHE["mtimes"] = tuple(
            _CACHE_MAP[nome]["mtime"] for nome in ("combustivel", "manutencao", "hoteis", "pedagio")
        )
        _OVERVIEW_CACHE["dados"] = detalhes
    return detalhes


@app.route("/data/overview")
def data_overview():
    ano = _parse_int(request.args.get("ano"))
    mes = _parse_int(request.args.get("mes"), min_value=1, max_value=12)
    meses_lista = _parse_mes_int_list(request.args.getlist("mes"))
    if mes is not None and mes not in meses_lista:
        meses_lista.append(mes)
    return jsonify(compute_overview_totals(ano=ano, mes=mes, meses_lista=meses_lista))


def _running_inside_streamlit() -> bool:
    try:
        from streamlit.runtime import exists as streamlit_runtime_exists
    except Exception:
        return False
    return bool(streamlit_runtime_exists())


if __name__ == "__main__":
    if _running_inside_streamlit():
        os.environ.setdefault("JR_SKIP_WARM_CACHE", "1")
        from streamlit_app import main as streamlit_main

        streamlit_main()
    else:
        app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
